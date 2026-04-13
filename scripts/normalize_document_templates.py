from __future__ import annotations

import os
import sys
from copy import copy
from pathlib import Path

import django
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / 'backend'

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from crm.contracts import TEMPLATE_REGISTRY  # noqa: E402


EMPTY_BORDER = Border(
    left=Side(style=None),
    right=Side(style=None),
    top=Side(style=None),
    bottom=Side(style=None),
)
EMPTY_FILL = PatternFill(fill_type=None)

HEADER_CLEAR_COORDS = [
    'D23',
    'D24',
    'D25',
    'D26',
    'D27',
    'D30',
    'D33',
]
ROW31_COLUMNS = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
ROW32_COLUMNS = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
DATA_COLUMNS = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
SUMMARY_CLEAR_COLUMNS = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
DOWNLOAD_TEMPLATE_GLOB = 'CELIK_KAPI_TEKLIF_SABLONU_V12*.xltx'


def resolve_cell(ws, coordinate):
    cell = ws[coordinate]
    if cell.__class__.__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col)
    return cell


def clear_cell_value(ws, coordinate):
    cell = resolve_cell(ws, coordinate)
    cell.value = ''
    cell.number_format = 'General'


def clear_visual_cell(ws, coordinate):
    cell = resolve_cell(ws, coordinate)
    cell.value = ''
    cell.border = EMPTY_BORDER
    cell.fill = EMPTY_FILL
    cell.number_format = 'General'


def clear_row(ws, row_number, columns, visual=False):
    for column in columns:
        coordinate = f'{column}{row_number}'
        if visual:
            clear_visual_cell(ws, coordinate)
        else:
            clear_cell_value(ws, coordinate)


def copy_layout(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def copy_row_layout(ws, source_row: int, target_row: int, columns):
    for column in columns:
        copy_layout(ws[f'{column}{source_row}'], ws[f'{column}{target_row}'])
        clear_cell_value(ws, f'{column}{target_row}')


def split_bank_and_iban(value: str):
    text = str(value or '').strip()
    if not text:
        return '', ''
    marker = text.find('TR')
    if marker <= 0:
        return text, ''
    return text[:marker].strip(), text[marker:].strip()


def normalize_bank_layout(ws, template_meta: dict):
    header_row = template_meta.get('bank_header_row')
    if header_row:
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row == header_row and merged_range.max_row == header_row and merged_range.min_col >= 11 and merged_range.max_col <= 14:
                ws.unmerge_cells(str(merged_range))
        copy_layout(ws[f'C{header_row}'], ws[f'K{header_row}'])
        copy_layout(ws[f'D{header_row}'], ws[f'L{header_row}'])
        copy_layout(ws[f'D{header_row}'], ws[f'M{header_row}'])
        ws[f'K{header_row}'] = ws[f'K{header_row}'].value or ''
        ws[f'L{header_row}'] = ''
        ws[f'M{header_row}'] = ''
        clear_visual_cell(ws, f'N{header_row}')
        target_merge = f'K{header_row}:M{header_row}'
        if all(str(existing) != target_merge for existing in ws.merged_cells.ranges):
            ws.merge_cells(target_merge)

    for row_number in template_meta.get('bank_rows', []):
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row == row_number and merged_range.max_row == row_number and merged_range.min_col >= 11 and merged_range.max_col <= 14:
                ws.unmerge_cells(str(merged_range))

        right_bank, right_iban = split_bank_and_iban(ws[f'K{row_number}'].value)
        if not right_iban and ws[f'L{row_number}'].value:
            right_iban = str(ws[f'L{row_number}'].value or '').strip()
        if not right_iban and ws[f'M{row_number}'].value:
            right_iban = str(ws[f'M{row_number}'].value or '').strip()

        left_name_cell = ws[f'C{row_number}']
        left_iban_cell = ws[f'D{row_number}']
        for target in [ws[f'K{row_number}'], ws[f'L{row_number}']]:
            copy_layout(left_name_cell, target)
        copy_layout(left_iban_cell, ws[f'M{row_number}'])

        ws[f'K{row_number}'] = right_bank
        ws[f'L{row_number}'] = ''
        ws[f'M{row_number}'] = right_iban
        clear_visual_cell(ws, f'N{row_number}')

        for target_merge in [f'K{row_number}:L{row_number}']:
            if all(str(existing) != target_merge for existing in ws.merged_cells.ranges):
                ws.merge_cells(target_merge)


def normalize_template(path: Path, template_meta: dict):
    workbook = load_workbook(path)
    workbook.template = True
    worksheet = workbook['1'] if '1' in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    for coordinate in HEADER_CLEAR_COORDS:
        clear_cell_value(worksheet, coordinate)
    clear_row(worksheet, 31, ROW31_COLUMNS, visual=False)

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row == 32 and merged_range.max_row == 32 and merged_range.min_col >= 4 and merged_range.max_col <= 13:
            worksheet.unmerge_cells(str(merged_range))
    copy_row_layout(worksheet, 31, 32, ROW32_COLUMNS)

    for block in template_meta.get('line_blocks', []):
        for row_number in range(block['start_row'], block['start_row'] + block['capacity']):
            worksheet.row_dimensions[row_number].hidden = False
            clear_row(worksheet, row_number, DATA_COLUMNS, visual=False)
        for summary_row in [block['subtotal_row'], block['tax_row'], block['grand_row']]:
            worksheet.row_dimensions[summary_row].hidden = False
            clear_row(worksheet, summary_row, SUMMARY_CLEAR_COLUMNS, visual=True)
            clear_cell_value(worksheet, f'L{summary_row}')
            clear_cell_value(worksheet, f'M{summary_row}')

    for row_number in template_meta.get('commercial_rows', []):
        clear_cell_value(worksheet, f'D{row_number}')

    normalize_bank_layout(worksheet, template_meta)

    workbook.save(path)


def main():
    seen_paths: set[Path] = set()
    for template in TEMPLATE_REGISTRY:
        source_path = Path(template['source_path'])
        if source_path in seen_paths:
            continue
        normalize_template(source_path, template)
        seen_paths.add(source_path)
        print(f'normalized {source_path.as_posix().encode("unicode_escape").decode("ascii")}')

    steel_template = next(item for item in TEMPLATE_REGISTRY if item['template_key'] == 'quote_steel')
    for download_copy in Path(r'C:/Users/Enes3078/Downloads').glob(DOWNLOAD_TEMPLATE_GLOB):
        normalize_template(download_copy, steel_template)
        print(f'normalized {download_copy.as_posix().encode("unicode_escape").decode("ascii")}')


if __name__ == '__main__':
    main()
