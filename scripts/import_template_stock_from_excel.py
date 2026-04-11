from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXCEL_EXTENSIONS = {'.xlsx', '.xltx', '.xlsm', '.xltm'}


CATEGORY_META = {
    'celik_kapi': {
        'name': 'Excel / Çelik Kapı',
        'section_key': 'ÇELİK KAPI',
        'unit': 'Adet',
        'tax': 20,
        'template_family': 'excel_celik_kapi',
    },
    'ic_oda': {
        'name': 'Excel / İç Oda Kapısı',
        'section_key': 'İÇ ODA KAPISI',
        'unit': 'Adet',
        'tax': 20,
        'template_family': 'excel_ic_oda',
    },
    'montaj': {
        'name': 'Excel / Montaj',
        'section_key': 'MONTAJ',
        'unit': 'Adet',
        'tax': 20,
        'template_family': 'excel_montaj',
    },
    'mobilya_mutfak': {
        'name': 'Excel / Mobilya / Mutfak',
        'section_key': 'MUTFAK',
        'unit': 'Metre',
        'tax': 10,
        'template_family': 'excel_mobilya_mutfak',
    },
    'mobilya_portmanto': {
        'name': 'Excel / Mobilya / Portmanto',
        'section_key': 'PORTMANTO',
        'unit': 'm²',
        'tax': 10,
        'template_family': 'excel_mobilya_portmanto',
    },
    'mobilya_banyo': {
        'name': 'Excel / Mobilya / Banyo',
        'section_key': 'BANYO',
        'unit': 'Adet',
        'tax': 10,
        'template_family': 'excel_mobilya_banyo',
    },
}

SKU_PREFIX = {
    'celik_kapi': 'EXCEL-CK',
    'ic_oda': 'EXCEL-IK',
    'montaj': 'EXCEL-MTJ',
    'mobilya_mutfak': 'EXCEL-MTF',
    'mobilya_portmanto': 'EXCEL-PRT',
    'mobilya_banyo': 'EXCEL-BNY',
}


def canonical_excel_files(base: Path) -> list[Path]:
    paths: list[Path] = []
    teklif = base / 'TEKLIF_TASLAKLARI'
    sozlesme = base / 'SOZLEŞME_TASLAKLARI' / 'PARÇALI_LISTE'
    for root in (teklif, sozlesme):
        if not root.exists():
            continue
        paths.extend(sorted(p for p in root.rglob('*') if p.suffix.lower() in EXCEL_EXTENSIONS))
    return paths


def norm(text: str) -> str:
    return re.sub(r'\s+', ' ', (text or '').strip()).lower()


def infer_group_key(path: Path, sheet_name: str, product_name: str) -> str | None:
    path_upper = ' '.join(part.upper() for part in path.parts)
    sheet_upper = sheet_name.upper()
    name_upper = product_name.upper()

    if 'MUTFAK' in sheet_upper or name_upper.startswith('MUTFAK'):
        return 'mobilya_mutfak'
    if 'PORTMANTO' in sheet_upper or 'GIYINME' in name_upper or 'GİYİNME' in name_upper or name_upper.startswith('PORT-'):
        return 'mobilya_portmanto'
    if 'BANYO' in sheet_upper:
        return 'mobilya_banyo'
    if 'MOBILYA' in path_upper and 'MUTFAK' in path_upper:
        return 'mobilya_mutfak'
    if 'MOBILYA' in path_upper and ('PORTMANTO' in path_upper or 'GIYINME' in path_upper or 'GİYİNME' in path_upper):
        return 'mobilya_portmanto'
    if 'MOBILYA' in path_upper and 'BANYO' in path_upper:
        return 'mobilya_banyo'
    if 'MONTAJ' in path_upper or 'MONTAJ' in sheet_upper:
        return 'montaj'
    if 'IC_ODA' in path_upper or 'İC_ODA' in path_upper or 'IC_KAPI' in path_upper or 'ICKAPI' in sheet_upper:
        return 'ic_oda'
    if 'CELIK' in path_upper or 'ÇELIK' in path_upper or 'CELIKKAPI' in sheet_upper:
        return 'celik_kapi'
    return None


def build_category_schema() -> list[dict]:
    base_fields = [
        ('original_code', 'Excel kodu', 'text'),
        ('source_unit', 'Excel birimi', 'text'),
        ('document_types', 'Belge türleri', 'text'),
        ('source_files', 'Kaynak dosyalar', 'textarea'),
        ('source_sheets', 'Kaynak sayfalar', 'textarea'),
        ('source_rows', 'Kaynak satırlar', 'textarea'),
        ('source_prices', 'Kaynak fiyatlar', 'textarea'),
    ]
    return [
        {
            'field_key': field_key,
            'label': label,
            'type': field_type,
            'required': False,
            'order': index + 1,
            'applies_to_documents': 'both',
        }
        for index, (field_key, label, field_type) in enumerate(base_fields)
    ]


def category_payload(group_key: str) -> dict:
    meta = CATEGORY_META[group_key]
    return {
        'name': meta['name'],
        'template_defaults': {
            'section_key': meta['section_key'],
            'unit': meta['unit'],
            'tax': meta['tax'],
            'discount': 0,
            'discount_secondary': 0,
            'template_family': meta['template_family'],
            'import_origin': 'excel_templates',
        },
        'attribute_schema': build_category_schema(),
    }


@dataclass
class ProductRecord:
    group_key: str
    original_code: str
    name: str
    price: float
    source_unit: str
    document_types: set[str] = field(default_factory=set)
    source_files: set[str] = field(default_factory=set)
    source_sheets: set[str] = field(default_factory=set)
    source_rows: list[str] = field(default_factory=list)
    source_prices: list[str] = field(default_factory=list)

    def merge_source(self, document_type: str, file_name: str, sheet_name: str, row_number: int, price: float):
        self.document_types.add(document_type)
        self.source_files.add(file_name)
        self.source_sheets.add(sheet_name)
        self.source_rows.append(f'{file_name}:{sheet_name}!{row_number}')
        self.source_prices.append(f'{file_name}:{sheet_name} -> {price:.2f}')


def infer_unit(group_key: str, product_name: str) -> str:
    if group_key == 'mobilya_mutfak':
        return 'Metre'
    if group_key == 'mobilya_portmanto':
        return 'm²'
    if group_key == 'mobilya_banyo':
        return 'Adet'
    if ' M²' in product_name.upper():
        return 'm²'
    return CATEGORY_META[group_key]['unit']


def document_type_for_path(path: Path) -> str:
    upper = str(path).upper()
    if 'TEKLIF' in upper:
        return 'Teklif'
    if 'SOZLE' in upper or 'SÖZLE' in upper:
        return 'Sözleşme'
    return 'Belge'


def parse_catalog_rows(base: Path) -> tuple[dict[str, dict], dict[tuple[str, str, str, float], ProductRecord]]:
    categories: dict[str, dict] = {}
    products: dict[tuple[str, str, str, float], ProductRecord] = {}

    for path in canonical_excel_files(base):
        workbook = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                continue
            headers = [str(cell).strip().lower() if cell is not None else '' for cell in first_row]
            if not any('kod' in header for header in headers):
                continue

            for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                code = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                price = row[2] if len(row) > 2 else None
                if code in (None, '') or name in (None, ''):
                    continue

                original_code = str(code).strip().strip('"')
                product_name = re.sub(r'\s+', ' ', str(name).strip())
                group_key = infer_group_key(path, sheet_name, product_name)
                if not group_key:
                    continue
                categories[group_key] = category_payload(group_key)
                price_number = round(float(price or 0), 2)
                unit = infer_unit(group_key, product_name)
                product_key = (group_key, original_code, product_name, price_number)
                document_type = document_type_for_path(path)

                if product_key not in products:
                    products[product_key] = ProductRecord(
                        group_key=group_key,
                        original_code=original_code,
                        name=product_name,
                        price=price_number,
                        source_unit=unit,
                    )
                products[product_key].merge_source(document_type, path.name, sheet_name, row_index, price_number)

    return categories, products


def payload_from_excel(base: Path) -> dict:
    categories, products = parse_catalog_rows(base)
    sku_counters: defaultdict[str, int] = defaultdict(int)
    product_payloads = []

    for product in sorted(products.values(), key=lambda item: (item.group_key, item.original_code, item.name, item.price)):
        base_sku = f"{SKU_PREFIX[product.group_key]}-{product.original_code}"
        sku_counters[base_sku] += 1
        sku = base_sku if sku_counters[base_sku] == 1 else f"{base_sku}-{sku_counters[base_sku]}"
        meta = CATEGORY_META[product.group_key]

        product_payloads.append(
            {
                'sku': sku,
                'name': product.name,
                'category_name': meta['name'],
                'price': product.price,
                'stock': 0,
                'reserved': 0,
                'reorder_point': 0,
                'template_defaults': {
                    'section_key': meta['section_key'],
                    'unit': product.source_unit,
                    'tax': meta['tax'],
                    'discount': 0,
                    'discount_secondary': 0,
                    'template_family': meta['template_family'],
                    'import_origin': 'excel_templates',
                },
                'attribute_values': {
                    'import_origin': 'excel_templates',
                    'original_code': product.original_code,
                    'source_unit': product.source_unit,
                    'document_types': ', '.join(sorted(product.document_types)),
                    'source_files': '\n'.join(sorted(product.source_files)),
                    'source_sheets': '\n'.join(sorted(product.source_sheets)),
                    'source_rows': '\n'.join(product.source_rows),
                    'source_prices': '\n'.join(product.source_prices),
                },
                'attribute_schema_override': [],
            }
        )

    return {
        'categories': [categories[key] for key in sorted(categories)],
        'products': product_payloads,
    }


def apply_payload_via_docker(payload: dict, workdir: Path):
    command = [
        'docker',
        'compose',
        'exec',
        '-T',
        'backend',
        'python',
        'manage.py',
        'import_template_catalog_json',
    ]
    completed = subprocess.run(
        command,
        cwd=workdir,
        input=json.dumps(payload, ensure_ascii=False),
        text=True,
        encoding='utf-8',
        errors='strict',
        capture_output=True,
        check=False,
    )
    if completed.returncode != 0:
        sys.stderr.write(completed.stderr or completed.stdout)
        raise SystemExit(completed.returncode)
    sys.stdout.write(completed.stdout)


def main():
    parser = argparse.ArgumentParser(description='Parse Excel templates and import stock products into Docker backend.')
    parser.add_argument('--excel-root', required=True, help='Root TASLAKLAR directory')
    parser.add_argument('--output-json', help='Optional path to save the generated payload')
    parser.add_argument('--apply-docker', action='store_true', help='Send payload to docker compose backend')
    args = parser.parse_args()

    excel_root = Path(args.excel_root)
    if not excel_root.exists():
        raise SystemExit(f'Excel root not found: {excel_root}')

    payload = payload_from_excel(excel_root)
    summary = {
        'categories': len(payload['categories']),
        'products': len(payload['products']),
    }
    print(json.dumps(summary, ensure_ascii=False))

    if args.output_json:
        output_path = Path(args.output_json)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    if args.apply_docker:
        repo_root = Path(__file__).resolve().parents[1]
        apply_payload_via_docker(payload, repo_root)


if __name__ == '__main__':
    main()
