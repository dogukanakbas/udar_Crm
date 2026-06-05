import copy
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import FileResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from crm.contracts import PDF_CONTENT_TYPE, _convert_xlsx_stream_to_pdf, resolve_product_document_defaults
from crm.models import Quote

from .models import ProductionReportTemplate, ProductionWorkOrder

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
PLACEHOLDER_PATTERN = re.compile(r'\{([A-Za-z0-9_.]+)\}')
LINE_TOKEN_PATTERN = re.compile(r'\{kalem1\.')


def list_production_report_placeholders():
    return [
        {
            'title': 'Genel',
            'items': [
                {'token': '{isEmriNo}', 'label': 'İş emri numarası'},
                {'token': '{sozlesmeNo}', 'label': 'Sözleşme / sipariş numarası'},
                {'token': '{cariUnvani}', 'label': 'Cari ünvanı'},
                {'token': '{musteriIsmi}', 'label': 'Müşteri ismi / yetkili'},
                {'token': '{temsilci}', 'label': 'Temsilci / hazırlayan'},
                {'token': '{siparisTarihi}', 'label': 'Sipariş tarihi'},
                {'token': '{teslimTarihi}', 'label': 'Teslim tarihi'},
                {'token': '{toplamAdet}', 'label': 'Toplam adet'},
                {'token': '{notlar}', 'label': 'Notlar ve ek açıklamalar'},
                {'token': '{teknikDetaylar}', 'label': 'Ürün teknik detayları'},
            ],
        },
        {
            'title': 'Kalem Satırı',
            'items': [
                {'token': '{kalem1.sn}', 'label': 'Sıra no'},
                {'token': '{kalem1.urunKodu}', 'label': 'Ürün kodu'},
                {'token': '{kalem1.urunAdi}', 'label': 'Ürün adı'},
                {'token': '{kalem1.urunTipi}', 'label': 'Ürün tipi / satış birimi'},
                {'token': '{kalem1.adet}', 'label': 'Adet'},
                {'token': '{kalem1.detay1}', 'label': 'Detay-1 / ölçü'},
                {'token': '{kalem1.detay2}', 'label': 'Detay-2 / renk'},
                {'token': '{kalem1.olcu}', 'label': 'Ölçü'},
                {'token': '{kalem1.renk}', 'label': 'Renk'},
                {'token': '{kalem1.aksesuar}', 'label': 'Aksesuar'},
                {'token': '{kalem1.aciklama}', 'label': 'Açıklama'},
                {'token': '{kalem1.teknikDetaylar}', 'label': 'Kalem teknik detayları'},
            ],
            'hint': 'Sınırsız kalem için şablonda bir örnek satıra {kalem1.*} placeholderlarını yazın; sistem satırı çoğaltır.',
        },
    ]


def build_work_order_report_export(work_order, template, output_format='xlsx', extra_notes=''):
    quote = _quote_for_work_order(work_order)
    lines = [_line_context_from_work_order(line, index) for index, line in enumerate(work_order.lines.all().order_by('sort_order', 'id'), start=1)]
    context = _general_context_from_work_order(work_order, quote, lines, extra_notes)
    return _render_template_export(template, context, lines, output_format, work_order.number or 'is_emri')


def build_quote_report_export(quote, template, output_format='xlsx', extra_notes=''):
    lines = [_line_context_from_quote(line, index) for index, line in enumerate(quote.lines.all().order_by('sort_order', 'id'), start=1)]
    context = _general_context_from_quote(quote, lines, extra_notes)
    return _render_template_export(template, context, lines, output_format, quote.number or 'sozlesme')


def response_for_export(export):
    return FileResponse(
        export['content'],
        as_attachment=True,
        filename=export['filename'],
        content_type=export['content_type'],
    )


def _render_template_export(template, general_context, line_contexts, output_format, base_name):
    if not isinstance(template, ProductionReportTemplate):
        raise ValueError('Geçerli rapor şablonu bulunamadı.')
    output_format = str(output_format or template.default_format or 'xlsx').lower()
    if output_format not in {'xlsx', 'pdf'}:
        raise ValueError('Çıktı tipi xlsx veya pdf olmalı.')
    if not template.file:
        raise ValueError('Rapor şablon dosyası bulunamadı.')

    try:
        workbook = load_workbook(template.file.path)
    except Exception as exc:
        raise ValueError('Excel rapor şablonu okunamadı.') from exc

    worksheet = workbook.active
    _expand_line_rows(worksheet, len(line_contexts))
    context = dict(general_context)
    for line in line_contexts:
        prefix = f"kalem{line['sn']}."
        for key, value in line.items():
            context[f'{prefix}{key}'] = value

    _apply_placeholders(worksheet, context)
    _prepare_report_print_layout(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_base = _safe_filename(base_name)
    if output_format == 'pdf':
        pdf = _convert_xlsx_stream_to_pdf(output, f'{safe_base}.xlsx')
        return {
            'content': pdf,
            'filename': f'{safe_base}-is-emri-raporu.pdf',
            'content_type': PDF_CONTENT_TYPE,
        }

    return {
        'content': output,
        'filename': f'{safe_base}-is-emri-raporu.xlsx',
        'content_type': XLSX_CONTENT_TYPE,
    }


def _quote_for_work_order(work_order):
    if str(work_order.source_type or '').lower() != 'contract' or not work_order.source_id:
        return None
    return (
        Quote.objects.filter(organization=work_order.organization, pk=work_order.source_id)
        .select_related('customer', 'prepared_by', 'owner')
        .prefetch_related('lines__product__category')
        .first()
    )


def _general_context_from_work_order(work_order, quote, lines, extra_notes):
    customer = getattr(quote, 'customer', None)
    technical_details = _technical_details_text(lines)
    notes = _join_notes(technical_details, work_order.notes, getattr(quote, 'notes', ''), extra_notes)
    
    temsilci = _resolve_prepared_by_name(quote) if quote else ''
    if not temsilci:
        temsilci = _user_label(work_order.created_by)

    return {
        'isEmriNo': work_order.number,
        'sozlesmeNo': work_order.source_number or getattr(quote, 'number', '') or '',
        'cariUnvani': getattr(customer, 'name', '') or work_order.customer_name,
        'musteriIsmi': _customer_person(quote) or getattr(customer, 'name', '') or work_order.customer_name,
        'temsilci': temsilci,
        'siparisTarihi': _format_date(getattr(quote, 'created_at', None) or work_order.created_at),
        'teslimTarihi': _format_date(work_order.due_date or getattr(quote, 'valid_until', None)),
        'toplamAdet': sum((_decimal(line.get('adet')) for line in lines), Decimal('0')),
        'notlar': notes,
        'teknikDetaylar': technical_details,
    }


def _general_context_from_quote(quote, lines, extra_notes):
    customer = getattr(quote, 'customer', None)
    technical_details = _technical_details_text(lines)
    return {
        'isEmriNo': '',
        'sozlesmeNo': quote.number,
        'cariUnvani': getattr(customer, 'name', '') or '',
        'musteriIsmi': _customer_person(quote) or getattr(customer, 'name', '') or '',
        'temsilci': _resolve_prepared_by_name(quote),
        'siparisTarihi': _format_date(quote.created_at),
        'teslimTarihi': _format_date(quote.valid_until),
        'toplamAdet': sum((_decimal(line.get('adet')) for line in lines), Decimal('0')),
        'notlar': _join_notes(technical_details, quote.notes, extra_notes),
        'teknikDetaylar': technical_details,
    }


def _line_context_from_work_order(line, index):
    details = line.details or {}
    technical_items = _line_technical_items(line)
    line_note = line.technical_notes or _details_pick(details, 'aciklama', 'açıklama', 'note', 'notes', 'technical_notes')
    product_code = (
        line.product_sku
        or getattr(line.product, 'sku', '')
        or _details_pick(details, 'product_sku', 'productSku', 'productCode', 'product_code', 'urunKodu', 'urun_kodu', 'kod', 'code', 'sku')
        or ''
    )
    product_name = (
        line.product_name
        or getattr(line.product, 'name', '')
        or _details_pick(details, 'product_name', 'productName', 'urunAdi', 'urun_adi', 'urunTipi', 'name')
        or product_code
    )
    return {
        'sn': index,
        'kod': product_code,
        'urunKodu': product_code,
        'urun_kodu': product_code,
        'sku': product_code,
        'productCode': product_code,
        'product_code': product_code,
        'urunAdi': product_name,
        'urun_adi': product_name,
        'productName': product_name,
        'product_name': product_name,
        'name': product_name,
        'urunTipi': product_name,
        'adet': line.quantity,
        'detay1': line.detail_1,
        'detay2': line.detail_2,
        'olcu': line.detail_1,
        'renk': line.detail_2,
        'aksesuar': _details_pick(details, 'aksesuar', 'accessory', 'hardware', 'kilit', 'kol'),
        'aciklama': _join_notes(line_note, '\n'.join(technical_items)),
        'teknikDetaylar': '\n'.join(technical_items),
    }


def _line_context_from_quote(line, index):
    details = line.details or {}
    detail_1 = details.get('primary') or details.get('detail_1') or details.get('detail1') or details.get('olcu') or details.get('ölçü') or ''
    detail_2 = details.get('secondary') or details.get('detail_2') or details.get('detail2') or details.get('renk') or ''
    technical_items = _line_technical_items(line)
    line_note = _details_pick(details, 'aciklama', 'açıklama', 'note', 'notes', 'technical_notes')
    product_code = (
        getattr(line.product, 'sku', '')
        or _details_pick(details, 'product_sku', 'productSku', 'productCode', 'product_code', 'urunKodu', 'urun_kodu', 'kod', 'code', 'sku')
        or ''
    )
    product_name = (
        line.name
        or getattr(line.product, 'name', '')
        or _details_pick(details, 'product_name', 'productName', 'urunAdi', 'urun_adi', 'urunTipi', 'name')
        or product_code
    )
    return {
        'sn': index,
        'kod': product_code,
        'urunKodu': product_code,
        'urun_kodu': product_code,
        'sku': product_code,
        'productCode': product_code,
        'product_code': product_code,
        'urunAdi': product_name,
        'urun_adi': product_name,
        'productName': product_name,
        'product_name': product_name,
        'name': product_name,
        'urunTipi': product_name,
        'adet': line.qty,
        'detay1': detail_1,
        'detay2': detail_2,
        'olcu': detail_1,
        'renk': detail_2,
        'aksesuar': _details_pick(details, 'aksesuar', 'accessory', 'hardware', 'kilit', 'kol'),
        'aciklama': _join_notes(line_note, '\n'.join(technical_items)),
        'teknikDetaylar': '\n'.join(technical_items),
    }


def _expand_line_rows(worksheet, line_count):
    template_row = None
    for row in worksheet.iter_rows():
        if any(isinstance(cell.value, str) and LINE_TOKEN_PATTERN.search(cell.value) for cell in row):
            template_row = row[0].row
            break
    if not template_row or line_count <= 1:
        return

    amount = line_count - 1
    max_row_before = worksheet.max_row
    merged_ranges = list(worksheet.merged_cells.ranges)
    row_dimensions = {
        row_index: _row_dimension_snapshot(worksheet.row_dimensions[row_index])
        for row_index in range(template_row + 1, max_row_before + 1)
    }
    template_row_dimension = _row_dimension_snapshot(worksheet.row_dimensions[template_row])

    for merged in merged_ranges:
        worksheet.unmerge_cells(str(merged))

    worksheet.insert_rows(template_row + 1, amount=amount)
    for source_row, snapshot in sorted(row_dimensions.items(), reverse=True):
        _apply_row_dimension_snapshot(worksheet.row_dimensions[source_row + amount], snapshot)
    for row_index in range(template_row + 1, template_row + line_count):
        _apply_row_dimension_snapshot(worksheet.row_dimensions[row_index], template_row_dimension)

    for row_index in range(template_row + 1, template_row + line_count):
        _copy_row(worksheet, template_row, row_index)

    _restore_shifted_merges(worksheet, merged_ranges, template_row, amount)

    for row_index in range(template_row, template_row + line_count):
        line_number = row_index - template_row + 1
        for cell in worksheet[row_index]:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace('{kalem1.', f'{{kalem{line_number}.')


def _row_dimension_snapshot(row_dimension):
    return {
        'height': row_dimension.height,
        'hidden': row_dimension.hidden,
        'outlineLevel': row_dimension.outlineLevel,
        'collapsed': row_dimension.collapsed,
    }


def _apply_row_dimension_snapshot(row_dimension, snapshot):
    row_dimension.height = snapshot.get('height')
    row_dimension.hidden = snapshot.get('hidden')
    row_dimension.outlineLevel = snapshot.get('outlineLevel') or 0
    row_dimension.collapsed = snapshot.get('collapsed') or False


def _restore_shifted_merges(worksheet, merged_ranges, template_row, inserted_count):
    restored = set()
    for merged in merged_ranges:
        if merged.min_row > template_row:
            ref = _merge_ref(
                merged.min_row + inserted_count,
                merged.min_col,
                merged.max_row + inserted_count,
                merged.max_col,
            )
        elif merged.min_row <= template_row < merged.max_row:
            ref = _merge_ref(merged.min_row, merged.min_col, merged.max_row + inserted_count, merged.max_col)
        else:
            ref = _merge_ref(merged.min_row, merged.min_col, merged.max_row, merged.max_col)
        if ref not in restored:
            worksheet.merge_cells(ref)
            restored.add(ref)

        if merged.min_row == template_row and merged.max_row == template_row:
            for offset in range(1, inserted_count + 1):
                copied_ref = _merge_ref(
                    merged.min_row + offset,
                    merged.min_col,
                    merged.max_row + offset,
                    merged.max_col,
                )
                if copied_ref not in restored:
                    worksheet.merge_cells(copied_ref)
                    restored.add(copied_ref)


def _merge_ref(min_row, min_col, max_row, max_col):
    return f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}'


def _copy_row(worksheet, source_row, target_row):
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    worksheet.row_dimensions[target_row].hidden = worksheet.row_dimensions[source_row].hidden
    for col in range(1, worksheet.max_column + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        target.value = source.value
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)

def _apply_placeholders(worksheet, context):
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or '{' not in cell.value or '}' not in cell.value:
                continue
            _replace_placeholder_in_cell(cell, context)


def _replace_placeholder_in_cell(cell, context):
    source = str(cell.value or '')
    matches = list(PLACEHOLDER_PATTERN.finditer(source))
    if not matches:
        return
    if len(matches) == 1 and matches[0].span() == (0, len(source)):
        cell.value = _typed_value(context.get(matches[0].group(1), ''))
        return
    rendered = source
    for match in matches:
        rendered = rendered.replace(match.group(0), _display_value(context.get(match.group(1), '')))
    cell.value = rendered


def _prepare_report_print_layout(worksheet):
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    worksheet.print_options.horizontalCentered = True


def _typed_value(value):
    if value in [None, '']:
        return ''
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.date()
    return value


def _display_value(value):
    if value in [None, '']:
        return ''
    if isinstance(value, Decimal):
        return _format_decimal(value)
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    return str(value)


def _format_decimal(value):
    value = Decimal(value)
    if value == value.to_integral_value():
        return f'{value:.0f}'
    return f'{value.normalize()}'


def _decimal(value):
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal('0')


def _user_label(user):
    if not user:
        return ''
    return user.get_full_name() or getattr(user, 'email', '') or getattr(user, 'username', '') or ''


def _format_date(value):
    if not value:
        return ''
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    return str(value)


def _resolve_prepared_by_name(quote):
    if not quote:
        return ''
    snapshot = dict((quote.contract_config or {}).get('prepared_by_snapshot') or {})
    snapshot_name = str(snapshot.get('name') or '').strip()
    related_name = _user_label(getattr(quote, 'prepared_by', None)) or _user_label(getattr(quote, 'owner', None))
    if snapshot_name and '@' not in snapshot_name:
        return snapshot_name
    return related_name or snapshot_name


def _customer_person(quote):
    if not quote:
        return ''
    snapshot = quote.contract_config.get('customer_snapshot') if isinstance(quote.contract_config, dict) else {}
    if not isinstance(snapshot, dict):
        snapshot = {}
    return (
        snapshot.get('authorized_person')
        or snapshot.get('authorizedPerson')
        or snapshot.get('contact_name')
        or snapshot.get('contactName')
        or ''
    )


def _details_pick(details, *keys):
    if not isinstance(details, dict):
        return ''
    candidates = [details]
    attributes = details.get('attributes')
    if isinstance(attributes, dict):
        candidates.append(attributes)
    for source in candidates:
        for key in keys:
            value = source.get(key)
            if value not in [None, '']:
                return value
    return ''


def _line_technical_items(line):
    details = dict(getattr(line, 'details', {}) or {})
    items = details.get('technicalItems') or details.get('technical_items')
    if not isinstance(items, list):
        product = getattr(line, 'product', None)
        section_key = getattr(line, 'section_key', '') or getattr(product, 'section_key', '')
        line_name = getattr(line, 'name', '') or getattr(line, 'product_name', '')
        items = resolve_product_document_defaults(product, fallback_section_key=section_key, line_name=line_name).get('technical_items') if product else []
    normalized = [str(item or '').strip() for item in items if str(item or '').strip()]
    technical_note = _details_pick(details, 'technical_notes', 'technicalItemsText', 'technical_items_text')
    if technical_note:
        normalized.append(str(technical_note).strip())
    return normalized


def _technical_details_text(lines):
    rows = []
    index = 1
    for line in lines:
        product_name = str(line.get('urunTipi') or '').strip()
        for item in str(line.get('teknikDetaylar') or '').splitlines():
            item = item.strip()
            if not item:
                continue
            prefix = f'Açıklama {index}:'
            rows.append(f'{prefix} {product_name} - {item}' if product_name else f'{prefix} {item}')
            index += 1
    return '\n'.join(rows)


def _join_notes(*values):
    return '\n'.join(str(value).strip() for value in values if str(value or '').strip())


def _safe_filename(value):
    raw = str(value or 'rapor').strip() or 'rapor'
    return re.sub(r'[^A-Za-z0-9_.-]+', '_', raw)
