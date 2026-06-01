from datetime import timedelta
from io import BytesIO
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from permissions import IsOrgMember, HasAPIPermission
from organizations.models import Organization, NumberRange, Warehouse
from core.events import push_event
from .inventory_service import InventoryError, adjust, allocate_opening_balance, as_decimal, stock_in, stock_out, transfer
from .models import InventoryLocation, Invoice, Product, Category, SalesOrder, PurchaseOrder, StockMovement, Vehicle, WarehouseStock
from .template_catalog_import import upsert_product_catalog, upsert_template_catalog
from .serializers import (
    InvoiceSerializer,
    ProductSerializer,
    CategorySerializer,
    SalesOrderSerializer,
    PurchaseOrderSerializer,
    StockMovementSerializer,
    VehicleSerializer,
    WarehouseSerializer,
    InventoryLocationSerializer,
    WarehouseStockSerializer,
)


def _ensure_org(request):
    org = getattr(getattr(request, 'user', None), 'organization', None) if request else None
    if org:
        return org
    org = Organization.objects.first()
    if not org:
        org = Organization.objects.create(name='Default Org')
    return org


def _normalize_reorder_payload(model, org, new_positions, item_label):
    if not isinstance(new_positions, list) or not new_positions:
        return None, Response({'detail': 'new_positions listesi gerekli'}, status=status.HTTP_400_BAD_REQUEST)

    normalized = []
    for index, position in enumerate(new_positions):
        if not isinstance(position, dict) or position.get('id') in [None, '']:
            return None, Response({'detail': f'{item_label} ID bilgisi eksik.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            item_id = int(position.get('id'))
        except (TypeError, ValueError):
            return None, Response({'detail': f'{item_label} ID bilgisi geçersiz.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            order = int(position.get('order', index))
        except (TypeError, ValueError):
            order = index
        normalized.append((item_id, max(order, 0)))

    ids = [item_id for item_id, _ in normalized]
    if len(ids) != len(set(ids)):
        return None, Response({'detail': f'Aynı {item_label.lower()} birden fazla kez gönderilemez.'}, status=status.HTTP_400_BAD_REQUEST)

    instances_by_id = {
        instance.id: instance
        for instance in model.objects.filter(organization=org, id__in=ids)
    }
    if set(ids) != set(instances_by_id):
        return None, Response(
            {'detail': f'Bazı {item_label.lower()} ID bilgileri geçersiz veya bu organizasyona ait değil.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    return [(instances_by_id[item_id], order) for item_id, order in normalized], None


class OrgScopedMixin:
    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(organization=org)
        return qs

    def perform_create(self, serializer):
        org = _ensure_org(getattr(self, 'request', None))
        serializer.save(organization=org)


class InvoiceViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = InvoiceSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'invoices.view'
    permission_map = {
        'create': 'invoices.edit',
        'update': 'invoices.edit',
        'partial_update': 'invoices.edit',
        'destroy': 'invoices.edit',
    }
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['number', 'customer_name', 'status']
    ordering_fields = ['issued_at', 'amount']
    queryset = Invoice.objects.all()


class ProductViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'products.view'
    permission_map = {
        'create': 'products.create',
        'update': 'products.edit',
        'partial_update': 'products.edit',
        'destroy': 'products.delete',
        'import_template_catalog': 'products.import',
        'bulk_delete': 'products.bulk.delete',
        'bulk_upsert': 'products.import',
        'reorder': 'products.edit',
    }
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['order', 'created_at', 'sku', 'name']
    ordering = ['order', 'id']  # Default ordering
    queryset = Product.objects.all()

    @action(detail=False, methods=['post'], url_path='import-template-catalog')
    def import_template_catalog(self, request):
        data = request.data if isinstance(request.data, dict) else {}
        categories_data = data.get('categories') or []
        products_data = data.get('products') or []
        if not isinstance(categories_data, list) or not isinstance(products_data, list):
            return Response({'detail': 'categories ve products liste olmalidir'}, status=status.HTTP_400_BAD_REQUEST)

        org = _ensure_org(request)

        with transaction.atomic():
            result = upsert_template_catalog(org, categories_data, products_data, user=request.user)
        return Response(result)

    @action(detail=False, methods=['post'], url_path='bulk-upsert')
    def bulk_upsert(self, request):
        data = request.data if isinstance(request.data, dict) else {}
        categories_data = data.get('categories') or []
        products_data = data.get('products') or []
        if not isinstance(categories_data, list) or not isinstance(products_data, list):
            return Response({'detail': 'categories ve products liste olmalidir'}, status=status.HTTP_400_BAD_REQUEST)

        org = _ensure_org(request)
        with transaction.atomic():
            result = upsert_product_catalog(
                org,
                categories_data,
                products_data,
                user=request.user,
                audit_entity='BulkProductImport',
                audit_entity_id='bulk-product-import',
                audit_action='imported',
                audit_field='bulk_product_import',
            )
        return Response(result)

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        data = request.data if isinstance(request.data, dict) else {}
        org = _ensure_org(request)
        ids = data.get('ids') or []
        delete_all = bool(data.get('all'))

        queryset = Product.objects.filter(organization=org)
        if delete_all:
            if data.get('confirm') is not True:
                return Response({'detail': 'Tüm ürünleri silmek için confirm=true gönderilmelidir.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            if not isinstance(ids, list) or not ids:
                return Response({'detail': 'Silinecek ürün id listesi veya all=true gerekli.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(id__in=ids)

        product_count = queryset.count()
        stock_movement_count = StockMovement.objects.filter(organization=org, product__in=queryset).count()
        with transaction.atomic():
            queryset.delete()

        return Response(
            {
                'deleted_products': product_count,
                'deleted_stock_movements': stock_movement_count,
            }
        )

    @action(detail=False, methods=['post'], url_path='reorder')
    def reorder(self, request):
        """Reorder products by updating their order field. Accepts list of {id, order} tuples."""
        data = request.data if isinstance(request.data, dict) else {}
        org = _ensure_org(request)
        new_positions = data.get('new_positions') or []

        normalized_positions, error = _normalize_reorder_payload(Product, org, new_positions, 'Ürün')
        if error is not None:
            return error

        with transaction.atomic():
            products = []
            for product, order in normalized_positions:
                product.order = order
                products.append(product)
            Product.objects.bulk_update(products, ['order'])

        return Response({'detail': 'Ürünler başarıyla sıralandı.', 'count': len(normalized_positions)})


class CategoryViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'products.view'
    permission_map = {
        'create': 'products.create',
        'update': 'products.edit',
        'partial_update': 'products.edit',
        'destroy': 'products.delete',
        'reorder': 'products.edit',
    }
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['order', 'created_at', 'name']
    ordering = ['order', 'id']  # Default ordering
    queryset = Category.objects.all()

    @action(detail=False, methods=['post'], url_path='reorder')
    def reorder(self, request):
        """Reorder categories by updating their order field. Accepts list of {id, order} tuples."""
        data = request.data if isinstance(request.data, dict) else {}
        org = _ensure_org(request)
        new_positions = data.get('new_positions') or []

        normalized_positions, error = _normalize_reorder_payload(Category, org, new_positions, 'Kategori')
        if error is not None:
            return error

        with transaction.atomic():
            categories = []
            for category, order in normalized_positions:
                category.order = order
                categories.append(category)
            Category.objects.bulk_update(categories, ['order'])

        return Response({'detail': 'Kategoriler başarıyla sıralandı.', 'count': len(normalized_positions)})


class SalesOrderViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = SalesOrderSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'orders.view'
    permission_map = {
        'create': 'orders.create',
        'update': 'orders.edit',
        'partial_update': 'orders.edit',
        'destroy': 'orders.delete',
    }
    queryset = SalesOrder.objects.all()

    def create(self, request, *args, **kwargs):
        org = _ensure_org(request)
        number_range, _ = NumberRange.objects.get_or_create(organization=org, doc_type='SO', defaults={'prefix': 'SO-'})
        number = number_range.next_number()
        data = request.data.copy()
        data.setdefault('organization', org.id if hasattr(org, 'id') else org)
        data.setdefault('number', number)
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)


class PurchaseOrderViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = PurchaseOrderSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'orders.view'
    permission_map = {
        'create': 'orders.create',
        'update': 'orders.edit',
        'partial_update': 'orders.edit',
        'destroy': 'orders.delete',
    }
    queryset = PurchaseOrder.objects.all()


class StockMovementViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = StockMovementSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'warehouse_movements.view'
    permission_map = {
        'create': 'inventory.edit',
        'update': 'inventory.edit',
        'partial_update': 'inventory.edit',
        'destroy': 'inventory.edit',
    }
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'movement_type']
    ordering_fields = ['created_at', 'quantity']
    queryset = StockMovement.objects.all()

    def create(self, request, *args, **kwargs):
        return Response({'detail': 'Stok hareketleri Depo ekranındaki işlem uçlarından oluşturulmalıdır.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def update(self, request, *args, **kwargs):
        return Response({'detail': 'Stok hareketleri değiştirilemez.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    partial_update = update
    destroy = update


class WarehouseViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = WarehouseSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'warehouses.view'
    permission_map = {'create': 'warehouses.manage', 'update': 'warehouses.manage', 'partial_update': 'warehouses.manage', 'destroy': 'warehouses.manage'}
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code', 'name', 'city', 'responsible', 'branch__name']
    ordering_fields = ['code', 'name']
    queryset = Warehouse.objects.all()

    def get_queryset(self):
        return super().get_queryset().select_related('branch').annotate(
            stock_total=Sum('stocks__quantity'),
            location_count=Count('inventorylocation', distinct=True),
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.stocks.filter(quantity__gt=0).exists():
            return Response({'detail': 'Bakiyesi bulunan depo arşivlenemez.'}, status=status.HTTP_400_BAD_REQUEST)
        instance.is_active = False
        instance.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class InventoryLocationViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = InventoryLocationSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'warehouses.view'
    permission_map = {'create': 'warehouse_locations.manage', 'update': 'warehouse_locations.manage', 'partial_update': 'warehouse_locations.manage', 'destroy': 'warehouse_locations.manage'}
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code', 'name', 'warehouse__code', 'warehouse__name']
    ordering_fields = ['warehouse__code', 'code']
    queryset = InventoryLocation.objects.all()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.stocks.filter(quantity__gt=0).exists():
            return Response({'detail': 'Bakiyesi bulunan raf arşivlenemez.'}, status=status.HTTP_400_BAD_REQUEST)
        instance.is_active = False
        instance.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class WarehouseStockViewSet(OrgScopedMixin, viewsets.ReadOnlyModelViewSet):
    serializer_class = WarehouseStockSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'warehouse_stock.view'
    permission_map = {
        'stock_in': 'warehouse_stock.operate', 'stock_out': 'warehouse_stock.operate', 'adjust': 'warehouse_stock.operate',
        'transfer': 'warehouse_stock.transfer', 'allocate_opening_balance': 'warehouse_stock.allocate',
        'import_count': 'warehouse_stock.import', 'export': 'warehouse_stock.export',
    }
    queryset = WarehouseStock.objects.all()

    def get_queryset(self):
        qs = super().get_queryset().select_related('warehouse', 'location', 'product', 'product__category')
        q = self.request.query_params.get('q', '').strip()
        if q:
            qs = qs.filter(Q(product__sku__icontains=q) | Q(product__name__icontains=q) | Q(product__category__name__icontains=q) |
                           Q(detail_1_override__icontains=q) | Q(detail_2_override__icontains=q))
        for param, field in [('warehouse', 'warehouse_id'), ('location', 'location_id'), ('product', 'product_id')]:
            value = self.request.query_params.get(param)
            if value:
                qs = qs.filter(**{field: value})
        return qs

    def _objects(self, request, *, transfer_mode=False):
        try:
            org = _ensure_org(request)
            product = Product.objects.get(pk=request.data.get('product_id'), organization=org)
            location = InventoryLocation.objects.select_related('warehouse').get(pk=request.data.get('location_id'), organization=org)
            if not transfer_mode:
                return (org, product, location), None
            target = InventoryLocation.objects.select_related('warehouse').get(pk=request.data.get('target_location_id'), organization=org)
            return (org, product, location, target), None
        except (Product.DoesNotExist, InventoryLocation.DoesNotExist, TypeError, ValueError):
            return None, Response({'detail': 'Ürün veya raf seçimi geçersiz.'}, status=status.HTTP_400_BAD_REQUEST)

    def _run(self, request, service, **kwargs):
        try:
            movement = service(user=request.user, **kwargs)
        except (InventoryError, Product.DoesNotExist, InventoryLocation.DoesNotExist) as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        push_event({'type': 'inventory.changed', 'product_id': movement.product_id})
        return Response(StockMovementSerializer(movement, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='stock-in')
    def stock_in(self, request):
        objects, error = self._objects(request)
        if error: return error
        org, product, location = objects
        return self._run(request, stock_in, organization=org, product=product, location=location, quantity=request.data.get('quantity'),
                         reference=request.data.get('reference', ''), note=request.data.get('note', ''))

    @action(detail=False, methods=['post'], url_path='stock-out')
    def stock_out(self, request):
        objects, error = self._objects(request)
        if error: return error
        org, product, location = objects
        return self._run(request, stock_out, organization=org, product=product, location=location, quantity=request.data.get('quantity'),
                         reference=request.data.get('reference', ''), note=request.data.get('note', ''))

    @action(detail=False, methods=['post'])
    def adjust(self, request):
        objects, error = self._objects(request)
        if error: return error
        org, product, location = objects
        return self._run(request, adjust, organization=org, product=product, location=location, target_quantity=request.data.get('quantity'),
                         reference=request.data.get('reference', ''), note=request.data.get('note', ''),
                         detail_1_override=request.data.get('detail_1_override'), detail_2_override=request.data.get('detail_2_override'))

    @action(detail=False, methods=['post'])
    def transfer(self, request):
        objects, error = self._objects(request, transfer_mode=True)
        if error: return error
        org, product, location, target = objects
        return self._run(request, transfer, organization=org, product=product, location_from=location, location_to=target,
                         quantity=request.data.get('quantity'), reference=request.data.get('reference', ''), note=request.data.get('note', ''))

    @action(detail=False, methods=['post'], url_path='allocate-opening-balance')
    def allocate_opening_balance(self, request):
        org = _ensure_org(request)
        try:
            product = Product.objects.get(pk=request.data.get('product_id'), organization=org)
            allocate_opening_balance(organization=org, product=product, allocations=request.data.get('allocations') or [], user=request.user)
        except (InventoryError, Product.DoesNotExist, InventoryLocation.DoesNotExist) as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        push_event({'type': 'inventory.changed', 'product_id': product.id})
        return Response({'detail': 'Açılış bakiyesi depolara aktarıldı.'})

    @action(detail=False, methods=['get'])
    def export(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Depo Sayımı'
        sheet.append(['Depo Kodu', 'Raf Kodu', 'Ürün Kodu', 'Ürün Adı', 'Detay-1', 'Detay-2', 'Hedef Miktar'])
        for row in self.get_queryset():
            data = self.get_serializer(row).data
            sheet.append([data['warehouse_code'], data['location_code'], data['product_sku'], data['product_name'], data['detail_1'], data['detail_2'], float(row.quantity)])
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="depo-sayim-sablonu.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import-count')
    def import_count(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Excel dosyası zorunludur.'}, status=status.HTTP_400_BAD_REQUEST)
        org = _ensure_org(request)
        try:
            rows = list(load_workbook(file, data_only=True).active.iter_rows(min_row=2, values_only=True))
            changed = 0
            for warehouse_code, location_code, sku, _name, detail_1, detail_2, target in rows:
                if not sku or target in [None, '']:
                    continue
                location = InventoryLocation.objects.select_related('warehouse').get(organization=org, warehouse__code=str(warehouse_code).strip(), code=str(location_code).strip())
                product = Product.objects.get(organization=org, sku=str(sku).strip())
                stock = WarehouseStock.objects.filter(organization=org, location=location, product=product).first()
                if not stock or stock.quantity != as_decimal(target, 'Hedef miktar') or stock.detail_1_override != str(detail_1 or '') or stock.detail_2_override != str(detail_2 or ''):
                    adjust(organization=org, product=product, location=location, target_quantity=target, user=request.user, reference='EXCEL SAYIM',
                           note='Depo sayım Excel aktarımı', source_type='excel_count', detail_1_override=detail_1, detail_2_override=detail_2)
                    changed += 1
        except Exception as exc:
            return Response({'detail': f'Excel işlenemedi: {exc}'}, status=status.HTTP_400_BAD_REQUEST)
        push_event({'type': 'inventory.changed'})
        return Response({'changed_rows': changed})


class WarehouseDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsOrgMember]

    def get(self, request):
        from accounts.utils import user_has_perm
        if not user_has_perm(request.user, 'warehouses.view'):
            return Response({'detail': 'Bu işlem için yetkiniz yok.'}, status=status.HTTP_403_FORBIDDEN)
        org = _ensure_org(request)
        stocks = WarehouseStock.objects.filter(organization=org)
        movements = StockMovement.objects.filter(organization=org, created_at__gte=timezone.now() - timedelta(days=30))
        warehouse_id = request.query_params.get('warehouse')
        if warehouse_id:
            stocks = stocks.filter(warehouse_id=warehouse_id)
            movements = movements.filter(Q(warehouse_from_id=warehouse_id) | Q(warehouse_to_id=warehouse_id))
        return Response({
            'active_warehouses': Warehouse.objects.filter(organization=org, is_active=True).count(),
            'total_stock': stocks.aggregate(value=Sum('quantity'))['value'] or 0,
            'critical_stock': Product.objects.filter(organization=org, stock__lt=F('reorder_point')).count(),
            'recent_movements': movements.count(),
            'by_warehouse': list(stocks.values(name=F('warehouse__name')).annotate(value=Sum('quantity')).order_by('-value')),
            'movement_trend': list(movements.annotate(day=TruncDate('created_at')).values('day', 'movement_type').annotate(value=Sum('quantity')).order_by('day')),
            'top_products': list(stocks.values(name=F('product__name')).annotate(value=Sum('quantity')).order_by('-value')[:8]),
        })


class VehicleViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'logistics.view'
    permission_map = {
        'create': 'logistics.edit',
        'update': 'logistics.edit',
        'partial_update': 'logistics.edit',
        'destroy': 'logistics.edit',
    }
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['plate', 'driver', 'status']
    ordering_fields = ['last_update', 'distance_today']
    queryset = Vehicle.objects.all()
