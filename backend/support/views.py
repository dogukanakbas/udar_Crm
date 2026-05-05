from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from core.events import push_event
from rest_framework.exceptions import PermissionDenied
from django.db.models import Q, Exists, OuterRef, Max
import re
import os
import uuid
from permissions import IsOrgMember, HasAPIPermission, CommentOnlyRestriction, ViewOnlyRestriction
from audit.utils import log_entity_action
from .models import Ticket, TicketMessage, Task, TaskAttachment, TaskComment, TaskChecklist, TaskTimeEntry, TaskModel, TaskProductionEntry
from accounts.models import User, Team, TeamAssociate
from .models_automation import AutomationRule
from .utils import send_slack_webhook, send_email, generate_presigned_post, scan_file_with_clamav
from rest_framework import status
from .serializers import (
    TicketSerializer,
    TicketMessageSerializer,
    TaskSerializer,
    TaskAttachmentSerializer,
    TaskCommentSerializer,
    TaskChecklistSerializer,
    TaskModelSerializer,
    AutomationRuleSerializer,
    TaskTimeEntrySerializer,
)
from rest_framework.views import APIView
from rest_framework.response import Response
from django.utils import timezone
from datetime import datetime as dt_datetime
from openpyxl import load_workbook

from .workflow_utils import (
    apply_product_line_to_task,
    ensure_product_line_workflows,
    ensure_workflow_state,
    format_shortfall_reason_for_storage,
    workflow_team_id_list,
    parallel_queue_visible,
    workflow_stage_production_met,
    resolve_production_gate,
)
from .checklist_sync import sync_workflow_checklist


def assign_task_to_team_leader(task, team):
    """Ekibe düşen görev ekip havuzunda kalır (tek kullanıcıya otomatik atanmaz)."""
    task.assignee = None


def user_may_claim_task_as_leader(user, team):
    """Normal/sıralı akışta görev üstlenme: yalnızca ekip lideri (Admin/Manager dahil)."""
    if not team:
        return False
    role = getattr(user, 'role', '')
    if role in ('Admin', 'Manager'):
        return True
    lid = getattr(team, 'leader_id', None)
    return bool(lid and user.id == lid)


def user_can_see_claim_queue_entry(task, user):
    """my-team-queue: bu kullanıcı görevi üstlenebilecekleri listesinde görmeli mi?"""
    role = getattr(user, 'role', '')
    staff = role in ('Admin', 'Manager')
    member_team_ids = set(user.teams.values_list('id', flat=True))
    leader_team_ids = set(
        Team.objects.filter(organization_id=task.organization_id, leader_id=user.id).values_list('id', flat=True)
    )
    user_team_ids = member_team_ids | leader_team_ids
    wf = workflow_team_id_list(task)
    if getattr(task, 'workflow_parallel', False) and wf:
        if staff:
            return True
        return parallel_queue_visible(task, user, list(user_team_ids))
    # Kalem-bazlı bağımsız akış: kullanıcının üyesi olduğu ekipte açık kalem adımı varsa görünür.
    lines = list(getattr(task, 'product_lines', None) or [])
    if lines:
        for ln in lines:
            ids = [int(x) for x in (ln or {}).get('workflow_team_ids', []) if str(x).isdigit()]
            if not ids:
                continue
            state = dict((ln or {}).get('workflow_stage_state') or {})
            for tid in ids:
                if not staff and tid not in user_team_ids:
                    continue
                st = dict(state.get(str(tid)) or {})
                if not st.get('stage_done'):
                    return True
    if task.assignee_id and not staff:
        return False
    if task.assignee_id and staff:
        return True
    team = task.current_team or task.team
    return user_may_claim_task_as_leader(user, team)


class OrgScopedMixin:
    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(organization=org)
        return qs

    def perform_create(self, serializer):
        org = getattr(self.request.user, 'organization', None)
        serializer.save(organization=org)


FACTORY_CHECKLIST = [
    "PVC Dilimleme",
    "MDF Dilimleme",
    "CNC",
    "Laminasyon + Laminasyon Vakum",
    "Kanat Açma",
    "Kasa Hattı",
    "Pervaz Hattı",
    "Paketleme",
]

SIZE_TOKEN_RE = re.compile(r'\d+(?:[.,]\d+)?(?:\s*[*xX]\s*\d+(?:[.,]\d+)?){1,2}')


def _norm_text(v):
    return str(v or '').strip()


def _parse_int_qty(v):
    s = _norm_text(v).replace(',', '.')
    if not s:
        return 1
    try:
        n = float(s)
    except (TypeError, ValueError):
        return 1
    if n <= 0:
        return 1
    return int(round(n)) if n >= 1 else 1


def _extract_sizes(*texts):
    out = []
    seen = set()
    merged = " / ".join([_norm_text(t) for t in texts if _norm_text(t)])
    for m in SIZE_TOKEN_RE.finditer(merged):
        raw = m.group(0)
        norm = re.sub(r'\s+', '', raw).replace('X', '*').replace('x', '*')
        if norm and norm not in seen:
            seen.add(norm)
            out.append(norm)
    return out


def _extract_color(*texts):
    merged = " / ".join([_norm_text(t) for t in texts if _norm_text(t)])
    if not merged:
        return ''
    segs = [s.strip() for s in merged.split('/') if s and s.strip()]
    # Sondan başlayarak ölçü/seri olmayan ve harf içeren ilk parçayı renk kabul et.
    for seg in reversed(segs):
        has_letter = bool(re.search(r'[A-Za-zÇĞİÖŞÜçğıöşü]', seg))
        looks_code = bool(re.fullmatch(r'[A-Za-z]-?\d+(?:[.,]\d+)?', seg))
        looks_size = bool(SIZE_TOKEN_RE.search(seg))
        if has_letter and not looks_code and not looks_size:
            return seg
    return ''


def _map_excel_status(raw: str) -> str:
    s = _norm_text(raw).lower()
    if s in ('aktif', 'active', 'in-progress', 'in progress', 'devam ediyor'):
        return 'in-progress'
    if s in ('kapanmış', 'kapandi', 'kapandı', 'closed', 'done', 'tamamlandı'):
        return 'done'
    return 'todo'


def _task_has_fire_record(task):
    """Eksik üretim kapanışı için en az bir kalemde fire adedi+sebebi girilmiş olmalı."""
    lines = list(getattr(task, 'product_lines', None) or [])
    if not lines:
        return False
    for ln in lines:
        row = ln or {}
        try:
            fq = float(row.get('fire_qty') or 0)
        except (TypeError, ValueError):
            fq = 0.0
        fr = str(row.get('fire_reason') or '').strip()
        if fq > 0 and fr:
            return True
    return False

class TicketViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TicketSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, CommentOnlyRestriction, HasAPIPermission]
    required_perm = 'tickets.view'
    permission_map = {
        'create': 'tickets.edit',
        'update': 'tickets.edit',
        'partial_update': 'tickets.edit',
        'destroy': 'tickets.edit',
    }
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['subject', 'company_name', 'status']
    ordering_fields = ['updated_at', 'priority']
    queryset = Ticket.objects.all()

    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization, assignee=self.request.user)


class TicketMessageViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TicketMessageSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, CommentOnlyRestriction, HasAPIPermission]
    required_perm = 'tickets.view'
    queryset = TicketMessage.objects.all()

    def perform_create(self, serializer):
        message = serializer.save(author=self.request.user)
        # Mention notification: @username in ticket message
        mentions = set(re.findall(r'@([\w\.-]+)', message.message or ''))
        if mentions:
            users = User.objects.filter(username__in=list(mentions), organization=message.ticket.organization)
            if users:
                TaskComment.objects.create(
                    task=None,
                    author=None,
                    type='activity',
                    text=f"Mention (ticket): {', '.join([u.username for u in users])}",
                )
            for u in users:
                push_event(
                    {
                        "type": "notification.mention",
                        "ticket_id": message.ticket_id,
                        "ticket_subject": message.ticket.subject,
                        "message_id": message.id,
                        "mentioned": u.username,
                        "organization": message.ticket.organization_id,
                    }
                )
                if u.email:
                    send_email(
                        u.email,
                        f"Ticket mention: {message.ticket.subject}",
                        f"{self.request.user.username} sizi bir ticket mesajında bahsetti: {message.message}",
                    )


class TaskViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, CommentOnlyRestriction, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
        'destroy': 'tasks.edit',
        'import_excel': 'tasks.edit',
    }
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'status', 'priority']
    ordering_fields = ['due', 'priority', 'updated_at']

    def _backfill_line_workflows(self, tasks):
        for t in tasks:
            if ensure_product_line_workflows(t):
                t.save(
                    update_fields=[
                        'product_lines',
                        'workflow_team_ids',
                        'workflow_parallel',
                        'current_team',
                        'team',
                        'updated_at',
                    ]
                )

    def get_queryset(self):
        qs = Task.objects.all()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(organization=org)
        role = getattr(self.request.user, 'role', '')
        user = self.request.user
        # Worker: tüm bölümler/kullanıcılar yeni görevleri görsün (org genelinde)
        if role == 'Worker':
            pass  # Org içindeki tüm görevleri göster
        elif role not in ['Admin', 'Manager']:
            qs = qs.filter(Q(owner=user) | Q(assignee=user) | Q(team__members=user)).distinct()
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        self._backfill_line_workflows(queryset[:300])
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        self._backfill_line_workflows([instance])
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def perform_create(self, serializer):
        if getattr(self.request.user, 'role', '') == 'Worker':
            raise PermissionDenied("Worker rolündeki kullanıcılar görev atayamaz")
        instance = serializer.save(organization=self.request.user.organization, owner=self.request.user)
        wf_list_new = workflow_team_id_list(instance)
        if wf_list_new and not getattr(instance, 'workflow_parallel', False):
            # Sıralı akış: ilk aşamada usta başı «Üstlen» ile kabul etsin; formda atanan olsa bile havuza al
            instance.assignee = None
            instance.save(update_fields=['assignee', 'updated_at'])
            TaskComment.objects.create(
                task=instance,
                author=self.request.user,
                type='activity',
                text="Sıralı iş akışı — ilk aşama ekip usta başı görevi «Üstlen» ile başlatmalıdır.",
            )
        elif not instance.assignee_id:
            assign_task_to_team_leader(instance, instance.current_team or instance.team)
            if instance.assignee_id:
                instance.save(update_fields=['assignee'])
                instance.refresh_from_db()
        if getattr(instance, 'workflow_parallel', False) and instance.assignee_id and workflow_team_id_list(instance):
            ensure_workflow_state(instance)
            st = dict(instance.workflow_stage_state or {})
            ct = instance.current_team_id or instance.team_id
            if ct and str(ct) in st:
                st[str(ct)]['assignee_id'] = instance.assignee_id
                instance.workflow_stage_state = st
                instance.save(update_fields=['workflow_stage_state'])
        # İş akışı varsa checklist ekip adımlarına göre; yoksa sabit modda fabrika checklist
        if workflow_team_id_list(instance):
            sync_workflow_checklist(instance)
        elif getattr(instance, 'mode', 'manual') == 'fixed':
            TaskChecklist.objects.bulk_create(
                [
                    TaskChecklist(task=instance, title=title, order=idx)
                    for idx, title in enumerate(FACTORY_CHECKLIST)
                ]
            )
        log_entity_action(instance, 'created', user=self.request.user)
        push_event(
            {
                "type": "task.created",
                "task_id": instance.id,
                "organization": instance.organization_id,
                "title": instance.title,
                "assignee": getattr(instance.assignee, "email", None),
                "assignee_id": instance.assignee_id,
            }
        )
        # Assignee bildirimi ve aktivite kaydı
        if instance.assignee:
            TaskComment.objects.create(
                task=instance,
                author=self.request.user,
                type='activity',
                text=f"Görev {instance.assignee.username} kullanıcısına atandı",
            )
            if instance.assignee.email:
                send_email(
                    instance.assignee.email,
                    f"Görev ataması: {instance.title}",
                    f"Size yeni bir görev atandı: {instance.title}",
                )
        # Automation: if due is soon (handled by scheduled tasks); status change not relevant on create

    def perform_update(self, serializer):
        original = Task.objects.get(pk=serializer.instance.pk)
        user_role = getattr(self.request.user, 'role', '')
        # Sadece sahip/atanan/ekip değişimini kısıtla; diğer alanlar (durum, öncelik, tarihler) Worker için serbest
        restricted_fields = ['owner', 'assignee', 'team']
        if user_role not in ['Admin', 'Manager']:
            # Kısıtlı alanları reddetmek yerine sessizce yok say
            for f in restricted_fields:
                serializer.validated_data.pop(f, None)
        instance = serializer.save()
        if workflow_team_id_list(instance):
            sync_workflow_checklist(instance)
        # Ekip değişti ve atanmış kişi gönderilmediyse usta başına yönlendir
        if (
            'assignee' not in serializer.validated_data
            and ('current_team' in serializer.validated_data or 'team' in serializer.validated_data)
        ):
            assign_task_to_team_leader(instance, instance.current_team or instance.team)
            instance.save(update_fields=['assignee'])
        # Assignee değiştiyse bildirim ve aktivite
        if 'assignee' in serializer.validated_data and instance.assignee != original.assignee:
            if instance.assignee:
                TaskComment.objects.create(
                    task=instance,
                    author=self.request.user,
                    type='activity',
                    text=f"Görev {instance.assignee.username} kullanıcısına yeniden atandı",
                )
                if instance.assignee.email:
                    send_email(
                        instance.assignee.email,
                        f"Görev ataması: {instance.title}",
                        f"Görev size yeniden atandı: {instance.title}",
                    )
        # activity log on status change
        if 'status' in serializer.validated_data:
            TaskComment.objects.create(
                task=instance,
                author=self.request.user,
                type='activity',
                text=f"Durum: {original.status} -> {serializer.validated_data.get('status')}",
            )
            # Automation: task_status_changed
            self.run_automations(instance, 'task_status_changed', extra={'from': original.status, 'to': serializer.validated_data.get('status')})
            push_event(
                {
                    "type": "task.status",
                    "task_id": instance.id,
                    "organization": instance.organization_id,
                    "title": instance.title,
                    "status": serializer.validated_data.get('status'),
                    "assignee_id": instance.assignee_id,
                    "by": getattr(self.request.user, "email", None),
                }
            )

        # activity log on assignee / due / priority changes
        changes = []
        if 'assignee' in serializer.validated_data and instance.assignee_id != original.assignee_id:
            changes.append(f"Atanan: {getattr(original.assignee, 'email', '—')} -> {getattr(instance.assignee, 'email', '—')}")
        if 'due' in serializer.validated_data and instance.due != original.due:
            changes.append(f"Vade: {original.due} -> {instance.due}")
        if 'priority' in serializer.validated_data and instance.priority != original.priority:
            changes.append(f"Öncelik: {original.priority} -> {instance.priority}")
        if 'tags' in serializer.validated_data and instance.tags != original.tags:
            changes.append(f"Etiketler: {original.tags or []} -> {instance.tags or []}")
        if 'status' in serializer.validated_data and original.status != serializer.validated_data.get('status'):
            changes.append(f"Durum: {original.status} -> {serializer.validated_data.get('status')}")
        for change in changes:
            TaskComment.objects.create(task=instance, author=self.request.user, type='activity', text=change)
        log_entity_action(instance, 'updated', user=self.request.user)
        push_event(
            {
                "type": "task.updated",
                "task_id": instance.id,
                "organization": instance.organization_id,
                "title": instance.title,
                "assignee_id": instance.assignee_id,
                "by": getattr(self.request.user, "email", None),
                "changes": changes,
            }
        )

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        Excel'den tek görev + çoklu ürün kalemi taslağı üretir.
        Her satır: 1 ürün kalemi.
        """
        if getattr(request.user, 'role', '') == 'Worker':
            raise PermissionDenied("Worker rolündeki kullanıcılar görev içe aktaramaz")
        up = request.FILES.get('file')
        if not up:
            return Response({'detail': 'file gerekli'}, status=status.HTTP_400_BAD_REQUEST)
        fname = _norm_text(getattr(up, 'name', ''))
        if not fname.lower().endswith('.xlsx'):
            return Response({'detail': 'Yalnızca .xlsx kabul edilir'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(up, data_only=True, read_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Excel okunamadı'}, status=status.HTTP_400_BAD_REQUEST)

        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        headers_row = next(rows, None)
        headers = [_norm_text(h) for h in (headers_row or [])]
        idx = {h: i for i, h in enumerate(headers) if h}
        required = [
            'ÜRETİLECEK ÜRÜN İSMİ',
            'KATEGORİ KODU',
            'SİPARİŞ SERİ',
            'SİPARİŞ SIRA',
            'ÜRETİM MİKTARI',
            'SİPARİŞ AÇIKLAMA_1',
            'SİPARİŞ AÇIKLAMA_2',
        ]
        missing = [k for k in required if k not in idx]
        if missing:
            return Response(
                {'detail': f'Eksik başlık(lar): {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        product_lines = []
        refs = []
        parsed_statuses = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            prod_name = _norm_text(r[idx['ÜRETİLECEK ÜRÜN İSMİ']]) if idx['ÜRETİLECEK ÜRÜN İSMİ'] < len(r) else ''
            qty_raw = r[idx['ÜRETİM MİKTARI']] if idx['ÜRETİM MİKTARI'] < len(r) else None
            if not prod_name and qty_raw in (None, ''):
                continue
            cat_code = _norm_text(r[idx['KATEGORİ KODU']]) if idx['KATEGORİ KODU'] < len(r) else ''
            s_seri = _norm_text(r[idx['SİPARİŞ SERİ']]) if idx['SİPARİŞ SERİ'] < len(r) else ''
            s_sira = _norm_text(r[idx['SİPARİŞ SIRA']]) if idx['SİPARİŞ SIRA'] < len(r) else ''
            ac1 = _norm_text(r[idx['SİPARİŞ AÇIKLAMA_1']]) if idx['SİPARİŞ AÇIKLAMA_1'] < len(r) else ''
            ac2 = _norm_text(r[idx['SİPARİŞ AÇIKLAMA_2']]) if idx['SİPARİŞ AÇIKLAMA_2'] < len(r) else ''
            st_raw = _norm_text(r[idx['DURUMU']]) if 'DURUMU' in idx and idx['DURUMU'] < len(r) else ''
            sizes = _extract_sizes(ac1, ac2)
            color = _extract_color(ac1, ac2)
            qty = _parse_int_qty(qty_raw)
            unit_type = 'metre' if 'pvc' in prod_name.lower() else 'adet'
            brief = " / ".join([x for x in [prod_name, ac1, ac2] if x]).strip()[:600]
            if s_seri or s_sira:
                refs.append(f'{s_seri}-{s_sira}'.strip('-'))
            if st_raw:
                parsed_statuses.append(_map_excel_status(st_raw))
            product_lines.append(
                {
                    'mode': 'manual',
                    'unit_type': unit_type,
                    'model_code': cat_code,
                    'variant': '',
                    'quantity': qty,
                    'model_duration_minutes': 0,
                    'total_planned_minutes': 0,
                    'model_blade_depth': '',
                    'model_sizes': sizes,
                    'product_color': color,
                    'product_color_code': '',
                    'brief_intro': brief,
                    'fire_qty': 0,
                    'fire_reason': '',
                    'fire_image_data_url': '',
                    'qty_produced': 0,
                }
            )

        if not product_lines:
            return Response({'detail': 'Excel içinde içe aktarılacak satır bulunamadı'}, status=status.HTTP_400_BAD_REQUEST)

        uniq_refs = [x for i, x in enumerate(refs) if x and x not in refs[:i]]
        ref_text = ", ".join(uniq_refs[:5]) if uniq_refs else fname
        title = f'Excel Sipariş: {ref_text} ({len(product_lines)} kalem)'
        qty_total = max(1, sum(int(pl.get('quantity') or 1) for pl in product_lines))
        # Durumu kolonundan görev durumunu türet:
        # herhangi bir aktif varsa in-progress; hepsi done ise done; aksi todo
        derived_status = 'todo'
        if parsed_statuses:
            if any(x == 'in-progress' for x in parsed_statuses):
                derived_status = 'in-progress'
            elif all(x == 'done' for x in parsed_statuses):
                derived_status = 'done'
        return Response(
            {
                'draft': {
                    'title': title,
                    'status': derived_status,
                    'priority': 'medium',
                    'planned_hours': 0,
                    'planned_cost': 0,
                    'workflow_team_ids': [],
                    'workflow_parallel': False,
                    'workflow_stage_targets': [],
                    'quantity': qty_total,
                    'tags': [f'excel-import', f'kalem:{len(product_lines)}'] + ([f'ref:{uniq_refs[0]}'] if uniq_refs else []),
                    'product_lines': product_lines,
                },
                'imported_lines': len(product_lines),
                'refs': uniq_refs,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'])
    def claim(self, request, pk=None):
        """Worker ekibindeki bekleyen görevi üstlenir. Paralel akışta workflow içindeki uygun bölümde üstlenir."""
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        if getattr(task, 'workflow_parallel', False) and workflow_team_id_list(task):
            wf = workflow_team_id_list(task)
            user_teams = list(request.user.teams.values_list('id', flat=True))
            user_set = set(user_teams)
            ensure_workflow_state(task)
            state = dict(task.workflow_stage_state or {})
            claimed_tid = None
            for tid in wf:
                if tid not in user_set:
                    continue
                sec_team = Team.objects.filter(id=tid, organization_id=task.organization_id).first()
                # Paralel akışta sadece lider değil, bölüm üyesi olan herkes üstlenebilir.
                if not sec_team:
                    continue
                is_member = sec_team.members.filter(id=request.user.id).exists()
                is_leader = bool(getattr(sec_team, 'leader_id', None) and str(sec_team.leader_id) == str(request.user.id))
                if not (is_member or is_leader):
                    continue
                st = dict(state.get(str(tid), {}))
                if st.get('stage_done'):
                    continue
                aid = st.get('assignee_id')
                if aid == request.user.id:
                    return Response(TaskSerializer(task, context={'request': request}).data)
                if aid and aid != request.user.id:
                    continue
                st['assignee_id'] = request.user.id
                state[str(tid)] = st
                task.workflow_stage_state = state
                task.assignee = request.user
                task.current_team_id = tid
                task.save(update_fields=['workflow_stage_state', 'assignee', 'current_team', 'updated_at'])
                claimed_tid = tid
                break
            if not claimed_tid:
                raise PermissionDenied("Bu görevde üstlenebileceğiniz açık bölüm yok")
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=f"{request.user.username} görevi ({claimed_tid} bölümü) üstlendi",
            )
            push_event(
                {
                    "type": "task.claimed",
                    "task_id": task.id,
                    "organization": task.organization_id,
                    "by": getattr(request.user, "email", None),
                }
            )
            return Response(TaskSerializer(task, context={'request': request}).data)

        current_team = task.current_team or task.team
        if not current_team:
            raise PermissionDenied("Görev henüz bir ekibe atanmamış")
        if not user_may_claim_task_as_leader(request.user, current_team):
            raise PermissionDenied("Görevi yalnızca ilgili ekip usta başısı veya yönetici üstlenebilir.")
        if task.assignee_id == request.user.id:
            return Response(TaskSerializer(task, context={'request': request}).data)
        if task.assignee_id and task.assignee_id != request.user.id:
            raise PermissionDenied(
                "Bu görev başka bir çalışan tarafından üstlenilmiş. Bölüm tamamlanana veya usta başı görevi "
                '"Ekibe aç" ile havuza alana kadar tekrar üstlenemez.'
            )
        task.assignee = request.user
        task.save(update_fields=['assignee', 'updated_at'])
        TaskComment.objects.create(
            task=task,
            author=request.user,
            type='activity',
            text=f"{request.user.username} görevi üstlendi",
        )
        push_event(
            {
                "type": "task.claimed",
                "task_id": task.id,
                "organization": task.organization_id,
                "by": getattr(request.user, "email", None),
            }
        )
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='release-to-team')
    def release_to_team(self, request, pk=None):
        """Usta başı: görevi ekibe açar (assignee temizlenir, üyeler üstlenebilir)."""
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        team = task.current_team or task.team
        if not team:
            return Response({"detail": "Görevde ekip yok"}, status=status.HTTP_400_BAD_REQUEST)
        leader_id = getattr(team, 'leader_id', None)
        role = getattr(request.user, 'role', '')
        if role not in ('Admin', 'Manager'):
            if not leader_id or task.assignee_id != leader_id:
                raise PermissionDenied("Sadece usta başı veya yönetici görevi ekibe açabilir")
        task.assignee = None
        task.save(update_fields=['assignee', 'updated_at'])
        TaskComment.objects.create(
            task=task,
            author=request.user,
            type='activity',
            text=f"{request.user.username} görevi ekip havuzuna açtı (üyeler üstlenebilir)",
        )
        push_event(
            {
                "type": "task.released_to_team",
                "task_id": task.id,
                "organization": task.organization_id,
                "assignee_id": None,
                "by": getattr(request.user, "email", None),
            }
        )
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def handover(self, request, pk=None):
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        target_team_id = request.data.get('team')
        target_user_id = request.data.get('assignee')
        note = request.data.get('note', '')
        handover_type = request.data.get('type', 'manual')
        target_team = None
        if target_team_id:
            target_team = Team.objects.filter(id=target_team_id, organization=request.user.organization).first()
        target_user = None
        if target_user_id:
            target_user = User.objects.filter(id=target_user_id, organization=request.user.organization).first()
        
        from_team = task.current_team or task.team
        history = task.handover_history or []
        history.append(
            {
                "from_team": getattr(from_team, 'id', None),
                "from_team_name": getattr(from_team, 'name', None),
                "to_team": getattr(target_team, 'id', None),
                "to_team_name": getattr(target_team, 'name', None),
                "by": request.user.username,
                "note": note,
                "type": handover_type,
                "at": timezone.now().isoformat(),
            }
        )
        task.current_team = target_team or task.current_team or task.team
        if target_user:
            task.assignee = target_user
        else:
            assign_task_to_team_leader(task, task.current_team)
        task.handover_reason = note
        task.handover_at = timezone.now()
        task.handover_history = history
        task.save(update_fields=['current_team', 'assignee', 'handover_reason', 'handover_at', 'handover_history', 'updated_at'])
        TaskComment.objects.create(
            task=task,
            author=request.user,
            type='activity',
            text=f"Devir: {getattr(from_team, 'name', '—')} → {getattr(target_team, 'name', '—')} (not: {note})",
        )
        push_event(
            {
                "type": "task.handover",
                "task_id": task.id,
                "organization": task.organization_id,
                "title": task.title,
                "assignee_id": task.assignee_id,
                "to_team": getattr(target_team, 'id', None),
                "by": getattr(request.user, "email", None),
            }
        )
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def self_handover(self, request, pk=None):
        """Worker kendi görevini başka ekibe devredebilir (bölüm değişimi için)"""
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        if task.assignee_id != request.user.id:
            raise PermissionDenied("Sadece size atanan görevi devredebilirsiniz")
        
        target_team_id = request.data.get('team')
        reason = request.data.get('reason', 'Bölüm değişimi - başka alanda çalışıyorum')
        
        if not target_team_id:
            raise PermissionDenied("Hedef ekip belirtilmeli")
        
        target_team = Team.objects.filter(id=target_team_id, organization=request.user.organization).first()
        if not target_team:
            raise PermissionDenied("Hedef ekip bulunamadı")
        
        from_team = task.current_team or task.team
        history = task.handover_history or []
        history.append(
            {
                "from_team": getattr(from_team, 'id', None),
                "from_team_name": getattr(from_team, 'name', None),
                "to_team": target_team.id,
                "to_team_name": target_team.name,
                "by": request.user.username,
                "note": reason,
                "type": "self-initiated",
                "at": timezone.now().isoformat(),
            }
        )
        
        task.current_team = target_team
        assign_task_to_team_leader(task, target_team)
        task.handover_reason = reason
        task.handover_at = timezone.now()
        task.handover_history = history
        task.save(update_fields=['current_team', 'assignee', 'handover_reason', 'handover_at', 'handover_history', 'updated_at'])
        
        TaskComment.objects.create(
            task=task,
            author=request.user,
            type='activity',
            text=f"🔄 {request.user.username} görevi {target_team.name} ekibine devretti (Sebep: {reason})",
        )
        
        push_event(
            {
                "type": "task.self_handover",
                "task_id": task.id,
                "organization": task.organization_id,
                "from_team": getattr(from_team, 'id', None),
                "to_team": target_team.id,
                "by": getattr(request.user, "email", None),
                "reason": reason,
            }
        )
        
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='complete-stage')
    def complete_stage(self, request, pk=None):
        """
        Paralel akış: aktif ekipte bölüm tamamlanınca otomatik olarak sıradaki ekibe geçer.
        Sıralı iş akışı (workflow_team_ids, workflow_parallel kapalı): ekip üyesi «bitir» ile
        onaya gönderir; usta başı approve_section ile sıradaki ekibe devreder veya görevi kapatır.
        İş akışı tanımsızsa fabrika sırası ile devretme / tamamlama (önceki davranış).
        """
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")

        wf_ids_early = workflow_team_id_list(task)
        from_team_early = task.current_team or task.team
        # Sıralı çok adımlı akış: ekip üyesi (lider dahil) önce onaya gönderir — assignee zorunlu değil
        if (
            wf_ids_early
            and not getattr(task, 'workflow_parallel', False)
            and from_team_early
            and int(from_team_early.id) in wf_ids_early
        ):
            is_member_early = from_team_early.members.filter(id=request.user.id).exists()
            is_leader_early = bool(
                getattr(from_team_early, 'leader_id', None) and str(from_team_early.leader_id) == str(request.user.id)
            )
            is_staff_early = getattr(request.user, 'role', '') in ('Admin', 'Manager')
            if not (is_member_early or is_leader_early or is_staff_early):
                raise PermissionDenied("Bu üretim aşamasının ekibinde değilsiniz")
            if not task.assignee_id:
                raise PermissionDenied(
                    "Sıralı akışta önce bölüm usta başı görevi «Üstlen» ile kabul etmelidir; "
                    "ardından üretim kaydı ve onaya gönderme yapılabilir."
                )
            ensure_workflow_state(task)
            gate_err, shortfall_reason = resolve_production_gate(task, from_team_early.id, request.data)
            if gate_err:
                return gate_err
            state_early = dict(task.workflow_stage_state or {})
            key_e = str(int(from_team_early.id))
            st_e = dict(state_early.get(key_e, {}))
            if st_e.get('stage_done'):
                return Response({'detail': 'Bu aşama zaten tamamlandı'}, status=status.HTTP_400_BAD_REQUEST)
            if st_e.get('pending_approval'):
                return Response({'detail': 'Bu aşama zaten usta başı onayı bekliyor'}, status=status.HTTP_400_BAD_REQUEST)
            st_e['pending_approval'] = True
            if shortfall_reason:
                if not _task_has_fire_record(task):
                    return Response(
                        {'detail': 'Eksik üretimde aşama kapatmak için kalemlerden birine fire adedi ve fire sebebi girin.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                done_e = int(st_e.get('qty_done') or 0)
                st_e['production_shortfall_reason'] = format_shortfall_reason_for_storage(done_e, shortfall_reason)
            else:
                st_e.pop('production_shortfall_reason', None)
            state_early[key_e] = st_e
            task.workflow_stage_state = state_early
            task.save(update_fields=['workflow_stage_state', 'updated_at'])
            note_sf = " (hedef altı tamamlama — gerekçe kayıtlı)" if shortfall_reason else ""
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=f"{request.user.username} aşama çalışmasını tamamladı — usta başı onayı bekleniyor{note_sf}",
            )
            sync_workflow_checklist(task)
            return Response(TaskSerializer(task, context={'request': request}).data)

        if not getattr(task, 'workflow_parallel', False) and task.assignee_id != request.user.id:
            raise PermissionDenied("Sadece size atanan görevi bitirebilirsiniz")

        from_team_pre = task.current_team or task.team
        wf_ids_pre = workflow_team_id_list(task)
        if wf_ids_pre and from_team_pre and not getattr(task, 'workflow_parallel', False):
            ensure_workflow_state(task)
            gate_err_pre, _ = resolve_production_gate(task, from_team_pre.id, request.data)
            if gate_err_pre:
                return gate_err_pre

        if getattr(task, 'workflow_parallel', False) and workflow_team_id_list(task):
            wf = workflow_team_id_list(task)
            user_teams = set(request.user.teams.values_list('id', flat=True))
            ensure_workflow_state(task)
            state = dict(task.workflow_stage_state or {})
            active_tid = None
            try:
                current_tid = int(task.current_team_id) if task.current_team_id is not None else None
            except (TypeError, ValueError):
                current_tid = None
            # Otomatik paralel mod: yalnızca aktif akış ekibi çalışır; ekipten herhangi bir üye bitirebilir.
            if current_tid is not None and current_tid in wf and current_tid in user_teams:
                st_cur = state.get(str(current_tid), {}) or {}
                if not st_cur.get('stage_done'):
                    active_tid = current_tid
            if active_tid is None:
                # current_team boş/bozuksa güvenli geri dönüş: ilk açık ekipten devam et.
                for tid in wf:
                    if tid not in user_teams:
                        continue
                    st = state.get(str(tid), {}) or {}
                    if st.get('stage_done'):
                        continue
                    active_tid = tid
                    break
            if active_tid is None:
                raise PermissionDenied("Tamamlanacak uygun bir bölümünüz yok")
            st = dict(state[str(active_tid)])
            gate_err_p, shortfall_p = resolve_production_gate(task, active_tid, request.data)
            if gate_err_p:
                return gate_err_p
            # Paralel akışta lider onayı beklenmez: bölüm otomatik kapanır.
            st['pending_approval'] = False
            st['stage_done'] = True
            st['assignee_id'] = None
            if shortfall_p:
                if not _task_has_fire_record(task):
                    return Response(
                        {'detail': 'Eksik üretimde aşama kapatmak için kalemlerden birine fire adedi ve fire sebebi girin.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                done_p = int(st.get('qty_done') or 0)
                st['production_shortfall_reason'] = format_shortfall_reason_for_storage(done_p, shortfall_p)
            else:
                st.pop('production_shortfall_reason', None)
            state[str(active_tid)] = st
            task.workflow_stage_state = state
            try:
                active_idx = wf.index(active_tid)
            except ValueError:
                active_idx = -1
            all_done = all((state.get(str(tid)) or {}).get('stage_done') for tid in wf)
            next_tid = wf[active_idx + 1] if active_idx >= 0 and active_idx < len(wf) - 1 else None
            update_fields = ['workflow_stage_state', 'updated_at']
            task.assignee = None
            update_fields.append('assignee')
            if next_tid is not None:
                next_team = Team.objects.filter(id=next_tid, organization=request.user.organization).first()
                if next_team:
                    task.current_team = next_team
                    task.team = next_team
                    task.status = 'in-progress'
                    update_fields.extend(['current_team', 'team', 'status'])
            elif all_done:
                task.status = 'done'
                update_fields.append('status')
            task.save(update_fields=list(dict.fromkeys(update_fields)))
            note_p = " (hedef altı tamamlama — gerekçe kayıtlı)" if shortfall_p else ""
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=(
                    f"{request.user.username} bölüm çalışmasını tamamladı{note_p}"
                    + (f" — sıradaki ekip devreye alındı" if next_tid is not None else "")
                    + (" — görev tamamlandı" if next_tid is None and all_done else "")
                ),
            )
            if all_done:
                push_event(
                    {
                        "type": "task.status",
                        "task_id": task.id,
                        "organization": task.organization_id,
                        "title": task.title,
                        "status": "done",
                        "assignee_id": None,
                        "by": getattr(request.user, "email", None),
                    }
                )
            return Response(TaskSerializer(task, context={'request': request}).data)

        org = request.user.organization
        from_team = task.current_team or task.team

        # Tanımlı sıralı iş akışı: sıradaki ekibe geçiş yalnızca approve_section (usta başı onayı) ile
        workflow_ids = getattr(task, 'workflow_team_ids', None) or []
        workflow_ids = [int(x) for x in workflow_ids if x is not None and str(x).isdigit()]
        if workflow_ids and not getattr(task, 'workflow_parallel', False):
            return Response(
                {
                    'detail': (
                        'Sıralı iş akışında sonraki ekibe geçiş için usta başı «Bölümü onayla» '
                        'işlemini kullanmalıdır.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # İş akışı tanımsız: FACTORY_CHECKLIST sırası ile devretme veya tamamlama
        def team_matches_stage(team_name, stage_name):
            tn = (team_name or '').lower()
            sn = (stage_name or '').lower()
            if sn in tn or tn in sn or tn == sn:
                return True
            stage_first = sn.split()[0] if sn else ''
            team_first = tn.split()[0] if tn else ''
            return stage_first in tn or team_first in sn or stage_first == team_first

        ordered_teams = []
        for stage in FACTORY_CHECKLIST:
            for t in Team.objects.filter(organization=org):
                if team_matches_stage(t.name, stage):
                    ordered_teams.append(t)
                    break

        if not ordered_teams:
            ordered_teams = list(Team.objects.filter(organization=org).order_by('name'))

        # Mevcut ekip sıradaki konumunu bul
        current_idx = -1
        next_team = None
        if from_team:
            for i, t in enumerate(ordered_teams):
                if t.id == from_team.id:
                    current_idx = i
                    break
        # Sıradaki ekibe devret veya tamamla
        history = task.handover_history or []
        if next_team or (current_idx >= 0 and current_idx < len(ordered_teams) - 1):
            if not next_team:
                next_team = ordered_teams[current_idx + 1]
            history.append({
                "from_team": from_team.id if from_team else None,
                "from_team_name": from_team.name if from_team else None,
                "to_team": next_team.id,
                "to_team_name": next_team.name,
                "by": request.user.username,
                "note": "Aşama tamamlandı - sıradaki ekibe devir",
                "type": "stage-complete",
                "at": timezone.now().isoformat(),
            })
            task.current_team = next_team
            assign_task_to_team_leader(task, next_team)
            task.handover_reason = "Aşama tamamlandı"
            task.handover_at = timezone.now()
            task.handover_history = history
            task.save(update_fields=['current_team', 'assignee', 'handover_reason', 'handover_at', 'handover_history', 'updated_at'])
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=f"✅ {request.user.username} aşamayı tamamladı → {next_team.name} ekibine devredildi",
            )
            push_event({
                "type": "task.handover",
                "task_id": task.id,
                "organization": task.organization_id,
                "title": task.title,
                "assignee_id": task.assignee_id,
                "to_team": next_team.id,
                "by": getattr(request.user, "email", None),
            })
        else:
            # Son aşama: görev tamamlanır. Çoklu ürün olsa bile akış başa sarmaz.
            task.status = 'done'
            task.assignee = None
            task.save(update_fields=['status', 'assignee', 'updated_at'])
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=f"✅ {request.user.username} görevi tamamen tamamladı",
            )
            push_event(
                {
                    'type': 'task.status',
                    'task_id': task.id,
                    'organization': task.organization_id,
                    'title': task.title,
                    'status': 'done',
                    'assignee_id': None,
                    'by': getattr(request.user, 'email', None),
                }
            )

        sync_workflow_checklist(task)
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='approve-section')
    def approve_section(self, request, pk=None):
        """Usta başı / yönetici: paralel akışta bölüm onayı (tüm bölümler bitince done); sıralı akışta onay sonrası sıradaki ekibe devir veya görevin kapanması."""
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        team_raw = request.data.get('team')
        if team_raw is None:
            return Response({'detail': 'team (ekip id) gerekli'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            team_id = int(team_raw)
        except (TypeError, ValueError):
            return Response({'detail': 'Geçersiz ekip'}, status=status.HTTP_400_BAD_REQUEST)
        team = Team.objects.filter(id=team_id, organization=request.user.organization).first()
        if not team:
            return Response({'detail': 'Ekip bulunamadı'}, status=status.HTTP_404_NOT_FOUND)
        leader_id = getattr(team, 'leader_id', None)
        role = getattr(request.user, 'role', '')
        if role not in ('Admin', 'Manager'):
            if not leader_id or request.user.id != leader_id:
                raise PermissionDenied("Sadece bölüm usta başı veya yönetici onaylayabilir")
        wf = workflow_team_id_list(task)
        if team_id not in wf:
            return Response({'detail': 'Bu görevde bu bölüm yok'}, status=status.HTTP_400_BAD_REQUEST)
        ensure_workflow_state(task)
        state = dict(task.workflow_stage_state or {})
        st = dict(state.get(str(team_id), {}))
        if not st.get('pending_approval'):
            return Response({'detail': 'Onay bekleyen bölüm kaydı yok'}, status=status.HTTP_400_BAD_REQUEST)

        wf_parallel = getattr(task, 'workflow_parallel', False)
        if not wf_parallel and wf:
            ctid = task.current_team_id or task.team_id
            if ctid and int(ctid) != int(team_id):
                return Response(
                    {'detail': 'Yalnızca şu an aktif üretim aşamasının usta başısı onay verebilir'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        st['pending_approval'] = False
        st['stage_done'] = True
        st['assignee_id'] = None
        state[str(team_id)] = st
        task.workflow_stage_state = state

        qty_adjust_fields = []

        if wf_parallel:
            fresh = task.workflow_stage_state or {}
            all_done = all((fresh.get(str(tid)) or {}).get('stage_done') for tid in wf)
            extra_upd = ['workflow_stage_state', 'updated_at']
            if all_done:
                task.status = 'done'
                task.assignee = None
                extra_upd.extend(['status', 'assignee'])
            task.save(update_fields=extra_upd)
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=f"✅ {request.user.username} bölüm ({team.name}) onayladı"
                + (" — görev tamamlandı" if all_done else ""),
            )
            if all_done:
                push_event(
                    {
                        "type": "task.status",
                        "task_id": task.id,
                        "organization": task.organization_id,
                        "title": task.title,
                        "status": "done",
                        "assignee_id": None,
                        "by": getattr(request.user, "email", None),
                    }
                )
            sync_workflow_checklist(task)
            return Response(TaskSerializer(task, context={'request': request}).data)

        org = request.user.organization
        try:
            idx = wf.index(team_id)
        except ValueError:
            return Response({'detail': 'Geçersiz akış adımı'}, status=status.HTTP_400_BAD_REQUEST)

        if idx < len(wf) - 1:
            next_id = wf[idx + 1]
            next_team_obj = Team.objects.filter(id=next_id, organization=org).first()
            if not next_team_obj:
                return Response({'detail': 'Sıradaki ekip bulunamadı'}, status=status.HTTP_400_BAD_REQUEST)
            history = list(task.handover_history or [])
            history.append(
                {
                    "from_team": team.id,
                    "from_team_name": team.name,
                    "to_team": next_team_obj.id,
                    "to_team_name": next_team_obj.name,
                    "by": request.user.username,
                    "note": "Aşama onaylandı - sıradaki ekibe devir",
                    "type": "stage-approved",
                    "at": timezone.now().isoformat(),
                }
            )
            task.current_team = next_team_obj
            task.team = next_team_obj
            task.assignee = None
            task.handover_reason = "Aşama onaylandı"
            task.handover_at = timezone.now()
            task.handover_history = history
            task.status = 'in-progress'
            _handover_fields = [
                'workflow_stage_state',
                'current_team',
                'team',
                'assignee',
                'handover_reason',
                'handover_at',
                'handover_history',
                'status',
                'updated_at',
            ] + qty_adjust_fields
            task.save(update_fields=list(dict.fromkeys(_handover_fields)))
            TaskComment.objects.create(
                task=task,
                author=request.user,
                type='activity',
                text=(
                    f"✅ {request.user.username} ({team.name}) aşamasını onayladı → "
                    f"{next_team_obj.name} devam ediyor"
                ),
            )
            push_event(
                {
                    "type": "task.handover",
                    "task_id": task.id,
                    "organization": task.organization_id,
                    "title": task.title,
                    "assignee_id": task.assignee_id,
                    "to_team": next_team_obj.id,
                    "by": getattr(request.user, "email", None),
                }
            )
            sync_workflow_checklist(task)
            return Response(TaskSerializer(task, context={'request': request}).data)

        # Eski görevleri kalem-bazlı workflow formatına otomatik dönüştür.
        if ensure_product_line_workflows(task):
            task.save(
                update_fields=[
                    'product_lines',
                    'workflow_team_ids',
                    'workflow_parallel',
                    'current_team',
                    'team',
                    'updated_at',
                ]
            )
        lines = list(getattr(task, 'product_lines', None) or [])
        active = int(getattr(task, 'active_product_index', 0) or 0)
        # NOT: Çoklu ürün olsa bile, son ekip onayından sonra workflow başa sarmaz.

        task.status = 'done'
        task.assignee = None
        _done_fields = ['workflow_stage_state', 'status', 'assignee', 'updated_at'] + qty_adjust_fields
        task.save(update_fields=list(dict.fromkeys(_done_fields)))
        TaskComment.objects.create(
            task=task,
            author=request.user,
            type='activity',
            text=f"✅ {request.user.username} son aşamayı ({team.name}) onayladı — görev tamamlandı",
        )
        push_event(
            {
                'type': 'task.status',
                'task_id': task.id,
                'organization': task.organization_id,
                'title': task.title,
                'status': 'done',
                'assignee_id': None,
                'by': getattr(request.user, "email", None),
            }
        )
        sync_workflow_checklist(task)
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='log-production')
    def log_production(self, request, pk=None):
        """Günlük tamamlanan adet — isteğe bağlı siparişe quantity_produced yansır."""
        task = self.get_object()
        if task.organization_id != request.user.organization_id:
            raise PermissionDenied("Farklı organizasyon")
        try:
            qty = int(request.data.get('quantity', 0) or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty < 0:
            return Response({'detail': 'quantity negatif olamaz'}, status=status.HTTP_400_BAD_REQUEST)
        day_raw = request.data.get('entry_date')
        if day_raw:
            try:
                day = dt_datetime.strptime(str(day_raw)[:10], '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'entry_date YYYY-MM-DD olmalı'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            day = timezone.now().date()
        if task.sales_order_id:
            from erp.models import SalesOrder

            so_gate = SalesOrder.objects.filter(id=task.sales_order_id, organization_id=task.organization_id).first()
            if so_gate:
                ordered_gate = int(getattr(so_gate, 'order_quantity', 0) or 0)
                produced_gate = int(getattr(so_gate, 'quantity_produced', 0) or 0)
                if ordered_gate > 0 and produced_gate >= ordered_gate:
                    return Response(
                        {'detail': 'Sipariş hedef adedine ulaşıldı; yeni üretim girişi yapılamaz.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
        user = request.user
        wf = workflow_team_id_list(task)
        user_teams = list(user.teams.values_list('id', flat=True))
        user_team_set = set(user_teams)
        team_raw = request.data.get('team')
        tid = int(team_raw) if team_raw not in (None, '') else None
        if tid is None:
            # Sıralı workflow'ta hangi ekip aktifse, üretim kaydı o ekibe gitmeli.
            # Kullanıcı birden fazla workflow ekibinin üyesi ise eski mantık ilk uygun ekibi
            # seçip A/B verisini karıştırabiliyordu.
            if wf and not getattr(task, 'workflow_parallel', False) and getattr(task, 'current_team_id', None):
                try:
                    cur_tid = int(task.current_team_id)
                except (TypeError, ValueError):
                    cur_tid = None
                if cur_tid is not None and cur_tid in wf and cur_tid in user_teams:
                    tid = cur_tid
            # Paralel workflow: aynı kullanıcı birden fazla bölümde üyeyse, ekip seçimini
            # workflow_team_ids sırasına göre yap (takım tablosu sırasına göre değil).
            if tid is None and wf and getattr(task, 'workflow_parallel', False):
                ensure_workflow_state(task)
                state = dict(task.workflow_stage_state or {})
                open_for_user = []
                for wtid in wf:
                    if wtid not in user_team_set:
                        continue
                    st = state.get(str(wtid), {}) or {}
                    if st.get('stage_done'):
                        continue
                    aid = st.get('assignee_id')
                    try:
                        aid_claimable = aid in (None, '') or int(aid) == int(user.id)
                    except (TypeError, ValueError):
                        aid_claimable = aid in (None, '')
                    if not aid_claimable:
                        continue
                    open_for_user.append(wtid)
                if len(open_for_user) == 1:
                    tid = open_for_user[0]
                elif len(open_for_user) > 1:
                    return Response(
                        {
                            'detail': (
                                'Birden fazla bölümde üyesiniz; üretimi hangi ekip için girdiğinizi '
                                'body içinde `team` (ekip id) olarak gönderin.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            if tid is None:
                # Paralel akışta ekip yukarıda çözülmediyse rastgele sırayla seçme (yanlış bölüme yazım riski).
                if not (wf and getattr(task, 'workflow_parallel', False)):
                    for t in user_teams:
                        if wf and t in wf:
                            tid = t
                            break
                        if not wf and t == (task.current_team_id or task.team_id):
                            tid = t
                            break
        if tid is None:
            return Response({'detail': 'Bölüm (ekip) belirlenemedi'}, status=status.HTTP_400_BAD_REQUEST)
        # Sıralı akışta ekip seçimi istemciden bağımsız olarak daima aktif aşamadır.
        # Böylece bir önceki aşamadan kalan/stale `team` değeri yanlış yetki hatası üretmez.
        if wf and not getattr(task, 'workflow_parallel', False) and task.current_team_id:
            try:
                tid = int(task.current_team_id)
            except (TypeError, ValueError):
                return Response({'detail': 'Aktif bölüm bilgisi geçersiz'}, status=status.HTTP_400_BAD_REQUEST)
        if wf and tid not in wf:
            return Response({'detail': 'Bu görev akışında bu ekip yok'}, status=status.HTTP_400_BAD_REQUEST)
        team_row = Team.objects.filter(id=tid, organization_id=task.organization_id).first()
        if not team_row:
            return Response({'detail': 'Ekip bulunamadı'}, status=status.HTTP_400_BAD_REQUEST)
        role_u = getattr(user, 'role', '')
        is_staff_u = role_u in ('Admin', 'Manager')
        is_member_u = team_row.members.filter(id=user.id).exists()
        is_leader_u = bool(getattr(team_row, 'leader_id', None) and str(team_row.leader_id) == str(user.id))
        is_assignee_u = bool(task.assignee_id and str(task.assignee_id) == str(user.id))
        # user.teams ilişkisini de dikkate al (bazı canlı verilerde members M2M senkronu gecikebiliyor).
        is_team_linked_u = bool(tid in user_team_set)
        if not (is_staff_u or is_member_u or is_leader_u or is_assignee_u or is_team_linked_u):
            raise PermissionDenied("Bu ekibin üyesi/lideri değilsiniz")
        if (
            wf
            and not getattr(task, 'workflow_parallel', False)
            and task.current_team_id
            and int(tid) == int(task.current_team_id)
            and not task.assignee_id
        ):
            return Response(
                {
                    'detail': (
                        "Sıralı iş akışında üretim kaydı için önce bölüm usta başının görevi "
                        "«Üstlen» ile kabul etmesi gerekir."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        lines = list(getattr(task, 'product_lines', None) or [])
        pl_raw = request.data.get('product_line_index')
        line_idx = None
        if pl_raw is not None and str(pl_raw).strip() != '':
            try:
                line_idx = int(pl_raw)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'product_line_index geçerli bir tam sayı olmalı'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if line_idx < 0 or (lines and line_idx >= len(lines)):
                return Response(
                    {'detail': 'product_line_index bu görevdeki kalem aralığında değil'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        if len(lines) > 1:
            if line_idx is None:
                return Response(
                    {
                        'detail': (
                            'Bu görevde birden fazla ürün kalemi var. '
                            'Üretimi hangi kalem için girdiğinizi belirtin (product_line_index, 0 ile başlar).'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif len(lines) == 1:
            if line_idx is None:
                line_idx = 0

        save_lines = False
        prev_q = 0
        active_line = dict(lines[line_idx] or {}) if (line_idx is not None and lines and 0 <= line_idx < len(lines)) else {}
        line_wf_ids = [int(x) for x in (active_line or {}).get('workflow_team_ids', []) if str(x).isdigit()]
        # Kalem-bazlı workflow: ekip doğrulaması + adım otomatik kapanışı bu kalem state'inde tutulur.
        if line_wf_ids:
            if tid not in line_wf_ids:
                return Response({'detail': 'Bu kalem akışında bu ekip yok'}, status=status.HTTP_400_BAD_REQUEST)
            lstate = dict(active_line.get('workflow_stage_state') or {})
            lst = dict(lstate.get(str(tid)) or {})
            prev_team_done = int(lst.get('qty_done') or 0)
            lst['qty_done'] = qty
            tgt = int(lst.get('qty_target') or max(1, int(active_line.get('quantity') or 1)))
            if qty >= tgt:
                lst['stage_done'] = True
                lst['pending_approval'] = False
                lst['assignee_id'] = None
            lstate[str(tid)] = lst
            # sıradaki açık ekip current_team olur
            next_tid = None
            for wtid in line_wf_ids:
                st_row = dict(lstate.get(str(wtid)) or {})
                if not st_row.get('stage_done'):
                    next_tid = int(wtid)
                    break
            active_line['workflow_stage_state'] = lstate
            active_line['current_team_id'] = next_tid
            lines[line_idx] = active_line
            task.product_lines = lines
            save_lines = True
            # tüm kalemlerin tüm adımları bittiyse görev kapanır
            all_lines_done = True
            for ln in lines:
                ids_ln = [int(x) for x in (ln or {}).get('workflow_team_ids', []) if str(x).isdigit()]
                if not ids_ln:
                    continue
                st_ln = dict((ln or {}).get('workflow_stage_state') or {})
                for wtid in ids_ln:
                    if not (st_ln.get(str(wtid)) or {}).get('stage_done'):
                        all_lines_done = False
                        break
                if not all_lines_done:
                    break
            if all_lines_done and task.status in ('todo', 'in-progress'):
                task.status = 'done'
                task.assignee = None
                task.save(update_fields=['product_lines', 'status', 'assignee', 'updated_at'])
                TaskComment.objects.create(
                    task=task,
                    author=user,
                    type='activity',
                    text='✅ Tüm kalemlerin akış adımları tamamlandı — görev otomatik kapatıldı',
                )
            else:
                task.save(update_fields=['product_lines', 'updated_at'])
            TaskProductionEntry.objects.create(
                task=task,
                user=user,
                team_id=tid,
                product_line_index=line_idx,
                entry_date=day,
                quantity=qty,
                note=str(request.data.get('note', '') or '')[:255],
            )
            if task.sales_order_id:
                from erp.models import SalesOrder
                so = SalesOrder.objects.filter(id=task.sales_order_id, organization_id=task.organization_id).first()
                if so:
                    ordered = int(getattr(so, 'order_quantity', 0) or 0)
                    produced = int(getattr(so, 'quantity_produced', 0) or 0)
                    delta_so = qty - prev_team_done
                    next_produced = max(0, produced + delta_so)
                    if ordered and next_produced > ordered:
                        next_produced = ordered
                    so.quantity_produced = next_produced
                    so.save(update_fields=['quantity_produced'])
            return Response(TaskSerializer(task, context={'request': request}).data)
        # Workflow varsa üretim "aşama/ekip" bazında tutulur (workflow_stage_state.qty_done).
        # product_lines.qty_produced yalnızca workflow olmayan görevlerde kullanılır.
        if not wf and line_idx is not None and lines and 0 <= line_idx < len(lines):
            row = dict(lines[line_idx] or {})
            try:
                prev_q = int(row.get('qty_produced') or 0)
            except (TypeError, ValueError):
                prev_q = 0
            row['qty_produced'] = qty
            lines[line_idx] = row
            task.product_lines = lines
            save_lines = True

        ensure_workflow_state(task)
        state = dict(task.workflow_stage_state or {})
        state_changed = False
        workflow_checklist_dirty = False
        prev_team_done = 0
        if str(tid) in state:
            try:
                prev_team_done = int((state[str(tid)] or {}).get('qty_done', 0) or 0)
            except (TypeError, ValueError):
                prev_team_done = 0
        if wf and str(tid) in state:
            st = dict(state[str(tid)])
            # Çoklu ürün kaleminde ekip üretimi kalem bazında tutulur.
            # qty_done: geriye uyum için "aktif kalem" sayısı (veya tek kalem)
            # qty_done_by_line: { "<index>": qty }
            qmap = dict(st.get('qty_done_by_line') or {})
            if line_idx is not None:
                qmap[str(int(line_idx))] = qty
            st['qty_done_by_line'] = qmap
            lines_count = len(lines or [])
            if lines_count <= 1 or line_idx is None:
                st['qty_done'] = qty
            else:
                total_done = 0
                for v in qmap.values():
                    try:
                        total_done += max(0, int(v or 0))
                    except (TypeError, ValueError):
                        continue
                st['qty_done'] = total_done
            state[str(tid)] = st
            task.workflow_stage_state = state
            state_changed = True
            workflow_checklist_dirty = True

        save_fields = []
        if state_changed:
            save_fields.extend(['workflow_stage_state', 'updated_at'])
        if save_lines:
            save_fields.append('product_lines')
        if save_fields:
            task.save(update_fields=list(dict.fromkeys(save_fields)))

        # Paralel akışta: ekip üretimi hedefi yakaladığında adımı otomatik tamamla.
        # Böylece ekip lideri onayı/ek buton tıklaması beklenmeden sonraki akış ilerler.
        if wf and getattr(task, 'workflow_parallel', False) and str(tid) in state:
            st_auto = dict((task.workflow_stage_state or {}).get(str(tid), {}) or {})
            if not st_auto.get('stage_done') and workflow_stage_production_met(task, tid):
                st_auto['pending_approval'] = False
                st_auto['stage_done'] = True
                st_auto['assignee_id'] = None
                state[str(tid)] = st_auto
                task.workflow_stage_state = state
                all_done = all((state.get(str(t)) or {}).get('stage_done') for t in wf)
                upd = ['workflow_stage_state', 'updated_at']
                next_tid = None
                if not all_done:
                    try:
                        i_cur = wf.index(int(tid))
                    except (TypeError, ValueError):
                        i_cur = -1
                    if i_cur >= 0:
                        for cand in wf[i_cur + 1:]:
                            if not (state.get(str(cand)) or {}).get('stage_done'):
                                next_tid = cand
                                break
                if all_done:
                    task.status = 'done'
                    task.assignee = None
                    upd.extend(['status', 'assignee'])
                elif next_tid is not None:
                    next_team = Team.objects.filter(id=next_tid, organization_id=task.organization_id).first()
                    if next_team:
                        task.current_team = next_team
                        task.team = next_team
                        task.assignee = None
                        task.status = 'in-progress'
                        upd.extend(['current_team', 'team', 'assignee', 'status'])
                task.save(update_fields=list(dict.fromkeys(upd)))
                workflow_checklist_dirty = True
                TaskComment.objects.create(
                    task=task,
                    author=user,
                    type='activity',
                    text=(
                        f"{user.username} üretim girişi ile bölüm hedefini tamamladı — adım otomatik kapatıldı"
                        + (f" — {next_team.name} ekibine devredildi" if (not all_done and next_tid is not None and 'next_team' in locals() and next_team) else "")
                        + (" — görev tamamlandı" if all_done else "")
                    ),
                )
                if all_done:
                    push_event(
                        {
                            'type': 'task.status',
                            'task_id': task.id,
                            'organization': task.organization_id,
                            'title': task.title,
                            'status': 'done',
                            'assignee_id': None,
                            'by': getattr(user, 'email', None),
                        }
                    )
        if wf and workflow_checklist_dirty:
            sync_workflow_checklist(task)

        TaskProductionEntry.objects.create(
            task=task,
            user=user,
            team_id=tid,
            product_line_index=line_idx,
            entry_date=day,
            quantity=qty,
            note=str(request.data.get('note', '') or '')[:255],
        )

        auto_done = False
        if not wf and lines:
            all_met = True
            for ln in lines:
                try:
                    tgt = max(1, int((ln or {}).get('quantity') or 1))
                except (TypeError, ValueError):
                    tgt = 1
                try:
                    prod = int((ln or {}).get('qty_produced') or 0)
                except (TypeError, ValueError):
                    prod = 0
                if prod < tgt:
                    all_met = False
                    break
            if all_met and task.status in ('todo', 'in-progress'):
                task.status = 'done'
                task.assignee = None
                task.save(update_fields=['status', 'assignee', 'updated_at'])
                TaskComment.objects.create(
                    task=task,
                    author=user,
                    type='activity',
                    text=(
                        f'✅ Tüm ürün kalemleri hedef üretime ulaştı — '
                        f'{user.username} (iş akışı olmayan görev, otomatik tamamlandı)'
                    ),
                )
                push_event(
                    {
                        'type': 'task.status',
                        'task_id': task.id,
                        'organization': task.organization_id,
                        'title': task.title,
                        'status': 'done',
                        'assignee_id': None,
                        'by': getattr(user, 'email', None),
                    }
                )
                auto_done = True
        if auto_done:
            sync_workflow_checklist(task)

        if task.sales_order_id:
            from erp.models import SalesOrder

            so = SalesOrder.objects.filter(id=task.sales_order_id, organization_id=task.organization_id).first()
            if so:
                ordered = int(getattr(so, 'order_quantity', 0) or 0)
                produced = int(getattr(so, 'quantity_produced', 0) or 0)
                if save_lines:
                    delta_so = qty - prev_q
                elif state_changed:
                    delta_so = qty - prev_team_done
                elif not wf:
                    # Ürün satırı + iş akışı yok: eski davranış (günlük dilim artışı)
                    if ordered > 0:
                        delta_so = min(qty, max(0, ordered - produced))
                    else:
                        delta_so = qty
                else:
                    delta_so = 0
                next_produced = produced + delta_so
                if next_produced < 0:
                    next_produced = 0
                if ordered and next_produced > ordered:
                    next_produced = ordered
                so.quantity_produced = next_produced
                so.save(update_fields=['quantity_produced'])
        return Response(TaskSerializer(task, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='production-report')
    def production_report(self, request):
        """Günlük üretim: ?date=YYYY-MM-DD (yönetici)"""
        if getattr(request.user, 'role', '') not in ('Admin', 'Manager'):
            raise PermissionDenied("Bu rapora yalnızca Admin veya Manager erişebilir")
        org = getattr(request.user, 'organization', None)
        if not org:
            return Response({'date': None, 'entries': [], 'entries_count': 0, 'by_team_entry_counts': {}})
        day_raw = request.query_params.get('date')
        if day_raw:
            try:
                d = dt_datetime.strptime(str(day_raw)[:10], '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'date YYYY-MM-DD olmalı'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            d = timezone.now().date()
        rows = (
            TaskProductionEntry.objects.filter(task__organization=org, entry_date=d)
            .select_related('task', 'user', 'team')
            .order_by('-created_at')
        )
        entries = [
            {
                'id': e.id,
                'task_id': e.task_id,
                'task_title': e.task.title,
                'user_id': e.user_id,
                'user_name': e.user.username if e.user else None,
                'team_id': e.team_id,
                'team_name': e.team.name if e.team else None,
                'product_line_index': e.product_line_index,
                'quantity': e.quantity,
                'note': e.note,
                'created_at': e.created_at.isoformat() if e.created_at else None,
            }
            for e in rows
        ]
        # quantity her kayıtta o an bildirilen mutlak üretimdir; satırların toplamı fiziksel çıktı değildir.
        by_team_entry_counts = {}
        for e in entries:
            key = e['team_name'] or '—'
            by_team_entry_counts[key] = by_team_entry_counts.get(key, 0) + 1
        return Response(
            {
                'date': str(d),
                'entries': entries,
                'entries_count': len(entries),
                'by_team_entry_counts': by_team_entry_counts,
            }
        )

    @action(detail=False, methods=['get'], url_path='my-team-queue')
    def my_team_queue(self, request):
        """
        Worker için: sıralı akışta yalnızca ilgili ekip usta başısı (havuzdaki görevler);
        paralel akışta ilgili bölüm üyeleri. Liste sonunda kullanıcıya göre filtrelenir.
        """
        user = request.user
        org = getattr(user, 'organization', None)
        if not org:
            return Response([])
        member_team_ids = set(user.teams.values_list('id', flat=True))
        leader_team_ids = set(Team.objects.filter(organization=org, leader_id=user.id).values_list('id', flat=True))
        user_team_ids = list(member_team_ids | leader_team_ids)
        if not user_team_ids:
            return Response([])
        assignee_in_current_team = Team.members.through.objects.filter(
            team_id=OuterRef('current_team_id'),
            user_id=OuterRef('assignee_id'),
        )
        base = Task.objects.filter(organization=org, current_team_id__in=user_team_ids).exclude(status='done')
        # Sıralı akış: yalnızca havuz (assignee boş) — lider atanınca görev "Al" kuyruğunda görünmez, çift tıklama/azdırma önlenir.
        qs_seq = base.filter(workflow_parallel=False, assignee__isnull=True)
        qs_par = base.filter(workflow_parallel=True).filter(Q(assignee__isnull=True) | Exists(assignee_in_current_team))
        qs = (qs_seq | qs_par).select_related('team', 'current_team').distinct().order_by('-updated_at')
        seen = {t.id for t in qs}
        out = list(qs)
        par_ids = (
            Task.objects.filter(organization=org, workflow_parallel=True)
            .exclude(status='done')
            .exclude(id__in=seen)
            .values_list('id', flat=True)
        )
        for tid in par_ids:
            t = (
                Task.objects.filter(id=tid)
                .select_related('team', 'current_team')
                .first()
            )
            if t and parallel_queue_visible(t, user, user_team_ids):
                out.append(t)
                seen.add(t.id)
        out.sort(key=lambda x: x.updated_at, reverse=True)
        # Kalem-bazlı akışta current_team farklı olsa bile kullanıcı kendi açık kalemlerini görsün.
        line_candidates = (
            Task.objects.filter(organization=org)
            .exclude(status='done')
            .exclude(id__in=seen)
            .select_related('team', 'current_team')
            .order_by('-updated_at')[:400]
        )
        for t in line_candidates:
            lines = list(getattr(t, 'product_lines', None) or [])
            if not lines:
                continue
            if user_can_see_claim_queue_entry(t, user):
                out.append(t)
                seen.add(t.id)
        out.sort(key=lambda x: x.updated_at, reverse=True)
        out = [t for t in out if user_can_see_claim_queue_entry(t, user)]
        return Response(TaskSerializer(out, many=True, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='worker-tracking')
    def worker_tracking(self, request):
        """
        Admin/Manager için worker tracking endpoint'i.
        Her worker'ın hangi departmanda çalıştığını gösterir.
        """
        # Sadece Admin ve Manager erişebilir
        if request.user.role not in ['Admin', 'Manager']:
            raise PermissionDenied("Bu endpoint'e sadece Admin ve Manager erişebilir")
        
        org = request.user.organization
        
        # Tüm worker'ları al
        workers = User.objects.filter(organization=org, role='Worker')
        
        tracking_data = []
        for worker in workers:
            # Worker'ın aktif görevlerini al
            active_tasks = Task.objects.filter(
                organization=org,
                assignee=worker,
                status__in=['todo', 'in-progress']
            ).select_related('team', 'current_team').order_by('-updated_at')
            
            # Worker'ın son self-handover'ını bul
            last_handover = None
            for task in active_tasks:
                if task.handover_history:
                    for h in reversed(task.handover_history):
                        if h.get('type') == 'self-initiated' and h.get('by') == worker.username:
                            last_handover = h
                            break
                    if last_handover:
                        break
            
            # Worker'ın ana ekibi
            primary_teams = list(worker.teams.values_list('name', flat=True))
            
            # Worker'ın şu an çalıştığı departman
            current_department = None
            if active_tasks.exists():
                task = active_tasks.first()
                current_department = task.current_team.name if task.current_team else (task.team.name if task.team else None)
            
            tracking_data.append({
                'worker_id': worker.id,
                'worker_name': worker.username,
                'worker_email': worker.email,
                'primary_teams': primary_teams,
                'current_department': current_department,
                'active_tasks_count': active_tasks.count(),
                'last_handover': last_handover,
                'last_activity': active_tasks.first().updated_at if active_tasks.exists() else None,
                'has_account': True,
            })
        
        field_staff = []
        for a in (
            TeamAssociate.objects.filter(organization=org, is_active=True)
            .prefetch_related('teams')
            .order_by('full_name', 'id')
        ):
            team_names = list(a.teams.values_list('name', flat=True))
            field_staff.append(
                {
                    'associate_id': a.id,
                    'worker_name': a.full_name,
                    'worker_email': '',
                    'contact': a.phone or '—',
                    'primary_teams': team_names,
                    'team_ids': list(a.teams.values_list('id', flat=True)),
                    'current_department': None,
                    'active_tasks_count': 0,
                    'last_handover': None,
                    'last_activity': a.updated_at.isoformat() if a.updated_at else None,
                    'has_account': False,
                    'notes': (a.notes or '')[:200],
                }
            )

        return Response(
            {
                'workers': tracking_data,
                'total_workers': len(tracking_data),
                'field_workers': field_staff,
                'total_field_workers': len(field_staff),
                'timestamp': timezone.now().isoformat(),
            }
        )

    @action(detail=False, methods=['get'], url_path='worker-detail')
    def worker_detail(self, request):
        """Çalışan detay: günlük/aylık süre, aktif/biten görevler. Query: ?worker_id=123"""
        if request.user.role not in ['Admin', 'Manager']:
            raise PermissionDenied("Bu endpoint'e sadece Admin ve Manager erişebilir")
        org = request.user.organization

        worker_id = request.query_params.get('worker_id')
        if not worker_id:
            return Response({'detail': 'worker_id gerekli'}, status=400)
        worker = User.objects.filter(organization=org, id=worker_id, role='Worker').first()
        if not worker:
            return Response({'detail': 'Çalışan bulunamadı'}, status=404)

        # Görevler: atanan veya sahip
        tasks = Task.objects.filter(organization=org).filter(
            Q(assignee=worker) | Q(owner=worker)
        ).select_related('team', 'current_team').order_by('-updated_at')

        active_tasks = list(tasks.filter(status__in=['todo', 'in-progress']).values(
            'id', 'title', 'status', 'priority', 'due', 'start', 'end', 'updated_at'
        ))
        completed_tasks = list(tasks.filter(status='done').values(
            'id', 'title', 'status', 'priority', 'due', 'start', 'end', 'updated_at'
        ))[:50]

        # Zaman kayıtlarından günlük/aylık toplam
        from .models import TaskTimeEntry
        entries = TaskTimeEntry.objects.filter(
            task__organization=org,
            user=worker
        ).only('started_at', 'ended_at')

        daily_hours = {}
        monthly_hours = {}
        for e in entries:
            start = e.started_at
            end = e.ended_at or timezone.now()
            hrs = max(0, (end - start).total_seconds() / 3600)
            day_key = start.date().isoformat()
            month_key = start.strftime('%Y-%m')
            daily_hours[day_key] = daily_hours.get(day_key, 0) + hrs
            monthly_hours[month_key] = monthly_hours.get(month_key, 0) + hrs

        return Response({
            'worker_id': worker.id,
            'worker_name': worker.username,
            'worker_email': worker.email,
            'primary_teams': list(worker.teams.values_list('name', flat=True)),
            'active_tasks': active_tasks,
            'completed_tasks': completed_tasks,
            'daily_hours': daily_hours,
            'monthly_hours': monthly_hours,
        })

    def perform_destroy(self, instance):
        log_entity_action(instance, 'deleted', user=self.request.user)
        instance.delete()

    def run_automations(self, task, trigger, extra=None):
        if getattr(task, '_automation_in_progress', False):
            return
        task._automation_in_progress = True
        org = getattr(self.request.user, 'organization', None)
        rules = AutomationRule.objects.filter(organization=org, is_active=True, trigger=trigger)
        for rule in rules:
            cond = rule.condition or {}
            if trigger == 'task_status_changed':
                if cond.get('from') and extra and extra.get('from') != cond.get('from'):
                    continue
                if cond.get('to') and extra and extra.get('to') != cond.get('to'):
                    continue
            if trigger == 'task_due_soon':
                continue
            try:
                if rule.action == 'add_comment':
                    text = rule.action_payload.get('comment') if isinstance(rule.action_payload, dict) else None
                    TaskComment.objects.create(task=task, author=None, type='activity', text=text or 'Otomasyon')
                elif rule.action == 'set_assignee':
                    assignee_id = rule.action_payload.get('assignee') if isinstance(rule.action_payload, dict) else None
                    if assignee_id:
                        task.assignee_id = assignee_id
                        task.save(update_fields=['assignee'])
                elif rule.action == 'add_tag':
                    payload = rule.action_payload if isinstance(rule.action_payload, dict) else {}
                    tag = payload.get('tag')
                    if tag:
                        current = task.tags or []
                        if tag not in current:
                            task.tags = current + [tag]
                            task.save(update_fields=['tags'])
                            TaskComment.objects.create(task=task, author=None, type='activity', text=f"Etiket eklendi: {tag}")
                elif rule.action == 'set_field':
                    payload = rule.action_payload if isinstance(rule.action_payload, dict) else {}
                    field = payload.get('field')
                    value = payload.get('value')
                    allowed = ['priority', 'status']
                    if field in allowed and value:
                        setattr(task, field, value)
                        task.save(update_fields=[field])
                        TaskComment.objects.create(task=task, author=None, type='activity', text=f"{field} -> {value}")
                elif rule.action == 'notify':
                    payload = rule.action_payload if isinstance(rule.action_payload, dict) else {}
                    msg = payload.get('message') or f"Task {task.title}: durum {extra.get('from')} -> {extra.get('to')}"
                    webhook = payload.get('webhook')
                    email_to = payload.get('email') or (task.assignee.email if task.assignee else None)
                    TaskComment.objects.create(task=task, author=None, type='activity', text='(Notify) Otomasyon tetiklendi')
                    send_slack_webhook(msg, webhook_url=webhook)
                    if email_to:
                        send_email(email_to, f"Görev bildirimi: {task.title}", msg)
                    push_event(
                        {
                            "type": "notification.automation",
                            "task_id": task.id,
                            "task_title": task.title,
                            "message": msg,
                            "organization": task.organization_id,
                            "rule_id": rule.id,
                        }
                    )
                elif rule.action == 'multi_notify':
                    payload = rule.action_payload if isinstance(rule.action_payload, dict) else {}
                    msg = payload.get('message') or f"Task {task.title}: durum {extra.get('from')} -> {extra.get('to')}"
                    emails = payload.get('emails') or []
                    hooks = payload.get('webhooks') or []
                    TaskComment.objects.create(task=task, author=None, type='activity', text='(Multi notify) Otomasyon tetiklendi')
                    for h in hooks:
                        send_slack_webhook(msg, webhook_url=h)
                    for em in emails:
                        send_email(em, f"Görev bildirimi: {task.title}", msg)
                    push_event(
                        {
                            "type": "notification.automation",
                            "task_id": task.id,
                            "task_title": task.title,
                            "message": msg,
                            "organization": task.organization_id,
                            "rule_id": rule.id,
                        }
                    )
            except Exception as e:
                TaskComment.objects.create(task=task, author=None, type='activity', text=f"Otomasyon hata: {e}")
        task._automation_in_progress = False


class TaskAttachmentViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TaskAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, CommentOnlyRestriction, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'destroy': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
    }
    queryset = TaskAttachment.objects.select_related('task')

    def perform_update(self, serializer):
        # only description/metadata allowed
        serializer.save()

    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(task__organization=org)
        return qs

    def perform_create(self, serializer):
        file_obj = self.request.FILES.get('file')
        # versioning: if same original name exists for this task, increment
        latest = (
            TaskAttachment.objects.filter(task=serializer.validated_data['task'], original_name=getattr(file_obj, 'name', ''))
            .order_by('-version')
            .first()
        )
        next_version = (latest.version + 1) if latest else 1
        serializer.save(
            uploaded_by=self.request.user,
            original_name=getattr(file_obj, 'name', ''),
            content_type=getattr(file_obj, 'content_type', ''),
            size=getattr(file_obj, 'size', 0),
            version=next_version,
            parent=latest if latest else None,
        )
        push_event(
            {
                "type": "task.attachment",
                "task_id": serializer.validated_data['task'].id,
                "organization": serializer.validated_data['task'].organization_id,
                "by": getattr(self.request.user, "email", None),
                "name": getattr(file_obj, 'name', ''),
            }
        )

    def perform_destroy(self, instance):
        user = self.request.user
        # Fine RBAC: owner/assignee or admin/manager can delete
        if user.role not in ['Admin', 'Manager']:
            if not (instance.task.owner_id == user.id or instance.task.assignee_id == user.id):
                raise PermissionDenied("Bu eki silme yetkiniz yok")
        instance.delete()
        push_event(
            {
                "type": "task.attachment.deleted",
                "task_id": instance.task_id,
                "organization": instance.task.organization_id,
                "by": getattr(self.request.user, "email", None),
                "attachment_id": instance.id,
            }
        )


class TaskCommentViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TaskCommentSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, CommentOnlyRestriction, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'destroy': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
    }
    queryset = TaskComment.objects.select_related('task', 'author')

    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(task__organization=org)
        return qs

    def perform_create(self, serializer):
        comment = serializer.save(author=self.request.user)
        # SSE event for new comment
        push_event(
            {
                "type": "task.comment",
                "task_id": comment.task_id,
                "organization": comment.task.organization_id,
                "by": getattr(self.request.user, "email", None),
            }
        )
        # Mention tespiti (@username) + bildirim/eposta
        mentions = set(re.findall(r'@([\\w\\.-]+)', comment.text or ''))
        if mentions:
            users = User.objects.filter(username__in=list(mentions), organization=comment.task.organization)
            if users:
                TaskComment.objects.create(
                    task=comment.task,
                    author=None,
                    type='activity',
                    text=f"Mention: {', '.join([u.username for u in users])}",
                )
            for u in users:
                push_event(
                    {
                        "type": "notification.mention",
                        "task_id": comment.task_id,
                        "task_title": comment.task.title,
                        "comment_id": comment.id,
                        "mentioned": u.username,
                        "organization": comment.task.organization_id,
                    }
                )
                if u.email:
                    send_email(
                        u.email,
                        f"Mention: {comment.task.title}",
                        f"{self.request.user.username} sizi bir yorumda bahsetti: {comment.text}",
                    )

    def perform_destroy(self, instance):
        user = self.request.user
        # Only Admin/Manager or task owner/assignee can delete comments
        if user.role not in ['Admin', 'Manager']:
            if not (instance.task.owner_id == user.id or instance.task.assignee_id == user.id):
                raise permissions.PermissionDenied("Bu yorumu silme yetkiniz yok")
        instance.delete()


class TaskTimeEntryViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TaskTimeEntrySerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'destroy': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
    }
    queryset = TaskTimeEntry.objects.select_related('task', 'user')

    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(task__organization=org)
        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class AutomationRuleViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = AutomationRuleSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'tasks.edit'
    queryset = AutomationRule.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
          qs = qs.filter(organization=org)
        return qs

    def perform_create(self, serializer):
        org = getattr(self.request.user, 'organization', None)
        serializer.save(organization=org)

    @action(detail=False, methods=['get'])
    def help(self, request):
        data = {
            "triggers": [
                {"code": "task_status_changed", "payload": {"from": "todo", "to": "done"}},
                {"code": "task_due_soon", "payload": {"hours": 24}},
                {"code": "task_created", "payload": {}},
            ],
            "actions": [
                {"code": "add_comment", "payload": {"comment": "Metin"}},
                {"code": "set_assignee", "payload": {"assignee": 1}},
                {"code": "add_tag", "payload": {"tag": "SLA"}},
                {"code": "set_field", "payload": {"field": "priority", "value": "high"}},
                {"code": "notify", "payload": {"message": "metin", "email": "a@b.com", "webhook": "https://..."}},
                {"code": "multi_notify", "payload": {"message": "metin", "emails": ["a@b.com"], "webhooks": ["https://..."]}},
            ],
        }
        return Response(data)

    @action(detail=False, methods=['post'])
    def test(self, request):
        """
        Dry-run automation: evaluates condition and returns would-be actions without side effects.
        Body: {trigger, condition, action, action_payload, sample_task: {...}, extra: {...}}
        """
        trigger = request.data.get('trigger')
        condition = request.data.get('condition') or {}
        action = request.data.get('action')
        payload = request.data.get('action_payload') or {}
        task_data = request.data.get('sample_task') or {}
        extra = request.data.get('extra') or {}
        # simple condition check
        def cond_ok():
            if trigger == 'task_status_changed':
                if condition.get('from') and extra.get('from') != condition.get('from'):
                    return False
                if condition.get('to') and extra.get('to') != condition.get('to'):
                    return False
            if trigger == 'task_due_soon':
                hours = condition.get('hours', 24)
                due = task_data.get('due') or task_data.get('end')
                if not due:
                    return False
                diff = (timezone.datetime.fromisoformat(due).timestamp() - timezone.now().timestamp()) / 3600
                if diff > hours:
                    return False
            return True

        if not cond_ok():
            return Response({"would_run": False, "reason": "Condition mismatch"}, status=200)

        simulated = {}
        if action == 'add_comment':
            simulated['comment'] = payload.get('comment') or 'Otomasyon yorumu'
        if action == 'set_assignee':
            simulated['assignee'] = payload.get('assignee')
        if action == 'add_tag':
            simulated['tag'] = payload.get('tag')
        if action == 'set_field':
            simulated['field'] = payload.get('field')
            simulated['value'] = payload.get('value')
        if action == 'notify':
            simulated['message'] = payload.get('message')
            simulated['email'] = payload.get('email')
            simulated['webhook'] = payload.get('webhook')
        if action == 'multi_notify':
            simulated['message'] = payload.get('message')
            simulated['emails'] = payload.get('emails')
            simulated['webhooks'] = payload.get('webhooks')
        return Response({"would_run": True, "action": action, "simulated": simulated}, status=200)


class UploadPresignView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsOrgMember]

    def post(self, request, *args, **kwargs):
        file_name = request.data.get('file_name') or request.data.get('filename') or 'upload.bin'
        content_type = request.data.get('content_type', 'application/octet-stream')
        size = int(request.data.get('size') or 0)
        strategy = request.data.get('strategy', 'direct')

        allowed = {'image/png', 'image/jpeg', 'image/jpg', 'image/webp', 'application/pdf'}
        max_size_mb = int(os.getenv('PRESIGN_MAX_SIZE_MB', 50 if strategy == 'chunk' else 10))

        if content_type not in allowed:
            return Response({"detail": "Yalnızca PNG/JPG/WEBP/PDF kabul edilir"}, status=status.HTTP_400_BAD_REQUEST)
        if size and size > max_size_mb * 1024 * 1024:
            return Response({"detail": f"Dosya {max_size_mb}MB üstü olamaz"}, status=status.HTTP_400_BAD_REQUEST)

        provider = os.getenv('PRESIGN_PROVIDER', 'direct').lower()
        bucket = os.getenv('PRESIGN_BUCKET') or os.getenv('AWS_STORAGE_BUCKET_NAME')
        endpoint = os.getenv('PRESIGN_ENDPOINT') or os.getenv('MINIO_ENDPOINT')
        access_key = os.getenv('PRESIGN_ACCESS_KEY') or os.getenv('MINIO_ACCESS_KEY') or os.getenv('AWS_ACCESS_KEY_ID')
        secret_key = os.getenv('PRESIGN_SECRET_KEY') or os.getenv('MINIO_SECRET_KEY') or os.getenv('AWS_SECRET_ACCESS_KEY')
        region = os.getenv('PRESIGN_REGION') or os.getenv('AWS_REGION') or 'us-east-1'
        prefix = os.getenv('PRESIGN_PREFIX', 'uploads').strip('/')

        if provider in ['s3', 'minio'] and bucket and access_key and secret_key and endpoint:
            key = f"{prefix}/tasks/{uuid.uuid4().hex}/{file_name}" if prefix else f"tasks/{uuid.uuid4().hex}/{file_name}"
            try:
                presigned = generate_presigned_post(
                    endpoint=endpoint.replace('https://', '').replace('http://', ''),
                    bucket=bucket,
                    key=key,
                    access_key=access_key,
                    secret_key=secret_key,
                    secure=endpoint.startswith('https://'),
                    region=region,
                    content_type=content_type,
                    max_size=max_size_mb * 1024 * 1024,
                )
                resp = {
                    "upload_url": presigned['url'],
                    "fields": presigned['fields'],
                    "max_size_mb": max_size_mb,
                    "content_type": content_type,
                    "filename": file_name,
                    "strategy": provider,
                    "part_size": max_size_mb * 1024 * 1024,
                    "key": key,
                }
                return Response(resp)
            except Exception as e:
                return Response({"detail": f"Presign başarısız: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Fallback: backend'e direkt yükleme (isteğe bağlı ClamAV taraması için flag)
        resp = {
            'upload_url': request.build_absolute_uri('/api/task-attachments/'),
            'fields': {},
            'max_size_mb': max_size_mb,
            'content_type': content_type,
            'filename': file_name,
            'strategy': 'direct',
            'part_size': 5 * 1024 * 1024,  # 5MB öneri
            'clamav': os.getenv('CLAMAV_ENABLED', 'false').lower() == 'true',
        }
        return Response(resp)

class TaskModelViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    """Sabit görev modelleri (AY-01 vb.) - Admin/Manager yönetir."""
    serializer_class = TaskModelSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
        'destroy': 'tasks.edit',
    }
    queryset = TaskModel.objects.all()
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['order', 'code']
    ordering = ['order', 'code']

    def get_queryset(self):
        qs = super().get_queryset()
        org = getattr(self.request.user, 'organization', None)
        if org:
            qs = qs.filter(organization=org)
        return qs

    def perform_create(self, serializer):
        org = getattr(self.request.user, 'organization', None)
        serializer.save(organization=org)


class TaskChecklistViewSet(OrgScopedMixin, viewsets.ModelViewSet):
    serializer_class = TaskChecklistSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgMember, ViewOnlyRestriction, HasAPIPermission]
    required_perm = 'tasks.view'
    permission_map = {
        'create': 'tasks.edit',
        'destroy': 'tasks.edit',
        'update': 'tasks.edit',
        'partial_update': 'tasks.edit',
        'reorder': 'tasks.edit',
    }
    queryset = TaskChecklist.objects.select_related('task')

    def get_queryset(self):
        # OrgScopedMixin filter(organization=org) yapılamıyor; doğrudan task üzerinden filtrele
        org = getattr(self.request.user, 'organization', None)
        qs = TaskChecklist.objects.select_related('task', 'workflow_team')
        if org:
            qs = qs.filter(task__organization=org)
        return qs

    def perform_create(self, serializer):
        validated = serializer.validated_data
        task = validated['task']
        extra = {}
        if validated.get('workflow_team') is None:
            req_data = getattr(self.request, 'data', {}) or {}
            if 'order' not in req_data:
                m = TaskChecklist.objects.filter(task=task).aggregate(m=Max('order'))['m']
                extra['order'] = (m + 1) if m is not None else 0
        serializer.save(**extra)

    def perform_update(self, serializer):
        inst = serializer.instance
        if inst.workflow_team_id:
            vd = serializer.validated_data
            vd.pop('done', None)
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.workflow_team_id:
            return Response(
                {'detail': 'İş akışına bağlı checklist maddesi silinemez; sıra görev akışından gelir.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'], url_path='reorder')
    def reorder(self, request):
        """Sıra güncelle: { "task": "<task_id>", "order": ["<id1>", "<id2>", ...] }"""
        task_id = request.data.get('task')
        order_ids = request.data.get('order', [])
        if not task_id or not order_ids:
            return Response({'detail': 'task ve order gerekli'}, status=status.HTTP_400_BAD_REQUEST)
        org = getattr(request.user, 'organization', None)
        items = TaskChecklist.objects.filter(task_id=task_id)
        if org:
            items = items.filter(task__organization=org)
        for idx, item_id in enumerate(order_ids):
            items.filter(id=item_id).update(order=idx)
        return Response({'ok': True})

