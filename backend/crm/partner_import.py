import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from .models import BusinessPartner


TURKEY_PROVINCES = {
    'ADANA': 'Adana',
    'ADIYAMAN': 'Adıyaman',
    'AFYON': 'Afyonkarahisar',
    'AFYONKARAHISAR': 'Afyonkarahisar',
    'AĞRI': 'Ağrı',
    'AMASYA': 'Amasya',
    'ANKARA': 'Ankara',
    'ANTALYA': 'Antalya',
    'ARTVİN': 'Artvin',
    'AYDIN': 'Aydın',
    'BALIKESİR': 'Balıkesir',
    'BİLECİK': 'Bilecik',
    'BİNGÖL': 'Bingöl',
    'BİTLİS': 'Bitlis',
    'BOLU': 'Bolu',
    'BURDUR': 'Burdur',
    'BURSA': 'Bursa',
    'ÇANAKKALE': 'Çanakkale',
    'ÇANKIRI': 'Çankırı',
    'ÇORUM': 'Çorum',
    'DENİZLİ': 'Denizli',
    'DİYARBAKIR': 'Diyarbakır',
    'DÜZCE': 'Düzce',
    'EDİRNE': 'Edirne',
    'ELAZIĞ': 'Elazığ',
    'ERZİNCAN': 'Erzincan',
    'ERZURUM': 'Erzurum',
    'ESKİŞEHİR': 'Eskişehir',
    'GAZİANTEP': 'Gaziantep',
    'GİRESUN': 'Giresun',
    'GÜMÜŞHANE': 'Gümüşhane',
    'HAKKARİ': 'Hakkari',
    'HATAY': 'Hatay',
    'IĞDIR': 'Iğdır',
    'ISPARTA': 'Isparta',
    'İSTANBUL': 'İstanbul',
    'İZMİR': 'İzmir',
    'KAHRAMANMARAŞ': 'Kahramanmaraş',
    'K.MARAŞ': 'Kahramanmaraş',
    'KARAMAN': 'Karaman',
    'KARS': 'Kars',
    'KASTAMONU': 'Kastamonu',
    'KAYSERİ': 'Kayseri',
    'KIRIKKALE': 'Kırıkkale',
    'KIRKLARELİ': 'Kırklareli',
    'KIRŞEHİR': 'Kırşehir',
    'KİLİS': 'Kilis',
    'KOCAELİ': 'Kocaeli',
    'KONYA': 'Konya',
    'KÜTAHYA': 'Kütahya',
    'MALATYA': 'Malatya',
    'MANİSA': 'Manisa',
    'MARDİN': 'Mardin',
    'MERSİN': 'Mersin',
    'MUĞLA': 'Muğla',
    'MUŞ': 'Muş',
    'NEVŞEHİR': 'Nevşehir',
    'NİĞDE': 'Niğde',
    'ORDU': 'Ordu',
    'OSMANİYE': 'Osmaniye',
    'RİZE': 'Rize',
    'SAKARYA': 'Sakarya',
    'SAMSUN': 'Samsun',
    'SİİRT': 'Siirt',
    'SİNOP': 'Sinop',
    'SİVAS': 'Sivas',
    'ŞANLIURFA': 'Şanlıurfa',
    'Ş.URFA': 'Şanlıurfa',
    'TEKİRDAĞ': 'Tekirdağ',
    'TOKAT': 'Tokat',
    'TRABZON': 'Trabzon',
    'TUNCELİ': 'Tunceli',
    'UŞAK': 'Uşak',
    'VAN': 'Van',
    'YALOVA': 'Yalova',
    'YOZGAT': 'Yozgat',
    'ZONGULDAK': 'Zonguldak',
}

COUNTRY_ALIASES = {
    'TÜRKİYE': 'Türkiye',
    'TURKIYE': 'Türkiye',
    'TURKEY': 'Türkiye',
    'IRAK': 'Irak',
    'IRAQ': 'Irak',
    'CEZAYİR': 'Cezayir',
    'FAS': 'Fas',
    'GÜRCİSTAN': 'Gürcistan',
    'KOSOVA': 'Kosova',
    'LÜBNAN': 'Lübnan',
    'LİBYA': 'Libya',
    'SUDAN': 'Sudan',
    'TÜRKMENİSTAN': 'Türkmenistan',
    'GİNE': 'Gine',
}


def _clean(value):
    if value is None:
        return ''
    text = str(value).replace('\xa0', ' ').strip()
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r'\s+', ' ', text)


def _upper_tr(value):
    return _clean(value).replace('i', 'İ').upper()


def _email_from(*values):
    for value in values:
        match = re.search(r'[\w.+-]+@[\w.-]+\.[A-Za-zÇĞİÖŞÜçğıöşü]{2,}', _clean(value))
        if match:
            return match.group(0).strip().lower()
    return ''


def _phone_part(value):
    text = _clean(value)
    if not text:
        return ''
    text = re.split(r'[;,/]| - ', text)[0].strip()
    digits = re.sub(r'\D+', '', text)
    if len(digits) < 7:
        return ''
    return digits


def _phone(mobile, area, landline):
    parts = []
    mobile_digits = _phone_part(mobile)
    if mobile_digits:
        parts.append(mobile_digits)
    landline_digits = _phone_part(landline)
    area_digits = _phone_part(area)
    if landline_digits:
        if area_digits and not landline_digits.startswith(area_digits):
            landline_digits = f'{area_digits}{landline_digits}'
        if landline_digits not in parts:
            parts.append(landline_digits)
    return ' / '.join(parts)


def _currency(value):
    normalized = _upper_tr(value)
    if normalized in {'TL', 'TRY', 'TRL'}:
        return 'TRY'
    if normalized in {'USD', 'DOLAR', '$'}:
        return 'USD'
    if normalized in {'EUR', 'EURO'}:
        return 'EUR'
    return 'TRY'


def _country_and_city(location, address):
    combined = _upper_tr(f'{location} {address}')
    country = ''
    for key, label in COUNTRY_ALIASES.items():
        if re.search(rf'(^|\s){re.escape(key)}($|\s)', combined):
            country = label
            break
    city = ''
    for key, label in TURKEY_PROVINCES.items():
        if re.search(rf'(^|[\s/,-]){re.escape(key)}($|[\s/,-])', combined):
            city = label
            break
    if city and not country:
        country = 'Türkiye'
    if not country:
        country = 'Türkiye'
    if country != 'Türkiye' and not city:
        loc = _clean(location)
        for key in COUNTRY_ALIASES:
            loc = re.sub(rf'\b{re.escape(key)}\b', '', loc, flags=re.IGNORECASE).strip()
        city = loc[:100]
    return country, city


def _address(address, postcode, location):
    parts = [_clean(address)]
    if _clean(postcode):
        parts.append(f'Posta Kodu: {_clean(postcode)}')
    if _clean(location):
        parts.append(_clean(location))
    return ' / '.join(part for part in parts if part)


def _xlsx_path(uploaded_file):
    suffix = Path(uploaded_file.name or '').suffix.lower()
    temp_dir = tempfile.TemporaryDirectory()
    source = Path(temp_dir.name) / f'import{suffix or ".xls"}'
    with source.open('wb') as handle:
        for chunk in uploaded_file.chunks():
            handle.write(chunk)
    if suffix == '.xlsx':
        return temp_dir, source
    if suffix != '.xls':
        raise ValueError('Yalnızca .xls veya .xlsx cari listesi yükleyebilirsiniz.')
    soffice = shutil.which('libreoffice') or shutil.which('soffice')
    if not soffice:
        raise ValueError('.xls dosyasını okuyabilmek için LibreOffice gereklidir.')
    result = subprocess.run(
        [soffice, '--headless', '--convert-to', 'xlsx', '--outdir', temp_dir.name, str(source)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=60,
        check=False,
    )
    converted = source.with_suffix('.xlsx')
    if result.returncode != 0 or not converted.exists():
        raise ValueError(f'Excel dosyası dönüştürülemedi: {(result.stderr or result.stdout).strip()}')
    return temp_dir, converted


def import_business_partners_from_excel(organization, uploaded_file, update_existing=True):
    temp_dir, path = _xlsx_path(uploaded_file)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        created = updated = skipped = 0
        errors = []
        rows = []
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cells = [_clean(cell) for cell in row[:18]]
            if len(cells) < 18:
                cells += [''] * (18 - len(cells))
            name = cells[1]
            if not name or name.startswith('---------------'):
                skipped += 1
                continue
            country, city = _country_and_city(cells[9] or cells[13], cells[7])
            payload = {
                'name': name[:255],
                'currency': _currency(cells[2]),
                'tax_office': cells[3][:100],
                'tax_number': re.sub(r'\D+', '', cells[4])[:64],
                'phone': _phone(cells[5], cells[10], cells[11] or cells[14])[:50],
                'email': _email_from(cells[6], cells[16], cells[15]),
                'address': _address(cells[7], cells[8] or cells[12], cells[9] or cells[13]),
                'city': city[:100],
                'country': country[:50],
                'group': '',
                'owner': '',
            }
            lookup = BusinessPartner.objects.filter(organization=organization)
            existing = None
            if payload['tax_number']:
                existing = lookup.filter(tax_number=payload['tax_number']).first()
            if existing is None:
                existing = lookup.filter(name__iexact=payload['name']).first()
            try:
                if existing:
                    if update_existing:
                        for key, value in payload.items():
                            setattr(existing, key, value)
                        existing.save()
                        updated += 1
                    else:
                        skipped += 1
                else:
                    BusinessPartner.objects.create(organization=organization, **payload)
                    created += 1
                rows.append({'row': row_number, 'name': payload['name'], 'city': payload['city'], 'country': payload['country']})
            except Exception as exc:
                errors.append({'row': row_number, 'name': name, 'detail': str(exc)})
        return {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors[:50], 'preview': rows[:20]}
    finally:
        temp_dir.cleanup()
