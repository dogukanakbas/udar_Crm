from __future__ import annotations

from copy import copy, deepcopy
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import math
import os
import re
import shutil
import subprocess
import tempfile
import unicodedata
from django.conf import settings
from accounts.price_lists import get_org_price_list_label
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
PDF_CONTENT_TYPE = 'application/pdf'
EMPTY_BORDER = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
EMPTY_FILL = PatternFill(fill_type=None)
PLACEHOLDER_PATTERN = re.compile(r'\{([A-Za-z0-9_.]+)\}')

DEFAULT_TERMS = [
    '1- Alıcı özellikleri belirtilen ürünlerin satın alma şartlarını kabul eder.',
    '2- Satıcı onaylanan kalemlerin üretimini ve sevkini kabul eder.',
    '3- Onay sonrası teknik ve finansal değişiklikler yeni mutabakat gerektirir.',
    '4- Ödemesi tamamlanmayan siparişler üretime alınmayabilir.',
    '5- Mücbir sebepler kaynaklı gecikmelerde satıcı sorumlu tutulamaz.',
    '6- Teslimden sonra yazılı bildirim gelmezse ürünler eksiksiz kabul edilir.',
    '7- Alıcı gerekli vergi ve yetki belgelerini eksiksiz sunar.',
    '8- İmalat hataları garanti koşulları ve teknik şartname kapsamındadır.',
    '9- İhtilaflarda Malatya Mahkemeleri ve İcra Daireleri yetkilidir.',
]
DEFAULT_TERMS_TEXT = '\n'.join(DEFAULT_TERMS)
DEFAULT_GENERAL_TERMS = DEFAULT_TERMS
DEFAULT_CONTRACT_NOTES = [
    'Ölçüler net ölçü üzerinden değerlendirilmiştir.',
    'Üretim öncesi teyit alınacaktır.',
]
DEFAULT_CONTRACT_NOTES_TEXT = '\n'.join(DEFAULT_CONTRACT_NOTES)

SECTION_FAMILY_MAP = {
    'steel_door': 'steel',
    'interior_door': 'interior',
    'kitchen': 'furniture',
    'wardrobe': 'furniture',
    'bathroom': 'furniture',
    'accessory': 'furniture',
    'laminate': 'furniture',
    'service': 'service',
}
SECTION_KEY_ALIASES = {
    'celik kapi': 'steel_door',
    'celik kapi grubu': 'steel_door',
    'steel door': 'steel_door',
    'ic oda kapi': 'interior_door',
    'ic oda kapisi': 'interior_door',
    'ic oda kapi grubu': 'interior_door',
    'interior door': 'interior_door',
    'mutfak': 'kitchen',
    'mutfak grubu': 'kitchen',
    'kitchen': 'kitchen',
    'portmanto': 'wardrobe',
    'portmanto grubu': 'wardrobe',
    'vestiyer': 'wardrobe',
    'wardrobe': 'wardrobe',
    'banyo': 'bathroom',
    'banyo dolaplari': 'bathroom',
    'bathroom': 'bathroom',
    'aksesuar': 'accessory',
    'aksesuarlar': 'accessory',
    'accessory': 'accessory',
    'parke': 'laminate',
    'laminat': 'laminate',
    'laminate': 'laminate',
    'hizmet': 'service',
    'hizmetler': 'service',
    'montaj': 'service',
    'hizmetler montaj': 'service',
    'hizmetler montaj grubu': 'service',
    'service': 'service',
}

DEFAULT_SELLER_PROFILES = [
    {
        'key': 'ORTKA',
        'short_name': 'ORTKA',
        'display_name': 'ORTKA YAPI ELEMANLARI ÜRETİM SAN. LTD. ŞTİ.',
        'legal_name': 'ORTKA YAPI ELEMANLARI ÜRETİM SAN. LTD. ŞTİ.',
        'tax_office': 'BEYDAĞI',
        'tax_number': '6480365207',
        'address': '2. OSB 6. CADDE NO:6 YEŞİLYURT / MALATYA',
        'city': 'Malatya',
        'country': 'Türkiye',
        'phone': '444 0 932',
        'email': 'muhasebe@aykakapi.com.tr',
        'website': '',
        'kep_address': '',
        'logo_url': '/media/seller-logos/default/ortka-logo.png',
        'signature_name': '',
        'signature_title': '',
        'signature_label': '',
        'notes': '',
        'bank_iban_label': 'Türk Lirası Hesapları',
        'bank_iban_2_label': 'Dolar Hesapları',
        'is_active': True,
        'sort_order': 0,
        'bank_accounts': [
            {'bank': 'Türkiye İş Bankası', 'iban': 'TR24 0006 4000 0018 6003 9367 45', 'currency': 'TRY'},
            {'bank': 'Ziraat Bankası', 'iban': 'TR07 0001 0021 6935 0399 4450 09', 'currency': 'TRY'},
            {'bank': 'Albaraka Türk K.Bank.', 'iban': 'TR66 0020 3000 0056 3735 0000 02', 'currency': 'TRY'},
            {'bank': 'Vakıflar Bankası', 'iban': 'TR57 0001 5001 5800 7284 2692 06', 'currency': 'TRY'},
        ],
    },
    {
        'key': 'AYKA',
        'short_name': 'AYKA',
        'display_name': 'AYKA KAPI SANAYİ TİCARET ANONİM ŞİRKETİ',
        'legal_name': 'AYKA KAPI SANAYİ TİCARET ANONİM ŞİRKETİ',
        'tax_office': 'BEYDAĞI',
        'tax_number': '1210461108',
        'address': '2. OSB 6. CADDE NO:6 YEŞİLYURT / MALATYA',
        'city': 'Malatya',
        'country': 'Türkiye',
        'phone': '444 0 932',
        'email': 'muhasebe@aykakapi.com.tr',
        'website': '',
        'kep_address': '',
        'logo_url': '/media/seller-logos/default/ayka-logo.png',
        'signature_name': '',
        'signature_title': '',
        'signature_label': '',
        'notes': '',
        'bank_iban_label': 'Türk Lirası Hesapları',
        'bank_iban_2_label': 'Dolar Hesapları',
        'is_active': True,
        'sort_order': 1,
        'bank_accounts': [
            {'bank': 'Garanti BBVA Bankası', 'iban': 'TR14 0006 2000 1120 0006 2913 29', 'currency': 'TRY'},
            {'bank': 'Ziraat Bankası', 'iban': 'TR72 0001 0021 6994 2088 8850 01', 'currency': 'TRY'},
            {'bank': 'Albaraka Türk K.Bank.', 'iban': 'TR25 0020 3000 0770 5276 0000 01', 'currency': 'TRY'},
            {'bank': 'Vakıflar Bankası', 'iban': '', 'currency': 'TRY'},
        ],
    },
]

CURRENCY_SYMBOLS = {
    'TRY': '₺',
    'USD': '$',
    'EUR': '€',
}

ASSETS_DIR = Path(__file__).resolve().parent / 'assets'
SELLER_MASTER_TEMPLATE_KEY = 'seller_master'
SELLER_MASTER_TEMPLATE_PATH = ASSETS_DIR / 'contract-templates' / 'MASTER' / 'SELLER_MASTER_TEMPLATE.xlsx'
PRODUCT_GROUP_ANCHORS = {'urunGruplari', 'kalemTablolari', 'urunTablolari'}
YEKUN_SUMMARY_GROUPS = [
    ('steel_door', 'ÇELİK KAPI GRUBU'),
    ('interior_door', 'İÇ ODA KAPI GRUBU'),
    ('kitchen', 'MUTFAK GRUBU'),
    ('wardrobe', 'PORTMANTO GRUBU'),
    ('bathroom', 'BANYO DOLAPLARI'),
    ('accessory', 'AKSESUARLAR'),
    ('laminate', 'PARKE'),
    ('service', 'HİZMETLER'),
]
SELLER_MASTER_BODY_FONT_SIZE = 12
SELLER_MASTER_TABLE_FONT_SIZE = 11
SELLER_MASTER_HEADER_FONT_SIZE = 14
DEFAULT_DYNAMIC_DOCUMENT_COLUMNS = ['Kod', 'Satış Birimi', 'Ölçü / Gövde', 'Renk / Kapak', 'Miktar', 'Liste Fiyatı', '1-İskonto %', '2-İskonto %', 'Birim', 'Birim Net Fiyatı', 'Tutar']
SPECIAL_PRODUCT_DOCUMENT_GROUPS = [
    {
        'terms': ['saft kapagi', 'saft kapak', 'rogar kapagi', 'rogar kapak', 'menhol', 'menhole'],
        'section_key': 'SAFT_ROGAR_KAPAGI',
        'label': 'Şaft / Rögar Kapağı',
        'template_family': 'saft_rogar_kapagi',
        'document_order': 20,
        'document_columns': DEFAULT_DYNAMIC_DOCUMENT_COLUMNS,
    },
]


def _document_group_order_fallback(section_key, label):
    normalized_key = _normalize_document_group_text(section_key)
    normalized_label = _normalize_document_group_text(label)
    combined = f'{normalized_key} {normalized_label}'
    if 'celik' in combined and 'kapi' in combined:
        return 10
    if ('saft' in combined or 'rogar' in combined) and ('kapak' in combined or 'kapagi' in combined):
        return 20
    if 'ic' in combined and 'oda' in combined:
        return 30
    if 'mutfak' in combined:
        return 40
    if 'mobilya' in combined or 'dolap' in combined or 'vestiyer' in combined:
        return 50
    if 'montaj' in combined or 'hizmet' in combined:
        return 90
    return 999


def _asset_path(filename: str) -> Path:
    for root in (ASSETS_DIR / 'quote-templates', ASSETS_DIR / 'contract-templates', ASSETS_DIR):
        matches = list(root.rglob(filename))
        if matches:
            return matches[0]
    raise FileNotFoundError(filename)


TEMPLATE_REGISTRY = [
    {
        'document_type': 'Quote',
        'template_key': 'quote_steel',
        'family': 'steel',
        'source_path': _asset_path('CELIK_KAPI_TEKLIF_SABLONU_V12.xltx'),
        'supported_section_keys': ['steel_door'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['steel_door'], 'start_row': 38, 'capacity': 5, 'subtotal_row': 43, 'tax_row': 44, 'grand_row': 45}],
        'commercial_rows': [61, 62, 63, 64, 65, 66, 67],
        'terms_rows': [71, 72, 73, 74, 75],
        'bank_header_row': 79,
        'bank_rows': [80, 81, 82, 83],
        'signature_row': 85,
    },
    {
        'document_type': 'Quote',
        'template_key': 'quote_interior',
        'family': 'interior',
        'source_path': _asset_path('IC_ODA_TEKLIF_SABLONU_V12.xltx'),
        'supported_section_keys': ['interior_door'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['interior_door'], 'start_row': 38, 'capacity': 7, 'subtotal_row': 45, 'tax_row': 46, 'grand_row': 47}],
        'commercial_rows': [61, 62, 63, 64, 65, 66, 67],
        'terms_rows': [71, 72, 73, 74, 75],
        'bank_header_row': 79,
        'bank_rows': [80, 81, 82, 83],
        'signature_row': 85,
    },
    {
        'document_type': 'Quote',
        'template_key': 'quote_furniture',
        'family': 'furniture',
        'source_path': _asset_path('MOBILYA_TEKLIF_SABLONU_V12.xltx'),
        'supported_section_keys': ['kitchen', 'wardrobe', 'bathroom', 'accessory', 'laminate'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [
            {'section_keys': ['kitchen'], 'start_row': 38, 'capacity': 3, 'subtotal_row': 41, 'tax_row': 42, 'grand_row': 43},
            {'section_keys': ['wardrobe'], 'start_row': 59, 'capacity': 3, 'subtotal_row': 62, 'tax_row': 63, 'grand_row': 64},
            {'section_keys': ['bathroom', 'accessory', 'laminate'], 'start_row': 71, 'capacity': 3, 'subtotal_row': 74, 'tax_row': 75, 'grand_row': 76},
        ],
        'commercial_rows': [82, 83, 84, 85, 86, 87, 88],
        'terms_rows': [92, 93, 94, 95, 96],
        'bank_header_row': 100,
        'bank_rows': [101, 102, 103, 104],
        'signature_row': 106,
    },
    {
        'document_type': 'Quote',
        'template_key': 'quote_service',
        'family': 'service',
        'source_path': _asset_path('MONTAJ_TEKLIF_SABLONU_V12.xltx'),
        'supported_section_keys': ['service'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['service'], 'start_row': 38, 'capacity': 9, 'subtotal_row': 47, 'tax_row': 48, 'grand_row': 49}],
        'commercial_rows': [63, 64, 65, 66, 67, 68, 69],
        'terms_rows': [73, 74, 75, 76, 77],
        'bank_header_row': 81,
        'bank_rows': [82, 83, 84, 85],
        'signature_row': 87,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_steel',
        'family': 'steel',
        'source_path': _asset_path('CELIK_KAPI_SOZLESME_SABLONU_V1.xltx'),
        'supported_section_keys': ['steel_door'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['steel_door'], 'start_row': 38, 'capacity': 5, 'subtotal_row': 43, 'tax_row': 44, 'grand_row': 45}],
        'commercial_rows': [59, 60, 61, 62, 63, 64, 65],
        'terms_rows': [69, 70, 71, 72, 73, 74, 75, 76, 77],
        'bank_header_row': 80,
        'bank_rows': [81, 82, 83, 84],
        'signature_row': 86,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_interior',
        'family': 'interior',
        'source_path': _asset_path('IC_KAPI_SOZLESME_SABLONU_V11.xltx'),
        'supported_section_keys': ['interior_door'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['interior_door'], 'start_row': 38, 'capacity': 7, 'subtotal_row': 45, 'tax_row': 46, 'grand_row': 47}],
        'commercial_rows': [61, 62, 63, 64, 65, 66, 67],
        'terms_rows': [71, 72, 73, 74, 75, 76, 77, 78, 79],
        'bank_header_row': 82,
        'bank_rows': [83, 84, 85, 86],
        'signature_row': 88,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_furniture',
        'family': 'furniture',
        'source_path': _asset_path('MOBILYA_SOZLESME_SABLONU_V11.xltx'),
        'supported_section_keys': ['kitchen', 'wardrobe', 'bathroom', 'accessory', 'laminate'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [
            {'section_keys': ['kitchen'], 'start_row': 38, 'capacity': 3, 'subtotal_row': 41, 'tax_row': 42, 'grand_row': 43},
            {'section_keys': ['wardrobe'], 'start_row': 56, 'capacity': 3, 'subtotal_row': 59, 'tax_row': 60, 'grand_row': 61},
            {'section_keys': ['bathroom', 'accessory', 'laminate'], 'start_row': 68, 'capacity': 3, 'subtotal_row': 71, 'tax_row': 72, 'grand_row': 73},
        ],
        'commercial_rows': [79, 80, 81, 82, 83, 84, 85],
        'terms_rows': [89, 90, 91, 92, 93, 94, 95, 96, 97],
        'bank_header_row': 102,
        'bank_rows': [103, 104, 105, 106],
        'signature_row': 108,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_service',
        'family': 'service',
        'source_path': _asset_path('MONTAJ_SOZLESME_SABLONU_V11.xltx'),
        'supported_section_keys': ['service'],
        'is_combined': False,
        'selection_priority': 10,
        'line_blocks': [{'section_keys': ['service'], 'start_row': 38, 'capacity': 9, 'subtotal_row': 47, 'tax_row': 48, 'grand_row': 49}],
        'commercial_rows': [64, 65, 66, 67, 68, 69, 70],
        'terms_rows': [73, 74, 75, 76, 77, 78, 79, 80, 81],
        'bank_header_row': 84,
        'bank_rows': [85, 86, 87, 88],
        'signature_row': 90,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_ck_ik_mob_montajli',
        'family': 'bundle',
        'source_path': _asset_path('CK_IK_MOB_MONTAJLI_SOZLESME_SABLONU_V12.xltx'),
        'supported_section_keys': ['steel_door', 'interior_door', 'kitchen', 'wardrobe', 'bathroom', 'accessory', 'laminate', 'service'],
        'required_families': ['steel', 'interior', 'furniture', 'service'],
        'is_combined': True,
        'selection_priority': 100,
        'line_blocks': [
            {'section_keys': ['steel_door'], 'start_row': 38, 'capacity': 5, 'subtotal_row': 43, 'tax_row': 44, 'grand_row': 45},
            {'section_keys': ['interior_door'], 'start_row': 63, 'capacity': 7, 'subtotal_row': 70, 'tax_row': 71, 'grand_row': 72},
            {'section_keys': ['kitchen', 'wardrobe', 'bathroom', 'accessory', 'laminate'], 'start_row': 80, 'capacity': 9, 'subtotal_row': 88, 'tax_row': 89, 'grand_row': 90},
            {'section_keys': ['service'], 'start_row': 112, 'capacity': 9, 'subtotal_row': 121, 'tax_row': 122, 'grand_row': 123},
        ],
        'commercial_rows': [127, 128, 129, 130, 131, 132],
        'terms_rows': [136, 137, 138, 139, 140, 141, 142, 143, 144],
        'bank_header_row': 147,
        'bank_rows': [148, 149, 150, 151],
        'signature_row': 153,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_ck_ik_montajli',
        'family': 'bundle',
        'source_path': _asset_path('CK_IK_MONTAJLI_SOZLESME_SABLONU_V12.xltx'),
        'supported_section_keys': ['steel_door', 'interior_door', 'service'],
        'required_families': ['steel', 'interior', 'service'],
        'is_combined': True,
        'selection_priority': 90,
        'line_blocks': [
            {'section_keys': ['steel_door'], 'start_row': 38, 'capacity': 5, 'subtotal_row': 43, 'tax_row': 44, 'grand_row': 45},
            {'section_keys': ['interior_door'], 'start_row': 63, 'capacity': 7, 'subtotal_row': 70, 'tax_row': 71, 'grand_row': 72},
            {'section_keys': ['service'], 'start_row': 80, 'capacity': 9, 'subtotal_row': 89, 'tax_row': 90, 'grand_row': 91},
        ],
        'commercial_rows': [113, 114, 115, 116, 117, 118],
        'terms_rows': [122, 123, 124, 125, 126, 127, 128, 129, 130],
        'bank_header_row': 133,
        'bank_rows': [134, 135, 136, 137],
        'signature_row': 139,
    },
    {
        'document_type': 'Contract',
        'template_key': 'contract_ck_ik_montajsiz',
        'family': 'bundle',
        'source_path': _asset_path('CK_IK_MONTAJSIZ_SOZLESME_SABLONU_V12.xltx'),
        'supported_section_keys': ['steel_door', 'interior_door'],
        'required_families': ['steel', 'interior'],
        'is_combined': True,
        'selection_priority': 80,
        'line_blocks': [
            {'section_keys': ['steel_door'], 'start_row': 38, 'capacity': 5, 'subtotal_row': 43, 'tax_row': 44, 'grand_row': 45},
            {'section_keys': ['interior_door'], 'start_row': 63, 'capacity': 7, 'subtotal_row': 70, 'tax_row': 71, 'grand_row': 72},
        ],
        'commercial_rows': [78, 79, 80, 81, 82, 83],
        'terms_rows': [87, 88, 89, 90, 91, 92, 93, 94, 95],
        'bank_header_row': 98,
        'bank_rows': [99, 100, 101, 102],
        'signature_row': 104,
    },
]

ALLOWED_TEMPLATE_EXTENSIONS = {'.xlsx', '.xltx', '.xlsm'}


def get_quote_price_list_label(quote):
    config = quote.contract_config or {}
    label = str(config.get('price_list_label') or config.get('priceListLabel') or '').strip()
    if label:
        return label
    return get_org_price_list_label(quote.organization, config.get('price_list_key') or config.get('priceListKey'))


def normalize_seller_company_key(value):
    raw = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    normalized = re.sub(r'[^A-Za-z0-9]+', '_', raw).strip('_').upper()
    return normalized[:50]


def _normalize_document_group_text(value):
    normalized_source = (
        str(value or '')
        .replace('ı', 'i')
        .replace('İ', 'I')
        .replace('ğ', 'g')
        .replace('Ğ', 'G')
        .replace('ü', 'u')
        .replace('Ü', 'U')
        .replace('ş', 's')
        .replace('Ş', 'S')
        .replace('ö', 'o')
        .replace('Ö', 'O')
        .replace('ç', 'c')
        .replace('Ç', 'C')
    )
    raw = unicodedata.normalize('NFKD', normalized_source).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', ' ', raw.lower()).strip()


def _canonical_section_key(value):
    raw = str(value or '').strip()
    if not raw:
        return ''
    if raw in SECTION_FAMILY_MAP:
        return raw
    return SECTION_KEY_ALIASES.get(_normalize_document_group_text(raw), raw)


def _turkish_upper(value):
    return str(value or '').replace('i', 'İ').replace('ı', 'I').upper()


def _special_product_document_group(product=None, line_name=''):
    sku = str(getattr(product, 'sku', '') or '')
    product_name = str(getattr(product, 'name', '') or '')
    haystack = _normalize_document_group_text(' '.join(part for part in [sku, product_name, line_name] if part))
    if not haystack:
        return None
    for group in SPECIAL_PRODUCT_DOCUMENT_GROUPS:
        if any(term in haystack for term in group['terms']):
            return group
    return None


def resolve_product_document_defaults(product=None, fallback_section_key='', line_name=''):
    product_defaults = dict(getattr(product, 'template_defaults', {}) or {}) if product else {}
    category = getattr(product, 'category', None) if product else None
    category_defaults = dict(getattr(category, 'template_defaults', {}) or {}) if category else {}
    special_group = _special_product_document_group(product, line_name)
    group_defaults = special_group or category_defaults or product_defaults
    detail_defaults = special_group or product_defaults or category_defaults
    section_key = (
        (special_group or {}).get('section_key')
        or category_defaults.get('section_key')
        or product_defaults.get('section_key')
        or fallback_section_key
        or 'service'
    )
    section_key = _canonical_section_key(section_key) or 'service'
    label = (
        (special_group or {}).get('label')
        or str(getattr(category, 'name', '') or '').strip()
        or _section_label(section_key)
        or section_key
    )
    try:
        order = int(group_defaults.get('document_order', category_defaults.get('document_order', 999)) or 999)
    except Exception:
        order = 999
    if order == 999:
        order = _document_group_order_fallback(section_key, label)
    columns = group_defaults.get('document_columns') or category_defaults.get('document_columns') or product_defaults.get('document_columns')
    if not isinstance(columns, list) or not columns:
        columns = DEFAULT_DYNAMIC_DOCUMENT_COLUMNS
    technical_items = detail_defaults.get('technical_items') or product_defaults.get('technical_items') or category_defaults.get('technical_items')
    if not isinstance(technical_items, list):
        technical_items = []
    return {
        'section_key': str(section_key or '').strip() or 'service',
        'label': str(label or '').strip() or _section_label(section_key),
        'order': order,
        'columns': [str(column or '').strip() for column in columns if str(column or '').strip()] or DEFAULT_DYNAMIC_DOCUMENT_COLUMNS,
        'technical_items': [str(item or '').strip() for item in technical_items if str(item or '').strip()],
        'template_family': group_defaults.get('template_family') or category_defaults.get('template_family') or product_defaults.get('template_family') or '',
    }


def _line_technical_items(line):
    details = dict(getattr(line, 'details', {}) or {})
    items = details.get('technicalItems') or details.get('technical_items')
    if not isinstance(items, list):
        product = getattr(line, 'product', None)
        items = resolve_product_document_defaults(product, fallback_section_key=line.section_key, line_name=line.name).get('technical_items') if product else []
    return [str(item or '').strip() for item in items if str(item or '').strip()]


def _normalize_seller_bank_account(account):
    item = dict(account or {})
    return {
        'bank': str(item.get('bank') or '').strip(),
        'iban': str(item.get('iban') or '').strip(),
        'currency': _normalize_currency_code(item.get('currency') or 'TRY'),
        'iban_2': str(item.get('iban_2') or item.get('iban2') or '').strip(),
        'currency_2': _normalize_currency_code(item.get('currency_2') or item.get('currency2') or 'USD'),
        'branch': str(item.get('branch') or '').strip(),
        'account_holder': str(item.get('account_holder') or item.get('accountHolder') or '').strip(),
    }


def _normalize_seller_profile(profile, fallback=None, sort_order=0):
    source = deepcopy(fallback or {})
    source.update(dict(profile or {}))
    key = normalize_seller_company_key(source.get('key') or source.get('short_name') or source.get('display_name'))
    return {
        'key': key,
        'short_name': str(source.get('short_name') or key).strip() or key,
        'display_name': str(source.get('display_name') or source.get('legal_name') or key).strip() or key,
        'legal_name': str(source.get('legal_name') or source.get('display_name') or key).strip() or key,
        'tax_office': str(source.get('tax_office') or '').strip(),
        'tax_number': str(source.get('tax_number') or '').strip(),
        'mersis_number': str(source.get('mersis_number') or '').strip(),
        'trade_registry_number': str(source.get('trade_registry_number') or '').strip(),
        'address': str(source.get('address') or '').strip(),
        'city': str(source.get('city') or '').strip(),
        'country': str(source.get('country') or '').strip(),
        'phone': str(source.get('phone') or '').strip(),
        'email': str(source.get('email') or '').strip(),
        'website': str(source.get('website') or '').strip(),
        'kep_address': str(source.get('kep_address') or '').strip(),
        'logo_url': str(source.get('logo_url') or '').strip(),
        'signature_name': str(source.get('signature_name') or '').strip(),
        'signature_title': str(source.get('signature_title') or '').strip(),
        'signature_label': str(source.get('signature_label') or '').strip(),
        'notes': str(source.get('notes') or '').strip(),
        'bank_iban_label': str(source.get('bank_iban_label') or source.get('bankIbanLabel') or '1. IBAN').strip() or '1. IBAN',
        'bank_iban_2_label': str(source.get('bank_iban_2_label') or source.get('bankIban2Label') or '2. IBAN').strip() or '2. IBAN',
        'is_active': bool(source.get('is_active', True)),
        'sort_order': int(source.get('sort_order', sort_order) or sort_order),
        'bank_accounts': [
            normalized
            for normalized in (_normalize_seller_bank_account(item) for item in (source.get('bank_accounts') or []))
            if normalized['bank'] or normalized['iban'] or normalized['iban_2'] or normalized['branch'] or normalized['account_holder']
        ],
    }


def _seller_bank_accounts_with_ibans(seller):
    return [
        account
        for account in seller.get('bank_accounts') or []
        if account.get('bank') or account.get('iban') or account.get('iban_2')
    ]


def _seller_has_second_iban(bank_accounts):
    return any(str((account or {}).get('iban_2') or '').strip() for account in bank_accounts or [])


def _bank_iban_display(account, slot=1, show_currency=True):
    iban = str(account.get('iban_2' if slot == 2 else 'iban') or '').strip()
    if not iban:
        return ''
    if not show_currency:
        return iban
    currency = account.get('currency_2' if slot == 2 else 'currency') or ('USD' if slot == 2 else 'TRY')
    return f'{_currency_symbol(currency)} {iban}'


def _seller_bank_iban_label(seller, slot=1):
    key = 'bank_iban_2_label' if slot == 2 else 'bank_iban_label'
    fallback = '2. IBAN' if slot == 2 else '1. IBAN'
    return str((seller or {}).get(key) or fallback).strip() or fallback


def get_default_seller_profiles():
    return [_normalize_seller_profile(profile, sort_order=index) for index, profile in enumerate(DEFAULT_SELLER_PROFILES)]


def get_seller_profiles(organization):
    settings = getattr(organization, 'contract_settings', {}) or {}
    profiles = settings.get('seller_profiles')
    defaults_by_key = {profile['key']: profile for profile in get_default_seller_profiles()}
    if isinstance(profiles, list) and profiles:
        normalized_profiles = []
        seen_keys = set()
        for index, profile in enumerate(profiles):
            fallback = defaults_by_key.get(normalize_seller_company_key((profile or {}).get('key')))
            normalized = _normalize_seller_profile(profile, fallback=fallback, sort_order=index)
            if not normalized['key'] or normalized['key'] in seen_keys:
                continue
            normalized_profiles.append(normalized)
            seen_keys.add(normalized['key'])
        if normalized_profiles:
            return sorted(normalized_profiles, key=lambda item: (item.get('sort_order', 0), item.get('short_name', item['key'])))
    return get_default_seller_profiles()


def save_seller_profiles(organization, profiles):
    normalized_profiles = [
        _normalize_seller_profile(profile, sort_order=index)
        for index, profile in enumerate(profiles or [])
        if normalize_seller_company_key((profile or {}).get('key') or (profile or {}).get('short_name') or (profile or {}).get('display_name'))
    ]
    settings = dict(getattr(organization, 'contract_settings', {}) or {})
    settings['seller_profiles'] = normalized_profiles
    organization.contract_settings = settings
    organization.save(update_fields=['contract_settings'])
    return normalized_profiles


def _template_override_settings(organization):
    settings_data = getattr(organization, 'contract_settings', {}) or {}
    overrides = settings_data.get('document_template_overrides') or {}
    return overrides if isinstance(overrides, dict) else {}


def _seller_template_override_settings(organization):
    settings_data = getattr(organization, 'contract_settings', {}) or {}
    overrides = settings_data.get('seller_document_template_overrides') or {}
    return overrides if isinstance(overrides, dict) else {}


def _template_override_entry(organization, template_key: str):
    entry = _template_override_settings(organization).get(template_key) or {}
    return entry if isinstance(entry, dict) else {}


def _seller_template_override_entry(organization, template_key: str, seller_company_key: str | None = None):
    seller_key = normalize_seller_company_key(seller_company_key)
    if not seller_key:
        return {}
    seller_overrides = _seller_template_override_settings(organization).get(seller_key) or {}
    if not isinstance(seller_overrides, dict):
        return {}
    entry = seller_overrides.get(template_key) or {}
    return entry if isinstance(entry, dict) else {}


def _seller_master_override_entry(organization, seller_company_key: str | None = None):
    return _seller_template_override_entry(organization, SELLER_MASTER_TEMPLATE_KEY, seller_company_key)


def _template_entry_path(entry):
    relative_path = str(entry.get('path') or '').strip()
    if not relative_path:
        return None
    candidate = Path(settings.MEDIA_ROOT) / relative_path
    return candidate if candidate.exists() else None


def _default_seller_master_template_path():
    if SELLER_MASTER_TEMPLATE_PATH.exists():
        return SELLER_MASTER_TEMPLATE_PATH

    target_dir = Path(settings.MEDIA_ROOT) / 'document-templates' / 'defaults'
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / 'seller-master-template.xlsx'

    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Belge'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    widths = {
        'A': 4, 'B': 15, 'C': 15, 'D': 20, 'E': 15, 'F': 15, 'G': 12,
        'H': 12, 'I': 12, 'J': 14, 'K': 15, 'L': 18, 'M': 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in range(1, 70):
        ws.row_dimensions[row].height = 20
    for row in [2, 3, 4, 5, 6, 7, 8]:
        ws.row_dimensions[row].height = 23
    for row in [11, 16, 23, 31, 37, 45, 52, 60]:
        ws.row_dimensions[row].height = 24

    title_fill = PatternFill('solid', fgColor='203864')
    soft_fill = PatternFill('solid', fgColor='EAF2F8')
    white_font = Font(color='FFFFFF', bold=True)
    label_font = Font(color='203864', bold=True)
    blue_font = Font(color='000066', bold=True)
    thin = Side(style='thin', color='000000')
    medium = Side(style='medium', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def merge_range(range_ref, value='', fill=None, font=None, alignment=None, outline=False, draw_border=True):
        ws.merge_cells(range_ref)
        cell = ws[range_ref.split(':', 1)[0]]
        cell.value = value
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if draw_border:
            for row in ws[range_ref]:
                for item in row:
                    item.border = section_border if outline else border
        return cell

    def section_header(row_number, title):
        merge_range(f'B{row_number}:M{row_number}', title, fill=title_fill, font=white_font, alignment=center, outline=True)

    def label_value(row_number, label, value):
        merge_range(f'B{row_number}:C{row_number}', label, fill=soft_fill, font=label_font, alignment=center, outline=True)
        merge_range(f'D{row_number}:M{row_number}', value, font=Font(color='000000'), alignment=left, outline=True)

    # Ust kisim mevcut teklif/sözlesme dosyalarina yakin kalir; logo sadece secili saticidan basilir.
    merge_range('B2:E5', '{saticiLogo}', font=Font(color='7A8798', bold=True, italic=True, size=12), alignment=center, draw_border=False)
    merge_range('J2:M3', '{seciliSatici.resmiUnvan}', font=Font(color='000066', bold=True, size=11), alignment=center, draw_border=False)
    merge_range('J4:M4', '{seciliSatici.kisaAd}', font=Font(color='000066', bold=True, size=10), alignment=center, draw_border=False)
    merge_range('J5:M5', '{seciliSatici.email} / {seciliSatici.telefon}', font=Font(color='000066', bold=True, size=8), alignment=center, draw_border=False)

    section_header(7, '{belgeTuru}')
    merge_range('K8:L8', 'TARİH', font=blue_font, alignment=center, outline=True)
    merge_range('M8:M8', '{olusturmaTarihi}', font=Font(color='000066', bold=True), alignment=center, outline=True)

    ws.merge_cells('A9:A26')
    ws['A9'] = 'T\nA\nR\nA\nF\nL\nA\nR'
    ws['A9'].font = Font(color='000066', bold=True, size=18)
    ws['A9'].alignment = center

    section_header(9, 'SATIŞI YAPAN FİRMA')
    label_value(10, 'ÜNVAN', '{seciliSatici.resmiUnvan}')
    label_value(11, 'VERGİ DAİRESİ / NO', '{seciliSatici.vergiDairesi} / {seciliSatici.vergiNo}')
    label_value(12, 'ADRES', '{seciliSatici.adres}')
    label_value(13, 'TELEFON / E-POSTA', '{seciliSatici.telefon} / {seciliSatici.email}')

    section_header(15, 'ALICI BİLGİLERİ')
    label_value(16, 'CARİ ÜNVANI', '{cariUnvani}')
    label_value(17, 'VERGİ DAİRESİ / NO', '{vergiDairesi} / {vergiNo}')
    label_value(18, 'ADRES', '{adres}')
    label_value(19, 'YETKİLİ', '{yetkili}')
    label_value(20, 'TELEFON / E-POSTA', '{telefon} / {email}')

    section_header(22, '')
    label_value(23, '{belgeTuru} NUMARASI', '{belgeNo}')
    label_value(24, 'HAZIRLAYAN', '{hazirlayan}')
    label_value(25, 'GEÇERLİLİK TARİHİ', '{gecerlilikTarihi}')
    label_value(26, 'F.LİSTESİ', '{fiyatListesiEtiketi} - Para Birimi: {paraBirimi}')

    section_header(28, 'KALEMLER')
    ws['B29'] = '{urunGruplari}'
    ws['B29'].alignment = left

    section_header(45, 'ÖDEME VE TESLİM')
    label_value(46, 'ÖDEME KOŞULU', '{odemeKosulu}')
    label_value(47, 'TESLİM TİPİ', '{teslimTipi}')
    label_value(48, 'TESLİM TARİHİ', '{teslimTarihi}')
    label_value(49, 'NOTLAR', '{notlar}')

    section_header(52, 'FİRMA ÜNVANI & IBANLAR')
    label_value(53, '{seciliSatici.banka1.ad}', '{seciliSatici.banka1.iban}')
    label_value(54, '{seciliSatici.banka2.ad}', '{seciliSatici.banka2.iban}')
    label_value(55, '{seciliSatici.banka3.ad}', '{seciliSatici.banka3.iban}')
    label_value(56, '{seciliSatici.banka4.ad}', '{seciliSatici.banka4.iban}')

    section_header(60, 'TARAFLARIN KAŞE İMZASI')
    merge_range('B61:G66', 'SATICI\n{seciliSatici.resmiUnvan}', alignment=center, outline=True)
    merge_range('H61:M66', 'ALICI\n{cariUnvani}', alignment=center, outline=True)

    workbook.save(target_path)
    return target_path


def _seller_master_template_path(organization, seller_company_key: str | None = None, include_default=True):
    custom_path = _template_entry_path(_seller_master_override_entry(organization, seller_company_key))
    if custom_path:
        return custom_path
    return _default_seller_master_template_path() if include_default else None


def _seller_master_library_entry(organization, profile):
    seller_key = profile['key']
    entry = _seller_master_override_entry(organization, seller_key)
    custom_path = _template_entry_path(entry)
    current_path = custom_path or _default_seller_master_template_path()
    return {
        'template_key': SELLER_MASTER_TEMPLATE_KEY,
        'seller_company_key': seller_key,
        'seller_short_name': profile.get('short_name') or seller_key,
        'seller_display_name': _normalize_display_name(profile.get('display_name') or profile.get('legal_name') or seller_key),
        'default_filename': _default_seller_master_template_path().name,
        'current_filename': current_path.name,
        'has_custom': bool(custom_path),
        'source_type': 'custom' if custom_path else 'default',
        'uploaded_at': entry.get('uploaded_at'),
        'uploaded_by': entry.get('uploaded_by'),
    }


def _custom_template_path(organization, template_key: str, seller_company_key: str | None = None):
    seller_path = _template_entry_path(_seller_template_override_entry(organization, template_key, seller_company_key))
    if seller_path:
        return seller_path
    return _template_entry_path(_template_override_entry(organization, template_key))


def _template_source_path(organization, template, seller_company_key: str | None = None):
    return _custom_template_path(organization, template['template_key'], seller_company_key) or template['source_path']


def _template_label(template):
    document_label = 'Teklif' if template['document_type'] == 'Quote' else 'Sözleşme'
    family_labels = {
        'quote_steel': 'Çelik grubu',
        'quote_interior': 'İç oda grubu',
        'quote_furniture': 'Mobilya grubu',
        'quote_service': 'Montaj grubu',
        'contract_steel': 'Çelik grubu',
        'contract_interior': 'İç oda grubu',
        'contract_furniture': 'Mobilya grubu',
        'contract_service': 'Montaj grubu',
        'contract_ck_ik_mob_montajli': 'CK + İK + MOB + Montaj',
        'contract_ck_ik_montajli': 'CK + İK + Montaj',
        'contract_ck_ik_montajsiz': 'CK + İK',
    }
    return f"{document_label} / {family_labels.get(template['template_key'], template['template_key'])}"


def _template_library_entry(organization, template, seller_company_key: str | None = None):
    seller_key = normalize_seller_company_key(seller_company_key)
    seller_entry = _seller_template_override_entry(organization, template['template_key'], seller_key)
    shared_entry = _template_override_entry(organization, template['template_key'])
    seller_path = _template_entry_path(seller_entry)
    shared_path = _template_entry_path(shared_entry)
    if seller_key:
        override_entry = seller_entry if seller_path else {}
        current_path = seller_path or shared_path or template['source_path']
        has_custom = bool(seller_path)
        uses_shared_custom = bool(shared_path and not seller_path)
    else:
        override_entry = shared_entry
        current_path = shared_path or template['source_path']
        has_custom = bool(shared_path)
        uses_shared_custom = False
    entry = {
        'template_key': template['template_key'],
        'document_type': template['document_type'],
        'label': _template_label(template),
        'default_filename': template['source_path'].name,
        'current_filename': current_path.name,
        'has_custom': has_custom,
        'source_type': 'custom' if has_custom else 'default',
        'uses_shared_custom': uses_shared_custom,
        'uploaded_at': override_entry.get('uploaded_at'),
        'uploaded_by': override_entry.get('uploaded_by'),
    }
    if seller_key:
        entry['seller_company_key'] = seller_key
    return entry


def list_template_library(organization):
    templates = []
    for template in TEMPLATE_REGISTRY:
        templates.append(_template_library_entry(organization, template))
    seller_templates = [
        {
            'seller_company_key': profile['key'],
            'seller_short_name': profile.get('short_name') or profile['key'],
            'seller_display_name': _normalize_display_name(profile.get('display_name') or profile.get('legal_name') or profile['key']),
            'master_template': _seller_master_library_entry(organization, profile),
        }
        for profile in get_seller_profiles(organization)
        if profile.get('is_active', True)
    ]
    return {'templates': templates, 'seller_templates': seller_templates}


def get_template_download(organization, template_key: str, variant: str = 'current', seller_company_key: str | None = None):
    seller_key = normalize_seller_company_key(seller_company_key)
    if not template_key or template_key == SELLER_MASTER_TEMPLATE_KEY:
        if seller_key and seller_key not in {profile['key'] for profile in get_seller_profiles(organization) if profile.get('is_active', True)}:
            raise ValueError('Satıcı firma bulunamadı')
        source_path = _default_seller_master_template_path() if variant == 'default' else _seller_master_template_path(organization, seller_key)
        if not source_path.exists():
            raise ValueError('Şablon dosyası bulunamadı')
        prefix = f'{seller_key}-' if seller_key else ''
        return {
            'path': source_path,
            'filename': f'{prefix}firma-sablonu{source_path.suffix or ".xlsx"}',
        }

    template = next((item for item in TEMPLATE_REGISTRY if item['template_key'] == template_key), None)
    if not template:
        raise ValueError('Şablon bulunamadı')
    source_path = template['source_path'] if variant == 'default' else _template_source_path(organization, template, seller_key)
    if not source_path.exists():
        raise ValueError('Şablon dosyası bulunamadı')
    suffix = source_path.suffix or '.xlsx'
    prefix = f'{seller_key}-' if seller_key else ''
    return {
        'path': source_path,
        'filename': f"{prefix}{template_key}-sablon{suffix}",
    }


def save_template_override(organization, template_key: str, uploaded_file, user=None, seller_company_key: str | None = None):
    template_key = str(template_key or SELLER_MASTER_TEMPLATE_KEY).strip() or SELLER_MASTER_TEMPLATE_KEY
    seller_key = normalize_seller_company_key(seller_company_key)
    is_seller_master = template_key == SELLER_MASTER_TEMPLATE_KEY
    if is_seller_master and not seller_key:
        raise ValueError('Satıcı firma zorunludur')
    if seller_key:
        active_seller_keys = {profile['key'] for profile in get_seller_profiles(organization) if profile.get('is_active', True)}
        if seller_key not in active_seller_keys:
            raise ValueError('Satıcı firma bulunamadı')
    if not is_seller_master:
        template = next((item for item in TEMPLATE_REGISTRY if item['template_key'] == template_key), None)
        if not template:
            raise ValueError('Şablon bulunamadı')

    extension = Path(getattr(uploaded_file, 'name', '') or '').suffix.lower()
    if extension not in ALLOWED_TEMPLATE_EXTENSIONS:
        raise ValueError('Yalnızca .xlsx, .xltx veya .xlsm dosyaları yüklenebilir')

    content = uploaded_file.read()
    if not content:
        raise ValueError('Yüklenen dosya boş görünüyor')

    try:
        load_workbook(BytesIO(content))
    except Exception as exc:
        raise ValueError('Excel şablonu okunamadı') from exc

    target_dir = Path(settings.MEDIA_ROOT) / 'document-templates' / f'org_{organization.id}'
    if seller_key:
        target_dir = target_dir / seller_key
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f'{template_key}{extension}'
    target_path.write_bytes(content)

    settings_data = dict(getattr(organization, 'contract_settings', {}) or {})
    override_entry = {
        'path': target_path.relative_to(settings.MEDIA_ROOT).as_posix(),
        'original_name': getattr(uploaded_file, 'name', target_path.name),
        'uploaded_at': datetime.now().isoformat(),
        'uploaded_by': getattr(user, 'id', None),
    }
    if seller_key:
        seller_overrides = dict(settings_data.get('seller_document_template_overrides') or {})
        seller_template_overrides = dict(seller_overrides.get(seller_key) or {})
        seller_template_overrides[template_key] = override_entry
        seller_overrides[seller_key] = seller_template_overrides
        settings_data['seller_document_template_overrides'] = seller_overrides
    else:
        overrides = dict(settings_data.get('document_template_overrides') or {})
        overrides[template_key] = override_entry
        settings_data['document_template_overrides'] = overrides
    organization.contract_settings = settings_data
    organization.save(update_fields=['contract_settings'])
    library = list_template_library(organization)
    if is_seller_master:
        for seller_group in library.get('seller_templates', []):
            if seller_group.get('seller_company_key') == seller_key:
                return seller_group.get('master_template')
    if seller_key:
        for seller_group in library.get('seller_templates', []):
            if seller_group.get('seller_company_key') == seller_key:
                return next((item for item in seller_group.get('templates', []) if item['template_key'] == template_key), None)
    return next((item for item in library.get('templates', []) if item['template_key'] == template_key), None)


def list_template_placeholders():
    return [
        {
            'group': 'Belge bilgileri',
            'description': 'Belgenin temel kimliği, tarihleri ve finansal özet alanları.',
            'items': [
                {'token': '{belgeNo}', 'label': 'Belge numarası'},
                {'token': '{belgeTuru}', 'label': 'Belge türü'},
                {'token': '{olusturmaTarihi}', 'label': 'Oluşturulma tarihi'},
                {'token': '{guncellenmeTarihi}', 'label': 'Güncellenme tarihi'},
                {'token': '{gecerlilikTarihi}', 'label': 'Geçerlilik tarihi'},
                {'token': '{teslimTarihi}', 'label': 'Teslim tarihi'},
                {'token': '{odemeKosulu}', 'label': 'Ödeme koşulu'},
                {'token': '{hazirlayan}', 'label': 'Hazırlayan kullanıcı'},
                {'token': '{saticiLogo}', 'label': 'Seçili satıcı firma logosu'},
                {'token': '{fiyatListesiEtiketi}', 'label': 'Merkezi fiyat listesi etiketi'},
                {'token': '{paraBirimi}', 'label': 'Para birimi sembolü'},
                {'token': '{paraBirimiKodu}', 'label': 'Para birimi kodu'},
                {'token': '{kur}', 'label': 'Kur bilgisi'},
                {'token': '{araToplam}', 'label': 'Ara toplam'},
                {'token': '{iskontoToplami}', 'label': 'Toplam iskonto'},
                {'token': '{kdvToplami}', 'label': 'KDV toplamı'},
                {'token': '{genelToplam}', 'label': 'Genel toplam'},
                {'token': '{yekun}', 'label': 'Yekün'},
                {'token': '{kdvOrani}', 'label': 'KDV oranı'},
                {'token': '{notlar}', 'label': 'Belge notları'},
            ],
        },
        {
            'group': 'Cari bilgileri',
            'description': 'Seçilen cari / müşteri kartından gelen bilgiler.',
            'items': [
                {'token': '{cariUnvani}', 'label': 'Cari ünvanı'},
                {'token': '{cariAdi}', 'label': 'Cari adı'},
                {'token': '{yetkili}', 'label': 'Yetkili kişi'},
                {'token': '{vergiDairesi}', 'label': 'Vergi dairesi'},
                {'token': '{vergiNo}', 'label': 'Vergi numarası'},
                {'token': '{adres}', 'label': 'Adres'},
                {'token': '{telefon}', 'label': 'Telefon'},
                {'token': '{email}', 'label': 'E-posta'},
                {'token': '{sehir}', 'label': 'Şehir'},
                {'token': '{ulke}', 'label': 'Ülke'},
                {'token': '{cariParaBirimi}', 'label': 'Cari para birimi'},
                {'token': '{cariOlcegi}', 'label': 'Cari ölçeği'},
            ],
        },
        {
            'group': 'Satıcı firmalar',
            'description': 'Yönetim panelindeki satıcı firmalar sıra numarasıyla doldurulur. Örnek: saticiFirma1, saticiFirma2.',
            'items': [
                {'token': '{seciliSatici.unvan}', 'label': 'Belgede seçili satıcı firma ünvanı'},
                {'token': '{seciliSatici.vergiDairesi}', 'label': 'Belgede seçili satıcı firma vergi dairesi'},
                {'token': '{seciliSatici.vergiNo}', 'label': 'Belgede seçili satıcı firma vergi numarası'},
                {'token': '{seciliSatici.adres}', 'label': 'Belgede seçili satıcı firma adresi'},
                {'token': '{seciliSatici.telefon}', 'label': 'Belgede seçili satıcı firma telefonu'},
                {'token': '{seciliSatici.email}', 'label': 'Belgede seçili satıcı firma e-postası'},
                {'token': '{saticiFirma1.unvan}', 'label': '1. satıcı firma ünvanı'},
                {'token': '{saticiFirma1.banka1.ad}', 'label': '1. satıcı firmanın 1. banka adı'},
                {'token': '{saticiFirma1.banka1.iban}', 'label': '1. satıcı firmanın 1. IBAN bilgisi'},
                {'token': '{saticiFirma1.banka1.paraBirimi}', 'label': '1. satıcı firmanın 1. IBAN para birimi'},
                {'token': '{saticiFirma1.banka1.iban2}', 'label': '1. satıcı firmanın aynı bankadaki 2. IBAN bilgisi'},
                {'token': '{saticiFirma1.banka1.paraBirimi2}', 'label': '1. satıcı firmanın aynı bankadaki 2. IBAN para birimi'},
                {'token': '{saticiFirma2.unvan}', 'label': '2. satıcı firma ünvanı'},
                {'token': '{saticiFirma2.banka1.ad}', 'label': '2. satıcı firmanın 1. banka adı'},
                {'token': '{saticiFirma2.banka1.iban}', 'label': '2. satıcı firmanın 1. IBAN bilgisi'},
                {'token': '{saticiFirma2.banka1.paraBirimi}', 'label': '2. satıcı firmanın 1. IBAN para birimi'},
                {'token': '{saticiFirma2.banka1.iban2}', 'label': '2. satıcı firmanın aynı bankadaki 2. IBAN bilgisi'},
                {'token': '{saticiFirma2.banka1.paraBirimi2}', 'label': '2. satıcı firmanın aynı bankadaki 2. IBAN para birimi'},
            ],
        },
        {
            'group': 'Ürün Grupları',
            'description': 'Kalem bazlı alanlar sıra numarasıyla kullanılır. Örnek: kalem1, kalem2.',
            'items': [
                {'token': '{kalem1.kod}', 'label': '1. kalem ürün kodu'},
                {'token': '{kalem1.urun}', 'label': '1. kalem ürün adı'},
                {'token': '{kalem1.miktar}', 'label': '1. kalem miktarı'},
                {'token': '{kalem1.birim}', 'label': '1. kalem birimi'},
                {'token': '{kalem1.birimFiyat}', 'label': '1. kalem liste fiyatı'},
                {'token': '{kalem1.netBirimFiyat}', 'label': '1. kalem net birim fiyatı'},
                {'token': '{kalem1.tutar}', 'label': '1. kalem toplam tutarı'},
                {'token': '{kalem1.olcu}', 'label': '1. kalem ölçü / detay 1'},
                {'token': '{kalem1.renk}', 'label': '1. kalem renk / detay 2'},
                {'token': '{kalem1.iskonto1}', 'label': '1. kalem 1. iskonto'},
                {'token': '{kalem1.iskonto2}', 'label': '1. kalem 2. iskonto'},
                {'token': '{kalem1.kdvOrani}', 'label': '1. kalem KDV oranı'},
            ],
        },
    ]


def list_document_exports(quote):
    return [
        {
            'template_key': SELLER_MASTER_TEMPLATE_KEY,
            'filename': f"{quote.number}-firma-sablonu.xlsx",
        }
    ]


def build_document_export(quote, template_key: str | None = None):
    requested_key = str(template_key or SELLER_MASTER_TEMPLATE_KEY).strip() or SELLER_MASTER_TEMPLATE_KEY
    if requested_key == SELLER_MASTER_TEMPLATE_KEY:
        return _build_seller_master_document_export(quote)

    outputs = []
    for template in _select_templates(quote, template_key=requested_key):
        quote._current_export_template = template
        workbook = load_workbook(_template_source_path(quote.organization, template, quote.seller_company_key))
        workbook.template = False
        sheet_name = str(template.get('sheet_name') or '').strip()
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else (workbook['1'] if '1' in workbook.sheetnames else workbook[workbook.sheetnames[0]])
        template_layout, layout_expanded = _expand_template_for_line_counts(worksheet, template, quote)
        template_layout['expanded_layout'] = layout_expanded
        if layout_expanded:
            normalized_stream = BytesIO()
            workbook.save(normalized_stream)
            normalized_stream.seek(0)
            workbook = load_workbook(normalized_stream)
            workbook.template = False
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else (workbook['1'] if '1' in workbook.sheetnames else workbook[workbook.sheetnames[0]])
        _fill_shared_header(worksheet, quote)
        _fill_line_blocks(worksheet, quote, template_layout)
        _fill_commercial_rows(worksheet, quote, template_layout)
        _fill_terms(worksheet, quote, template_layout)
        _fill_bank_accounts(worksheet, quote, template_layout)
        _fill_signature_block(worksheet, quote, template_layout)
        _apply_template_placeholders(worksheet, quote, template_layout)
        output = BytesIO()
        workbook.save(output)
        outputs.append((template, output.getvalue()))

    if not outputs:
        raise ValueError('No template output generated')
    if len(outputs) > 1:
        raise ValueError('Multiple templates selected; request a specific template_key')

    template, content = outputs[0]
    stream = BytesIO(content)
    stream.seek(0)
    return {
        'content': stream,
        'filename': f"{quote.number}-{template['template_key']}.xlsx",
        'content_type': XLSX_CONTENT_TYPE,
    }


def _build_seller_master_document_export(quote):
    template_path = _seller_master_template_path(quote.organization, quote.seller_company_key)
    workbook = load_workbook(template_path)
    workbook.template = False
    worksheet = workbook['Belge'] if 'Belge' in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    banner_images = _detach_template_banner_images(worksheet)
    template = {
        'template_key': SELLER_MASTER_TEMPLATE_KEY,
        'document_type': quote.document_type,
        'seller_master_layout': True,
    }
    quote._current_export_template = template
    _apply_template_placeholders(worksheet, quote, template)
    _apply_seller_master_header_branding(worksheet, quote)
    tail_row = _render_dynamic_product_group_tables(worksheet, quote)
    _render_seller_master_tail(worksheet, quote, tail_row, banner_images=banner_images)
    _apply_times_new_roman_font(worksheet)
    _prepare_pdf_print_layout(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return {
        'content': output,
        'filename': f"{quote.number}-firma-sablonu.xlsx",
        'content_type': XLSX_CONTENT_TYPE,
    }


def _prepare_pdf_print_layout(worksheet):
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = worksheet.sheet_properties.pageSetUpPr or PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.2
    worksheet.page_margins.bottom = 0.2
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'


def build_document_pdf_export(quote):
    xlsx_export = build_document_export(quote, template_key=SELLER_MASTER_TEMPLATE_KEY)
    pdf_stream = _convert_xlsx_stream_to_pdf(xlsx_export['content'], xlsx_export['filename'])
    return {
        'content': pdf_stream,
        'filename': f'{quote.number}.pdf',
        'content_type': PDF_CONTENT_TYPE,
    }


def _convert_xlsx_stream_to_pdf(xlsx_stream, xlsx_filename):
    converter = shutil.which('libreoffice') or shutil.which('soffice')
    if not converter:
        raise ValueError('PDF oluşturmak için LibreOffice bulunamadı.')

    safe_filename = re.sub(r'[^A-Za-z0-9_.-]+', '_', xlsx_filename or 'document.xlsx')
    if not safe_filename.lower().endswith('.xlsx'):
        safe_filename = f'{safe_filename}.xlsx'

    xlsx_stream.seek(0)
    with tempfile.TemporaryDirectory(prefix='udar_pdf_') as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_path = tmp_path / safe_filename
        output_path = input_path.with_suffix('.pdf')
        input_path.write_bytes(xlsx_stream.read())

        result = subprocess.run(
            [
                converter,
                '--headless',
                '--nologo',
                '--nofirststartwizard',
                '--convert-to',
                'pdf',
                '--outdir',
                str(tmp_path),
                str(input_path),
            ],
            cwd=str(tmp_path),
            env={**os.environ, 'HOME': str(tmp_path)},
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=90,
            check=False,
        )
        if result.returncode != 0 or not output_path.exists():
            detail = (result.stderr or result.stdout or '').strip()
            raise ValueError(f'PDF oluşturulamadı: {detail or "LibreOffice dönüştürme hatası"}')

        pdf_stream = BytesIO(output_path.read_bytes())
        pdf_stream.seek(0)
        return pdf_stream


def _build_reportlab_document_pdf_export(quote):
    from html import escape
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Image as PdfImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    def register_font(name, candidates):
        for candidate in candidates:
            path = Path(candidate)
            if path.exists():
                try:
                    pdfmetrics.registerFont(TTFont(name, str(path)))
                    return name
                except Exception:
                    continue
        return 'Helvetica'

    base_font = register_font('UdarSans', [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        'C:/Windows/Fonts/arial.ttf',
    ])
    bold_font = register_font('UdarSansBold', [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        'C:/Windows/Fonts/arialbd.ttf',
    ])
    if bold_font == 'Helvetica':
        bold_font = 'Helvetica-Bold'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=quote.number,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle('UdarNormal', parent=styles['Normal'], fontName=base_font, fontSize=8, leading=10)
    small = ParagraphStyle('UdarSmall', parent=normal, fontSize=7, leading=9)
    title_style = ParagraphStyle('UdarTitle', parent=normal, fontName=bold_font, fontSize=13, alignment=1, textColor=colors.white)
    heading = ParagraphStyle('UdarHeading', parent=normal, fontName=bold_font, fontSize=10, textColor=colors.HexColor('#203864'))

    def p(value, style=normal):
        return Paragraph(escape(str(value or '')).replace('\n', '<br/>'), style)

    def money(value):
        symbol = _currency_symbol(_quote_currency(quote))
        amount = Decimal(value or 0)
        return f'{symbol} {amount:,.2f}'

    def basic_table(data, widths=None, header_rows=0):
        table = Table(data, colWidths=widths, repeatRows=header_rows)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#9FB2C8')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EAF2F8')) if header_rows else ('TEXTCOLOR', (0, 0), (0, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), bold_font) if header_rows else ('TEXTCOLOR', (0, 0), (0, 0), colors.black),
        ]))
        return table

    seller = _selected_seller_profile(quote)
    customer = _customer_snapshot(quote.contract_config or {})
    logo_path = _seller_logo_media_path(seller)
    logo = ''
    if logo_path:
        try:
            logo = PdfImage(str(logo_path), width=48 * mm, height=22 * mm, kind='proportional')
        except Exception:
            logo = ''

    story = []
    story.append(Table(
        [[logo, p(f"{_normalize_display_name(seller.get('display_name', ''))}\n{seller.get('email', '')} / {seller.get('phone', '')}", normal)]],
        colWidths=[85 * mm, 170 * mm],
    ))
    story.append(Spacer(1, 4))
    document_title = 'SÖZLEŞME' if quote.document_type == 'Contract' else 'TEKLİF'
    title_table = Table([[p(document_title, title_style)]], colWidths=[255 * mm])
    title_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#203864'))]))
    story.append(title_table)
    story.append(Spacer(1, 6))

    doc_date = _timezone_fallback(quote.created_at).strftime('%d.%m.%Y')
    story.append(basic_table([
        [p('TARİH', heading), p(doc_date), p(f'{document_title} NUMARASI', heading), p(quote.number)],
        [p('HAZIRLAYAN', heading), p(_resolve_prepared_by_name(quote)), p('GEÇERLİLİK TARİHİ', heading), p(format_validity_text(quote.valid_until))],
    ], widths=[35 * mm, 92 * mm, 42 * mm, 86 * mm]))
    story.append(Spacer(1, 6))

    story.append(p('SATIŞI YAPAN FİRMA', heading))
    story.append(basic_table([
        [p('ÜNVAN', heading), p(_normalize_display_name(seller.get('display_name', '')))],
        [p('VERGİ DAİRESİ / NO', heading), p(_join_tax(seller.get('tax_office', ''), seller.get('tax_number', '')))],
        [p('ADRES', heading), p(seller.get('address', ''))],
        [p('TELEFON / E-POSTA', heading), p(_join_contact(seller.get('phone', ''), seller.get('email', '')))],
    ], widths=[45 * mm, 210 * mm]))
    story.append(Spacer(1, 6))

    story.append(p('ALICI BİLGİLERİ', heading))
    story.append(basic_table([
        [p('CARİ ÜNVANI', heading), p(customer.get('name') or getattr(quote.customer, 'name', ''))],
        [p('VERGİ DAİRESİ / NO', heading), p(_join_tax(customer.get('tax_office'), customer.get('tax_number')))],
        [p('ADRES', heading), p(customer.get('address') or '')],
        [p('YETKİLİ', heading), p(customer.get('authorized_person') or '')],
        [p('TELEFON / E-POSTA', heading), p(_join_contact(customer.get('phone'), customer.get('email')))],
    ], widths=[45 * mm, 210 * mm]))
    story.append(Spacer(1, 8))

    groups = _build_dynamic_line_groups(quote)
    for group in groups:
        rows = [[p('Kod', heading), p('Ürün', heading), p('Miktar', heading), p('Birim', heading), p('Liste Fiyatı', heading), p('Birim Net', heading), p('Tutar', heading)]]
        quantity_total = Decimal('0')
        for line in group['lines']:
            details = dict(line.details or {})
            product = getattr(line, 'product', None)
            qty = Decimal(line.qty or 0)
            quantity_total += qty
            rows.append([
                p(details.get('code') or getattr(product, 'sku', '') or '', small),
                p(line.name or '', small),
                p(f'{qty:,.2f}', small),
                p(line.unit or 'Adet', small),
                p(money(line.unit_price), small),
                p(money(_net_unit_price(line)), small),
                p(money(_line_subtotal(line)), small),
            ])
        rows.append(['', '', p(f'{quantity_total:,.2f}', small), '', '', p('Yekün', small), p(money(sum((_line_subtotal(line) + _line_tax(line) for line in group['lines']), Decimal('0'))), small)])
        story.append(p(_turkish_upper(group['label']), heading))
        story.append(basic_table(rows, widths=[28 * mm, 78 * mm, 22 * mm, 20 * mm, 33 * mm, 33 * mm, 33 * mm], header_rows=1))
        story.append(Spacer(1, 6))

    service_rows = [[p('Kod', heading), p('Ürün Adı', heading), p('Miktar', heading), p('Ara Toplam', heading), p('K.D.V. %', heading), p('K.D.V.', heading), p('Yekün', heading)]]
    subtotal_total = Decimal('0')
    tax_total = Decimal('0')
    grand_total = Decimal('0')
    qty_total = Decimal('0')
    for group in groups:
        for line in group['lines']:
            details = dict(line.details or {})
            product = getattr(line, 'product', None)
            subtotal = _line_subtotal(line)
            tax = _line_tax(line)
            tax_rate = Decimal(getattr(line, 'tax', 0) or 0)
            qty = Decimal(line.qty or 0)
            subtotal_total += subtotal
            tax_total += tax
            grand_total += subtotal + tax
            qty_total += qty
            service_rows.append([p(details.get('code') or getattr(product, 'sku', '') or '', small), p(line.name or '', small), p(f'{qty:,.2f}', small), p(money(subtotal), small), p(f'%{tax_rate:,.2f}', small), p(money(tax), small), p(money(subtotal + tax), small)])
    service_rows.append([p('TOPLAM', small), '', p(f'{qty_total:,.2f}', small), p(money(subtotal_total), small), '', p(money(tax_total), small), p(money(grand_total), small)])
    story.append(p('HİZMETLER', heading))
    story.append(basic_table(service_rows, widths=[26 * mm, 78 * mm, 22 * mm, 32 * mm, 22 * mm, 30 * mm, 32 * mm], header_rows=1))
    story.append(Spacer(1, 8))

    config = quote.contract_config or {}
    story.append(p('ÖDEME VE TESLİM', heading))
    story.append(basic_table([
        [p('ÖDEME KOŞULU', heading), p(quote.payment_terms or config.get('paymentOption') or config.get('payment_option') or '')],
        [p('TESLİM TİPİ', heading), p(config.get('deliveryType') or config.get('delivery_type') or '')],
        [p('TESLİM TARİHİ', heading), p(_format_date_text(quote.delivery_terms))],
    ], widths=[45 * mm, 210 * mm]))
    story.append(Spacer(1, 6))

    terms = parse_terms_text(config.get('termsText') or config.get('terms_text') or DEFAULT_TERMS_TEXT)
    if quote.document_type == 'Contract':
        terms += parse_contract_notes_text(config.get('contractNotesText') or config.get('contract_notes_text') or '')
    if terms:
        story.append(p('SÖZLEŞME KOŞULLARI' if quote.document_type == 'Contract' else 'TEKLİF KOŞULLARI', heading))
        story.append(basic_table([[p(term, small)] for term in terms], widths=[255 * mm]))
        story.append(Spacer(1, 6))

    bank_accounts = _seller_bank_accounts_with_ibans(seller)
    has_second_iban = _seller_has_second_iban(bank_accounts)
    if has_second_iban:
        bank_rows = [[p('Banka', heading), p(_seller_bank_iban_label(seller, 1), heading), p(_seller_bank_iban_label(seller, 2), heading)]]
        bank_widths = [55 * mm, 100 * mm, 100 * mm]
    else:
        bank_rows = [[p('Banka', heading), p('IBAN', heading)]]
        bank_widths = [55 * mm, 200 * mm]
    for account in bank_accounts or [{}]:
        row = [
            p(_normalize_bank_name(account.get('bank', '')), small),
            p(_bank_iban_display(account, 1, show_currency=has_second_iban), small),
        ]
        if has_second_iban:
            row.append(p(_bank_iban_display(account, 2), small))
        bank_rows.append(row)
    story.append(p('FİRMA ÜNVANI & IBANLAR', heading))
    story.append(basic_table(bank_rows, widths=bank_widths, header_rows=1))
    story.append(Spacer(1, 8))

    story.append(basic_table([[p(f"SATICI\n{_normalize_display_name(seller.get('display_name', ''))}", normal), p(f"ALICI\n{customer.get('name') or getattr(quote.customer, 'name', '') or ''}", normal)]], widths=[127 * mm, 128 * mm]))

    doc.build(story)
    buffer.seek(0)
    return {
        'content': buffer,
        'filename': f'{quote.number}.pdf',
        'content_type': PDF_CONTENT_TYPE,
    }


def _find_product_group_anchor(ws):
    for row in ws.iter_rows():
        for cell in row:
            raw = str(cell.value or '').strip()
            if not raw.startswith('{') or not raw.endswith('}'):
                continue
            token = raw[1:-1].strip()
            if token in PRODUCT_GROUP_ANCHORS:
                return cell.row, cell.column
    return None


def _apply_logo_placeholders(ws, quote):
    seller = _selected_seller_profile(quote)
    logo_path = _seller_logo_media_path(seller)
    if not logo_path:
        return
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').strip() != '{saticiLogo}':
                continue
            cell.value = ''
            _clear_logo_frame(ws, cell)
            try:
                image = XLImage(str(logo_path))
            except Exception:
                continue
            image = _fit_excel_image(image, 260, 86)
            ws.add_image(image, cell.coordinate)


def _fixed_ayka_seller_profile(organization):
    profiles = [profile for profile in get_seller_profiles(organization) if profile.get('is_active', True)] or get_default_seller_profiles()
    ayka_key = normalize_seller_company_key('AYKA')
    for profile in profiles:
        if normalize_seller_company_key(profile.get('key')) == ayka_key:
            return profile
    return next((profile for profile in get_default_seller_profiles() if profile.get('key') == ayka_key), profiles[0] if profiles else {})


def _set_header_text(ws, coordinate, value, *, bold=False, size=10):
    cell = _resolve_cell(ws, coordinate)
    cell.value = value
    cell.font = _seller_master_font(color='000066', bold=bold, size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)


def _add_header_logo(ws, profile, anchor, max_width, max_height):
    logo_path = _seller_logo_media_path(profile)
    if not logo_path:
        return
    try:
        image = XLImage(str(logo_path))
    except Exception:
        return
    image = _crop_excel_image_content(image, padding=0)
    image = _fit_excel_image(image, max_width, max_height)
    ws.add_image(image, anchor)


def _merge_header_range(ws, cell_range):
    _safe_unmerge(ws, cell_range)
    ws.merge_cells(cell_range)


def _prepare_seller_master_header_layout(ws):
    for cell_range in ['B2:E5', 'F2:I5', 'J2:M3', 'J2:M2', 'J3:M3', 'J4:M4', 'J5:M5']:
        _safe_unmerge(ws, cell_range)
    _merge_header_range(ws, 'B2:E5')
    _merge_header_range(ws, 'F2:I5')
    for cell_range in ['J2:M2', 'J3:M3', 'J4:M4', 'J5:M5']:
        _merge_header_range(ws, cell_range)
    for row in range(2, 6):
        ws.row_dimensions[row].height = 22


def _apply_seller_master_header_branding(ws, quote):
    fixed_left = _fixed_ayka_seller_profile(quote.organization)
    selected_seller = _selected_seller_profile(quote)
    _strip_header_logos(ws)
    _prepare_seller_master_header_layout(ws)

    for coordinate in ['B2', 'F2', 'J2', 'J3', 'J4', 'J5']:
        _set_cell_value(ws, coordinate, '')

    _clear_logo_frame(ws, ws['B2'])
    _clear_logo_frame(ws, ws['F2'])
    _add_header_logo(ws, fixed_left, 'B2', 330, 112)

    _set_header_text(ws, 'J2', _normalize_display_name(selected_seller.get('display_name', '')), bold=True, size=9)
    _set_header_text(ws, 'J3', _join_tax(selected_seller.get('tax_office', ''), selected_seller.get('tax_number', '')), size=8)
    _set_header_text(ws, 'J4', selected_seller.get('address', ''), size=8)
    _set_header_text(ws, 'J5', _join_contact(selected_seller.get('phone', ''), selected_seller.get('email', '')), size=8)


def _image_anchor_row(image):
    anchor = getattr(image, 'anchor', None)
    marker = getattr(anchor, '_from', None)
    if marker is None:
        return None
    try:
        return int(marker.row) + 1
    except (TypeError, ValueError):
        return None


def _detach_template_banner_images(ws):
    banners = []
    kept = []
    for image in list(getattr(ws, '_images', []) or []):
        anchor_row = _image_anchor_row(image)
        if anchor_row and anchor_row >= 30:
            banners.append(image)
        else:
            kept.append(image)
    ws._images = kept
    return banners


def _render_dynamic_product_group_tables(ws, quote):
    anchor = _find_product_group_anchor(ws)
    start_row, start_column = anchor or (ws.max_row + 2, 1)
    groups = _build_dynamic_line_groups(quote)
    required_rows = _dynamic_tables_row_count(quote, groups)
    table_width = _dynamic_table_physical_width(groups)
    end_column = start_column + table_width - 1
    if anchor and required_rows > 1:
        ws.insert_rows(start_row + 1, required_rows - 1)

    _unmerge_overlapping_range(
        ws,
        start_row,
        start_row + max(required_rows, 1) - 1,
        start_column,
        end_column,
    )
    for row in range(start_row, start_row + max(required_rows, 1)):
        for column in range(start_column, max(end_column, ws.max_column) + 1):
            _set_cell_value(ws, f'{get_column_letter(column)}{row}', '')

    if not groups:
        cell = ws.cell(start_row, start_column)
        cell.value = 'Belgeye eklenmiş ürün kalemi yok.'
        cell.font = Font(italic=True, color='666666')
        return start_row + 2

    current_row = start_row
    for group in groups:
        current_row = _write_dynamic_product_group(ws, quote, group, current_row, start_column, table_width)
        current_row += 1
    current_row = _write_service_summary_group(ws, quote, groups, current_row, start_column, table_width)
    return current_row + 1


def _dynamic_tables_row_count(quote, groups):
    if not groups:
        return 1
    total = 0
    for group in groups:
        technical_count = sum((len(_line_technical_items(line)) for line in group['lines']), 0)
        total += 2 + len(group['lines']) + 1 + 3 + technical_count + 1
    return total + _service_summary_row_count(quote, groups)


def _service_summary_row_count(quote, groups):
    if not groups:
        return 0
    return len(_service_expense_rows(quote, groups)) + 3


def _dynamic_table_physical_width(groups):
    logical_width = max((len(group.get('columns') or []) for group in groups), default=len(DEFAULT_DYNAMIC_DOCUMENT_COLUMNS))
    return max(logical_width, 12)


def _dynamic_column_spans(columns, physical_width):
    spans = [1 for _ in columns]
    extra = max(physical_width - len(columns), 0)
    priority_terms = ('urun', 'hizmet', 'satis', 'sat', 'ad', 'olcu', 'renk')
    normalized = [_normalize_column_key(column) for column in columns]
    for term in priority_terms:
        if extra <= 0:
            break
        for index, key in enumerate(normalized):
            if extra <= 0:
                break
            if term in key:
                spans[index] += 1
                extra -= 1
    index = 0
    while extra > 0 and spans:
        spans[index % len(spans)] += 1
        index += 1
        extra -= 1
    return spans


def _dynamic_span_width(ws, start_column, span):
    width = Decimal('0')
    for physical_column in range(start_column, start_column + span):
        letter = get_column_letter(physical_column)
        width += Decimal(str(ws.column_dimensions[letter].width or 10))
    return max(width, Decimal('8'))


def _dynamic_text_row_height(value, width_units, base=18, line_height=13, max_height=96):
    text = str(value or '').replace('\r\n', '\n').replace('\r', '\n')
    if not text:
        return base
    usable_chars = max(8, int(float(width_units) * 1.05))
    visual_lines = 0
    for raw_line in text.split('\n'):
        visual_lines += max(1, (len(raw_line) + usable_chars - 1) // usable_chars)
    return min(max_height, max(base, base + (visual_lines - 1) * line_height))


def _seller_master_font(color='000000', bold=False, italic=False, size=SELLER_MASTER_BODY_FONT_SIZE):
    return Font(name='Times New Roman', color=color, bold=bold, italic=italic, size=size)


def _dynamic_set_row_height(ws, row, height):
    current = ws.row_dimensions[row].height or 0
    ws.row_dimensions[row].height = max(current, height)


def _dynamic_is_numeric_column(header):
    key = _normalize_column_key(header)
    return any(term in key for term in ('miktar', 'adet', 'fiyat', 'tutar', 'iskonto', 'kdv', 'toplam', 'yekun'))


def _dynamic_product_group_column_alignment(columns, index):
    normalized = [_normalize_column_key(column) for column in columns]
    try:
        sales_unit_index = next(
            idx
            for idx, key in enumerate(normalized)
            if 'satisbirimi' in key or 'satbirimi' in key
        )
        net_unit_index = next(
            idx
            for idx, key in enumerate(normalized)
            if 'net' in key and ('birim' in key or 'fiyat' in key)
        )
    except StopIteration:
        header = columns[index] if 0 <= index < len(columns) else ''
        return 'right' if _dynamic_is_numeric_column(header) else 'left'

    if sales_unit_index < index <= net_unit_index:
        return 'center'

    header = columns[index] if 0 <= index < len(columns) else ''
    return 'right' if _dynamic_is_numeric_column(header) else 'left'


def _format_quantity(value):
    amount = Decimal(value or 0)
    if amount == amount.to_integral_value():
        return f'{int(amount):,}'.replace(',', '.')
    normalized = amount.normalize()
    text = f'{normalized:,.2f}'.rstrip('0').rstrip('.')
    return text.replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_dynamic_line_groups(quote):
    grouped = {}
    for line in quote.lines.select_related('product__category').order_by('sort_order', 'id'):
        if not _line_is_meaningful(line):
            continue
        group = _line_dynamic_group(line)
        key = group['key']
        if key not in grouped:
            grouped[key] = group
        grouped[key]['lines'].append(line)
    return sorted(grouped.values(), key=lambda item: (item['order'], item['label']))


def _line_dynamic_group(line):
    product = getattr(line, 'product', None)
    resolved = resolve_product_document_defaults(product, fallback_section_key=line.section_key, line_name=line.name)
    section_key = resolved['section_key']
    return {
        'key': section_key,
        'label': resolved['label'],
        'order': resolved['order'],
        'columns': resolved['columns'],
        'lines': [],
    }


def _section_label(section_key):
    labels = {
        'steel_door': 'Çelik Kapı',
        'interior_door': 'İç Oda Kapısı',
        'kitchen': 'Mutfak',
        'wardrobe': 'Vestiyer / Dolap',
        'bathroom': 'Banyo',
        'accessory': 'Aksesuar',
        'laminate': 'Laminat',
        'service': 'Hizmetler & Montaj',
    }
    return labels.get(section_key, section_key.replace('_', ' ').title())


def _write_dynamic_product_group(ws, quote, group, row, column, physical_width=None):
    columns = group['columns']
    column_count = max(physical_width or len(columns), len(columns), 4)
    last_column = column + column_count - 1
    spans = _dynamic_column_spans(columns, column_count)
    currency_code = _quote_currency(quote)
    dark_fill = PatternFill('solid', fgColor='203864')
    header_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='9FB2C8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    _unmerge_overlapping_range(ws, row, row, column, last_column)
    ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=last_column)
    title = ws.cell(row, column)
    title.value = _turkish_upper(group['label'])
    title.fill = dark_fill
    title.font = _seller_master_font(color='FFFFFF', bold=True, size=SELLER_MASTER_HEADER_FONT_SIZE)
    title.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(column, last_column + 1):
        ws.cell(row, col).border = border
    _dynamic_set_row_height(ws, row, 26)
    row += 1

    current_column = column
    for offset, header in enumerate(columns):
        span = spans[offset]
        if span > 1:
            _unmerge_overlapping_range(ws, row, row, current_column, current_column + span - 1)
            ws.merge_cells(start_row=row, start_column=current_column, end_row=row, end_column=current_column + span - 1)
        cell = ws.cell(row, current_column)
        cell.value = _turkish_upper(header)
        cell.fill = header_fill
        cell.font = _seller_master_font(color='203864', bold=True, size=SELLER_MASTER_BODY_FONT_SIZE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for physical_column in range(current_column, current_column + span):
            ws.cell(row, physical_column).fill = header_fill
            ws.cell(row, physical_column).border = border
        current_column += span
    _dynamic_set_row_height(ws, row, 26)
    row += 1

    for line in group['lines']:
        current_column = column
        required_height = 18
        for offset, header in enumerate(columns):
            span = spans[offset]
            if span > 1:
                _unmerge_overlapping_range(ws, row, row, current_column, current_column + span - 1)
                ws.merge_cells(start_row=row, start_column=current_column, end_row=row, end_column=current_column + span - 1)
            cell = ws.cell(row, current_column)
            value, value_type = _dynamic_column_value(line, header)
            if value_type == 'currency':
                cell.value = float(value) if value not in [None, ''] else ''
                cell.number_format = _currency_number_format(currency_code)
            else:
                cell.value = value
            cell.font = _seller_master_font(size=SELLER_MASTER_TABLE_FONT_SIZE)
            for physical_column in range(current_column, current_column + span):
                ws.cell(row, physical_column).border = border
            horizontal_alignment = _dynamic_product_group_column_alignment(columns, offset)
            cell.alignment = Alignment(
                horizontal=horizontal_alignment,
                vertical='center',
                wrap_text=True,
                shrink_to_fit=_dynamic_is_numeric_column(header) or horizontal_alignment == 'center',
            )
            if not _dynamic_is_numeric_column(header):
                required_height = max(required_height, _dynamic_text_row_height(value, _dynamic_span_width(ws, current_column, span), base=20, line_height=14))
            current_column += span
        _dynamic_set_row_height(ws, row, required_height)
        row += 1

    subtotal = sum((_line_subtotal(line) for line in group['lines']), Decimal('0'))
    tax = sum((_line_tax(line) for line in group['lines']), Decimal('0'))
    grand = subtotal + tax
    quantity_total = sum((Decimal(line.qty or 0) for line in group['lines']), Decimal('0'))
    summary_values = [
        ('Toplam Ürün:', _format_quantity(quantity_total), False),
        ('Ara Toplam', subtotal, True),
        (_summary_tax_label(group['lines']), tax, True),
        ('Yekün', grand, True),
    ]
    for label, amount, is_currency in summary_values:
        label_column = max(column, last_column - 1)
        left_end_column = max(column, label_column - 1)
        _unmerge_overlapping_range(ws, row, row, column, left_end_column)
        if left_end_column > column:
            ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=left_end_column)
        for col in range(column, left_end_column + 1):
            ws.cell(row, col).border = EMPTY_BORDER
        left_cell = ws.cell(row, column)
        left_cell.value = ''
        left_cell.border = EMPTY_BORDER
        left_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        label_cell = ws.cell(row, label_column)
        label_cell.value = _turkish_upper(label)
        label_cell.font = _seller_master_font(color='203864', bold=True, size=SELLER_MASTER_TABLE_FONT_SIZE)
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        amount_cell = ws.cell(row, last_column)
        amount_cell.value = float(amount) if is_currency else str(amount)
        amount_cell.number_format = _currency_number_format(currency_code) if is_currency else '@'
        amount_cell.font = _seller_master_font(bold=True, size=SELLER_MASTER_TABLE_FONT_SIZE)
        amount_cell.border = border
        amount_cell.alignment = Alignment(horizontal='right', vertical='center', shrink_to_fit=True)
        _dynamic_set_row_height(ws, row, _dynamic_text_row_height(label, _dynamic_span_width(ws, max(column, last_column - 1), 1), base=18, max_height=48))
        row += 1

    for line in group['lines']:
        for item in _line_technical_items(line):
            _unmerge_overlapping_range(ws, row, row, column, last_column)
            ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=last_column)
            cell = ws.cell(row, column)
            cell.value = f'* {line.name}: {item}'
            cell.font = _seller_master_font(color='203864', italic=True, size=SELLER_MASTER_TABLE_FONT_SIZE)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True)
            ws.row_dimensions[row].height = 18
            row += 1

    return row


def _service_expense_rows(quote, groups=None):
    def normalize_label(value):
        return _normalize_document_group_text(value or '')

    def line_service_key(line):
        product = getattr(line, 'product', None)
        product_id = str(getattr(product, 'id', '') or '').strip()
        if product_id:
            return f'product:{product_id}'
        details = getattr(line, 'details', None) or {}
        code = str(details.get('code') or getattr(line, 'sku', '') or '').strip()
        if code:
            return f'code:{code}'
        name = str(getattr(line, 'name', '') or '').strip().lower()
        return f'name:{name}' if name else ''

    def line_service_label(line):
        product = getattr(line, 'product', None)
        product_name = str(getattr(product, 'name', '') or getattr(line, 'name', '') or '').strip()
        category_name = str(getattr(getattr(product, 'category', None), 'name', '') or _section_label(getattr(line, 'section_key', '') or '') or '').strip()
        if product_name and normalize_label(product_name) != normalize_label(category_name):
            return product_name
        details = getattr(line, 'details', None) or {}
        code = str(details.get('code') or getattr(product, 'sku', '') or getattr(line, 'sku', '') or '').strip()
        primary = str(details.get('primary') or '').strip()
        secondary = str(details.get('secondary') or '').strip()
        detailed_name = ' - '.join([part for part in (code, primary, secondary) if part])
        return detailed_name or product_name or code or 'Ürün'

    line_labels_by_key = {}
    for group in groups or _build_dynamic_line_groups(quote):
        if str(group.get('key') or '').strip() == 'service':
            continue
        for line in group.get('lines') or []:
            key = line_service_key(line)
            if key and key not in line_labels_by_key:
                line_labels_by_key[key] = line_service_label(line)

    configured = []
    for item in (quote.contract_config or {}).get('service_expenses') or []:
        if not isinstance(item, dict):
            continue
        category_key = str(item.get('category_key') or item.get('categoryKey') or '').strip()
        label = str(item.get('category_label') or item.get('categoryLabel') or '').strip()
        if category_key in line_labels_by_key:
            label = line_labels_by_key[category_key]
        if not label:
            continue
        try:
            amount = Decimal(str(item.get('amount') or 0))
        except (InvalidOperation, TypeError, ValueError):
            amount = Decimal('0')
        try:
            quantity = Decimal(str(item.get('quantity') if item.get('quantity') not in [None, ''] else 1))
        except (InvalidOperation, TypeError, ValueError):
            quantity = Decimal('1')
        try:
            unit_amount = Decimal(str(item.get('unit_amount') if item.get('unit_amount') not in [None, ''] else item.get('unitAmount', '')))
        except (InvalidOperation, TypeError, ValueError):
            unit_amount = Decimal('-1')
        quantity = max(quantity, Decimal('0'))
        amount = max(amount, Decimal('0'))
        if unit_amount >= Decimal('0'):
            amount = max(unit_amount, Decimal('0')) * quantity
        try:
            tax_rate = Decimal(str(item.get('tax') or item.get('tax_rate') or item.get('taxRate') or 0))
        except (InvalidOperation, TypeError, ValueError):
            tax_rate = Decimal('0')
        tax_rate = max(tax_rate, Decimal('0'))
        configured.append({'label': label, 'quantity': quantity, 'amount': amount, 'tax_rate': tax_rate, 'tax': amount * (tax_rate / Decimal('100'))})
    if configured:
        return configured

    rows_by_key = {}
    for group in groups or []:
        if str(group.get('key') or '').strip() == 'service':
            continue
        for line in group.get('lines') or []:
            key = line_service_key(line)
            if not key:
                continue
            label = line_service_label(line)
            quantity = Decimal(getattr(line, 'qty', 0) or 0)
            if key not in rows_by_key:
                rows_by_key[key] = {'label': label, 'quantity': Decimal('0'), 'amount': Decimal('0'), 'tax_rate': Decimal('0'), 'tax': Decimal('0')}
            rows_by_key[key]['quantity'] += quantity
    return list(rows_by_key.values())


def _write_service_summary_group(ws, quote, groups, row, column, physical_width):
    column_count = max(physical_width, 6)
    last_column = column + column_count - 1
    currency_code = _quote_currency(quote)
    dark_fill = PatternFill('solid', fgColor='203864')
    header_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='9FB2C8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ['Ürün Adı', 'Miktar', 'Ara Toplam', 'K.D.V. %', 'K.D.V.', 'Yekün']
    spans = _dynamic_column_spans(headers, column_count)

    _unmerge_overlapping_range(ws, row, row, column, last_column)
    ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=last_column)
    title = ws.cell(row, column)
    title.value = 'HİZMETLER'
    title.fill = dark_fill
    title.font = _seller_master_font(color='FFFFFF', bold=True, size=SELLER_MASTER_HEADER_FONT_SIZE)
    title.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(column, last_column + 1):
        ws.cell(row, col).border = border
    _dynamic_set_row_height(ws, row, 26)
    row += 1

    current_column = column
    for index, header in enumerate(headers):
        span = spans[index]
        if span > 1:
            _unmerge_overlapping_range(ws, row, row, current_column, current_column + span - 1)
            ws.merge_cells(start_row=row, start_column=current_column, end_row=row, end_column=current_column + span - 1)
        cell = ws.cell(row, current_column)
        cell.value = _turkish_upper(header)
        cell.fill = header_fill
        cell.font = _seller_master_font(color='203864', bold=True, size=SELLER_MASTER_BODY_FONT_SIZE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for physical_column in range(current_column, current_column + span):
            ws.cell(row, physical_column).fill = header_fill
            ws.cell(row, physical_column).border = border
        current_column += span
    _dynamic_set_row_height(ws, row, 26)
    row += 1

    subtotal_total = Decimal('0')
    tax_total = Decimal('0')
    grand_total = Decimal('0')
    quantity_total = Decimal('0')
    for item in _service_expense_rows(quote, groups):
        subtotal = item['amount']
        tax = item['tax']
        grand = subtotal + tax
        subtotal_total += subtotal
        tax_total += tax
        grand_total += grand
        quantity = Decimal(item.get('quantity') or 0)
        tax_rate = Decimal(item.get('tax_rate') or 0)
        quantity_total += quantity
        row = _write_service_summary_row(ws, row, column, spans, [
            item['label'],
            quantity,
            subtotal,
            tax_rate,
            tax,
            grand,
        ], currency_code, border)

    row = _write_service_summary_row(ws, row, column, spans, [
        'TOPLAM',
        quantity_total,
        subtotal_total,
        '',
        tax_total,
        grand_total,
    ], currency_code, border, bold=True)
    return row


def _write_service_summary_row(ws, row, column, spans, values, currency_code, border, bold=False):
    current_column = column
    required_height = 18
    for index, value in enumerate(values):
        span = spans[index]
        if span > 1:
            _unmerge_overlapping_range(ws, row, row, current_column, current_column + span - 1)
            ws.merge_cells(start_row=row, start_column=current_column, end_row=row, end_column=current_column + span - 1)
        cell = ws.cell(row, current_column)
        cell.value = float(value) if isinstance(value, Decimal) else value
        if isinstance(value, Decimal):
            cell.number_format = _currency_number_format(currency_code)
            if index == 1:
                cell.number_format = '#,##0.##'
            elif index == 3:
                cell.number_format = '0.##"%"'
        cell.font = _seller_master_font(color='203864' if bold else '000000', bold=bold, size=SELLER_MASTER_TABLE_FONT_SIZE)
        is_amount_column = index >= 2
        is_quantity_column = index == 1
        cell.alignment = Alignment(
            horizontal='center' if is_quantity_column else ('right' if is_amount_column else 'left'),
            vertical='center',
            wrap_text=True,
            shrink_to_fit=is_amount_column,
        )
        for physical_column in range(current_column, current_column + span):
            ws.cell(row, physical_column).border = border
        if not is_amount_column:
            required_height = max(required_height, _dynamic_text_row_height(value, _dynamic_span_width(ws, current_column, span), base=20, line_height=14, max_height=84))
        current_column += span
    _dynamic_set_row_height(ws, row, required_height)
    return row + 1


def _render_seller_master_tail(ws, quote, start_row, banner_images=None):
    if start_row <= ws.max_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    _reset_seller_master_columns(ws)

    yekun_end = _write_yekun_summary_tail(ws, quote, start_row, start_col=2, end_col=7)
    payment_end = _write_payment_delivery_tail(ws, quote, start_row, start_col=8, end_col=13)

    terms_start = max(yekun_end, payment_end) + 1
    row = _write_terms_tail(ws, quote, terms_start)

    bank_start = row + 1
    row = _write_bank_tail(ws, quote, bank_start)

    signature_start = row + 2
    row = _write_signature_tail(ws, quote, signature_start)

    prepared_banner = _prepare_bottom_banner_tail(ws, banner_images or [])
    banner_start = _bottom_banner_footer_start_row(ws, bank_start, row + 1, prepared_banner)
    _write_bottom_banner_tail(ws, banner_start, prepared_banner)

    _add_manual_page_break_before(ws, bank_start)


def _reset_seller_master_columns(ws):
    widths = {
        'B': 13, 'C': 14, 'D': 17, 'E': 13, 'F': 13, 'G': 11,
        'H': 11, 'I': 11, 'J': 13, 'K': 14, 'L': 15, 'M': 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _tail_styles():
    title_fill = PatternFill('solid', fgColor='203864')
    soft_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='000000')
    medium = Side(style='medium', color='000000')
    return {
        'title_fill': title_fill,
        'soft_fill': soft_fill,
        'white_font': _seller_master_font(color='FFFFFF', bold=True, size=SELLER_MASTER_HEADER_FONT_SIZE),
        'label_font': _seller_master_font(color='203864', bold=True, size=SELLER_MASTER_BODY_FONT_SIZE),
        'text_font': _seller_master_font(size=SELLER_MASTER_BODY_FONT_SIZE),
        'thin_border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'section_border': Border(left=medium, right=medium, top=medium, bottom=medium),
        'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
    }


def _tail_merge(ws, row, start_col, end_col, value='', fill=None, font=None, alignment=None, border=None):
    _unmerge_overlapping_range(ws, row, row, start_col, end_col)
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col)
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        for column in range(start_col, end_col + 1):
            ws.cell(row, column).border = border
    return cell


def _tail_section_header(ws, row, title, start_col=2, end_col=13):
    styles = _tail_styles()
    _tail_merge(ws, row, start_col, end_col, _turkish_upper(title), styles['title_fill'], styles['white_font'], styles['center'], styles['section_border'])
    ws.row_dimensions[row].height = 28


def _tail_label_value(ws, row, label, value, start_col=2, end_col=13):
    styles = _tail_styles()
    width = max(2, end_col - start_col + 1)
    label_width = 3 if width >= 6 else max(1, width // 2)
    label_end_col = min(end_col - 1, start_col + label_width - 1)
    value_start_col = min(end_col, label_end_col + 1)
    _tail_merge(ws, row, start_col, label_end_col, _turkish_upper(label), styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
    _tail_merge(ws, row, value_start_col, end_col, value, None, styles['text_font'], styles['left'], styles['section_border'])
    ws.row_dimensions[row].height = 28


def _yekun_summary_label(group):
    key = str(group.get('key') or '').strip()
    label = str(group.get('label') or '').strip() or _section_label(key)
    if key == 'service':
        return 'HİZMETLER'
    normalized_label = _normalize_document_group_text(label)
    if normalized_label.endswith('grubu'):
        return _turkish_upper(label)
    return _turkish_upper(f'{label} Grubu')


def _add_yekun_summary_row(rows_by_key, key, label, amount):
    amount = Decimal(amount or 0)
    if amount == 0:
        return
    normalized_key = str(key or '').strip() or _normalize_document_group_text(label)
    if normalized_key in rows_by_key:
        rows_by_key[normalized_key]['amount'] += amount
        return
    rows_by_key[normalized_key] = {'label': label, 'amount': amount}


def _build_yekun_summary_rows(quote):
    rows_by_key = {}
    for group in _build_dynamic_line_groups(quote):
        amount = sum((_line_subtotal(line) + _line_tax(line) for line in group['lines']), Decimal('0'))
        _add_yekun_summary_row(rows_by_key, group.get('key'), _yekun_summary_label(group), amount)

    service_expense_total = sum((item['amount'] + item['tax'] for item in _service_expense_rows(quote)), Decimal('0'))
    _add_yekun_summary_row(rows_by_key, 'service', 'HİZMETLER', service_expense_total)
    return list(rows_by_key.values())


def _write_yekun_summary_tail(ws, quote, row, start_col=2, end_col=13):
    styles = _tail_styles()
    currency_code = _quote_currency(quote)
    summary_rows = _build_yekun_summary_rows(quote)
    label_end_col = min(end_col - 1, start_col + 2)
    amount_start_col = min(end_col, label_end_col + 1)

    _tail_section_header(ws, row, 'YEKÜN İCMAALLERİ', start_col, end_col)
    current_row = row + 1
    for summary in summary_rows:
        label = summary['label']
        amount = summary['amount']
        _tail_merge(ws, current_row, start_col, label_end_col, label, styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
        amount_cell = _tail_merge(ws, current_row, amount_start_col, end_col, float(amount), None, styles['text_font'], styles['left'], styles['section_border'])
        amount_cell.number_format = _currency_number_format(currency_code)
        amount_cell.alignment = Alignment(horizontal='right', vertical='center', shrink_to_fit=True)
        ws.row_dimensions[current_row].height = 26
        current_row += 1

    grand_total = sum((summary['amount'] for summary in summary_rows), Decimal('0'))
    _tail_merge(ws, current_row, start_col, label_end_col, 'TOPLAM YEKÜN', styles['title_fill'], styles['white_font'], styles['center'], styles['section_border'])
    total_cell = _tail_merge(ws, current_row, amount_start_col, end_col, float(grand_total), None, _seller_master_font(color='203864', bold=True, size=SELLER_MASTER_BODY_FONT_SIZE), styles['left'], styles['section_border'])
    total_cell.number_format = _currency_number_format(currency_code)
    total_cell.alignment = Alignment(horizontal='right', vertical='center', shrink_to_fit=True)
    ws.row_dimensions[current_row].height = 28
    return current_row + 1


def _config_text(config, *keys):
    for key in keys:
        value = config.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ''


def _write_payment_delivery_tail(ws, quote, row, start_col=2, end_col=13):
    config = quote.contract_config or {}
    payment_value = quote.payment_terms or _config_text(config, 'paymentOption', 'payment_option')
    delivery_type = _config_text(config, 'deliveryType', 'delivery_type')
    delivery_date = _format_date_text(quote.delivery_terms or _config_text(config, 'deliveryDate', 'delivery_date'))
    notes = quote.notes or _config_text(config, 'notes')

    _tail_section_header(ws, row, 'ÖDEME VE TESLİM', start_col, end_col)
    _tail_label_value(ws, row + 1, 'ÖDEME KOŞULU', payment_value, start_col, end_col)
    _tail_label_value(ws, row + 2, 'TESLİM TİPİ', delivery_type, start_col, end_col)
    _tail_label_value(ws, row + 3, 'TESLİM TARİHİ', delivery_date, start_col, end_col)
    if notes:
        _tail_label_value(ws, row + 4, 'NOTLAR', notes, start_col, end_col)
        return row + 5
    return row + 4


def _write_terms_tail(ws, quote, row):
    config = quote.contract_config or {}
    raw_terms = _config_text(config, 'termsText', 'terms_text') or DEFAULT_TERMS_TEXT
    terms = parse_terms_text(raw_terms)
    if quote.document_type == 'Contract':
        notes = parse_contract_notes_text(_config_text(config, 'contractNotesText', 'contract_notes_text'))
        terms = terms + notes

    title = 'SÖZLEŞME KOŞULLARI' if quote.document_type == 'Contract' else 'TEKLİF KOŞULLARI'
    _tail_section_header(ws, row, title)
    styles = _tail_styles()
    current_row = row + 1
    for term in terms:
        _tail_merge(ws, current_row, 2, 13, term, None, styles['text_font'], styles['left'], styles['thin_border'])
        ws.row_dimensions[current_row].height = max(26, min(72, 22 + (len(str(term)) // 90) * 15))
        current_row += 1
    if quote.document_type == 'Contract':
        contract_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)
        closing = f'İşbu sözleşme {contract_date.strftime("%d.%m.%Y")} tarihinde iki nüsha olarak imzalanmış ve yürürlüğe girmiştir.'
        _tail_merge(ws, current_row, 2, 13, closing, None, styles['text_font'], styles['left'], styles['thin_border'])
        ws.row_dimensions[current_row].height = max(26, min(72, 22 + (len(closing) // 90) * 15))
        current_row += 1
    return current_row


def _write_bank_tail(ws, quote, row):
    seller = _selected_seller_profile(quote)
    styles = _tail_styles()
    text_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, shrink_to_fit=True)
    bank_accounts = _seller_bank_accounts_with_ibans(seller)
    has_second_iban = _seller_has_second_iban(bank_accounts)

    _tail_section_header(ws, row, 'FİRMA ÜNVANI & IBANLAR')
    _tail_merge(ws, row + 1, 2, 13, _normalize_display_name(seller.get('display_name', '')), None, styles['label_font'], styles['center'], styles['section_border'])
    _tail_merge(ws, row + 2, 2, 4, 'BANKA', styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
    if has_second_iban:
        _tail_merge(ws, row + 2, 5, 8, _seller_bank_iban_label(seller, 1), styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
        _tail_merge(ws, row + 2, 9, 13, _seller_bank_iban_label(seller, 2), styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
    else:
        _tail_merge(ws, row + 2, 5, 13, 'IBAN', styles['soft_fill'], styles['label_font'], styles['center'], styles['section_border'])
    ws.row_dimensions[row + 1].height = 28
    ws.row_dimensions[row + 2].height = 24

    current_row = row + 3
    for account in bank_accounts or [{}]:
        _tail_merge(ws, current_row, 2, 4, _normalize_bank_name(account.get('bank', '')), None, styles['text_font'], text_alignment, styles['thin_border'])
        if has_second_iban:
            _tail_merge(ws, current_row, 5, 8, _bank_iban_display(account, 1), None, styles['text_font'], text_alignment, styles['thin_border'])
            _tail_merge(ws, current_row, 9, 13, _bank_iban_display(account, 2), None, styles['text_font'], text_alignment, styles['thin_border'])
        else:
            _tail_merge(ws, current_row, 5, 13, _bank_iban_display(account, 1, show_currency=False), None, styles['text_font'], text_alignment, styles['thin_border'])
        ws.row_dimensions[current_row].height = 26
        current_row += 1
    return current_row


def _write_signature_tail(ws, quote, row):
    customer = _customer_snapshot(quote.contract_config or {})
    seller = _selected_seller_profile(quote)
    styles = _tail_styles()
    _tail_section_header(ws, row, 'TARAFLARIN KAŞE İMZASI')
    signature_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    _tail_merge(ws, row + 1, 2, 7, f"SATICI\n{_normalize_display_name(seller.get('display_name', ''))}", None, styles['text_font'], signature_alignment, styles['section_border'])
    _tail_merge(ws, row + 1, 8, 13, f"ALICI\n{customer.get('name') or getattr(quote.customer, 'name', '') or ''}", None, styles['text_font'], signature_alignment, styles['section_border'])
    ws.row_dimensions[row + 1].height = 170
    return row + 2


def _prepare_bottom_banner_tail(ws, banner_images):
    if not banner_images:
        return None
    banner = banner_images[0]
    banner = _crop_excel_image_vertical_content(banner)
    target_width = _worksheet_printable_width_pixels(ws)
    banner = _scale_excel_image_to_width(banner, target_width)
    row_height = 22
    banner_height = float(getattr(banner, 'height', 0) or 0) * 0.75
    row_count = max(4, math.ceil(banner_height / row_height))
    return banner, row_height, row_count, banner_height


def _bottom_banner_footer_start_row(ws, page_start_row, min_row, prepared_banner):
    if not prepared_banner:
        return min_row

    _, row_height, _, banner_height = prepared_banner
    printable_height = _worksheet_printable_height_points(ws)
    used_height = _row_height_points(ws, page_start_row, min_row - 1)
    spacer_height = max(0, printable_height - used_height - banner_height)
    spacer_rows = max(0, math.floor(spacer_height / row_height))

    for row in range(min_row, min_row + spacer_rows):
        ws.row_dimensions[row].height = row_height
    partial_spacer_height = spacer_height - (spacer_rows * row_height)
    if partial_spacer_height >= 1:
        ws.row_dimensions[min_row + spacer_rows].height = partial_spacer_height
        spacer_rows += 1
    return min_row + spacer_rows


def _row_height_points(ws, start_row, end_row):
    if end_row < start_row:
        return 0
    total = 0
    for row in range(start_row, end_row + 1):
        total += float(ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15)
    return total


def _worksheet_printable_height_points(ws):
    orientation = str(getattr(ws.page_setup, 'orientation', '') or 'landscape').lower()
    paper_size = str(getattr(ws.page_setup, 'paperSize', '') or str(ws.PAPERSIZE_A4))
    if paper_size == str(ws.PAPERSIZE_A4):
        page_height = 8.27 if orientation == 'landscape' else 11.69
    else:
        page_height = 8.5 if orientation == 'landscape' else 11.0
    return max(1, (page_height - 0.4) * 72)


def _write_bottom_banner_tail(ws, row, prepared_banner):
    if not prepared_banner:
        return row
    banner, row_height, row_count, _ = prepared_banner
    for offset in range(row_count):
        ws.row_dimensions[row + offset].height = row_height
    ws.cell(row + row_count - 1, 13).value = ' '
    ws.add_image(banner, f'A{row}')
    return row + row_count


def _add_manual_page_break_before(ws, row):
    if row <= 1:
        return
    break_id = row - 1
    existing = {getattr(item, 'id', None) for item in getattr(ws.row_breaks, 'brk', []) or []}
    if break_id not in existing:
        ws.row_breaks.append(Break(id=break_id))


def _apply_times_new_roman_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ''):
                continue
            font = copy(cell.font)
            font.name = 'Times New Roman'
            if not font.sz or float(font.sz) < SELLER_MASTER_TABLE_FONT_SIZE:
                font.sz = SELLER_MASTER_TABLE_FONT_SIZE
            cell.font = font


def _dynamic_column_value(line, header):
    key = _normalize_column_key(header)
    details = dict(line.details or {})
    attributes = details.get('attributes') if isinstance(details.get('attributes'), dict) else {}
    primary, secondary = _resolve_detail_texts(details, attributes)
    net_unit = _net_unit_price(line)
    total = Decimal(line.qty or 0) * net_unit
    product = getattr(line, 'product', None)

    if 'kod' in key or 'sku' in key:
        return details.get('code') or getattr(product, 'sku', '') or '', 'text'
    if 'net' in key and ('birim' in key or 'fiyat' in key):
        return net_unit, 'currency'
    if 'tutar' in key or 'toplam' in key:
        return total, 'currency'
    if 'liste' in key or ('fiyat' in key and 'net' not in key):
        return Decimal(line.unit_price or 0), 'currency'
    if 'iskonto2' in key or '2iskonto' in key:
        return Decimal(getattr(line, 'discount_secondary', 0) or 0), 'number'
    if 'iskonto1' in key or '1iskonto' in key or key == 'iskonto':
        return Decimal(line.discount or 0), 'number'
    if 'kdv' in key:
        return Decimal(line.tax or 0), 'number'
    if 'miktar' in key or 'adet' in key:
        return Decimal(line.qty or 0), 'number'
    if 'olcu' in key or 'govde' in key:
        return primary, 'text'
    if 'renk' in key or 'kapak' in key or 'detay2' in key:
        return secondary, 'text'
    if key == 'birim' or key.endswith('birim'):
        return line.unit or 'Adet', 'text'
    if 'urun' in key or 'satisbirimi' in key or key in {'ad', 'adi'}:
        return line.name or '', 'text'
    if key in attributes:
        return attributes.get(key), 'text'
    return details.get(header) or details.get(key) or '', 'text'


def _normalize_column_key(value):
    return re.sub(r'[^a-z0-9]+', '', _normalize_document_group_text(value))


def _unmerge_overlapping_range(ws, min_row, max_row, min_col, max_col):
    for merged_range in list(ws.merged_cells.ranges):
        overlaps_row = not (merged_range.max_row < min_row or merged_range.min_row > max_row)
        overlaps_col = not (merged_range.max_col < min_col or merged_range.min_col > max_col)
        if overlaps_row and overlaps_col:
            try:
                ws.unmerge_cells(str(merged_range))
            except (KeyError, ValueError):
                try:
                    ws.merged_cells.ranges.remove(merged_range)
                except (KeyError, ValueError):
                    pass


def parse_terms_text(value):
    if isinstance(value, str):
        return [line.strip() for line in value.splitlines() if line.strip()]
    if isinstance(value, list):
        return [str(line).strip() for line in value if str(line).strip()]
    return []


def parse_contract_notes_text(value):
    return parse_terms_text(value)


def _select_templates(quote, template_key: str | None = None):
    config = quote.contract_config or {}
    requested_key = str(template_key or config.get('template_key') or '').strip()
    registry = [item for item in TEMPLATE_REGISTRY if item['document_type'] == quote.document_type]
    if requested_key:
        selected = next((item for item in registry if item['template_key'] == requested_key), None)
        if selected:
            return [selected]

    families = _quote_families(quote)
    if quote.document_type == 'Contract':
        combined = [
            item
            for item in registry
            if item.get('is_combined') and set(item.get('required_families') or []) == families
        ]
        if combined:
            return sorted(combined, key=lambda item: item['selection_priority'], reverse=True)[:1]

    family_order = ['steel', 'interior', 'furniture', 'service']
    selected = []
    for family in family_order:
        if family not in families:
            continue
        selected.append(next(item for item in registry if item['family'] == family and not item.get('is_combined')))
    return selected


def _quote_families(quote):
    families = set()
    for line in quote.lines.select_related('product__category').all():
        section_key = _canonical_section_key(_category_section_key(line) or line.section_key)
        family = SECTION_FAMILY_MAP.get(section_key) or _category_template_family(line)
        if family:
            families.add(family)
    return families


def _category_template_family(line):
    product = getattr(line, 'product', None)
    category = getattr(product, 'category', None) if product else None
    defaults = getattr(category, 'template_defaults', {}) or {}
    family = str(defaults.get('template_family') or '').strip()
    if family in {'steel', 'interior', 'furniture', 'service'}:
        return family
    normalized_family = _normalize_document_group_text(family)
    if normalized_family in {'celik kapi', 'steel door', 'celik'} or 'celik' in normalized_family or 'steel' in normalized_family:
        return 'steel'
    if normalized_family in {'ic oda', 'ic oda kapi', 'interior door'} or 'ic oda' in normalized_family or 'interior' in normalized_family:
        return 'interior'
    if normalized_family in {'mobilya', 'furniture', 'mutfak', 'portmanto', 'banyo'} or any(term in normalized_family for term in ['mobilya', 'furniture', 'mutfak', 'portmanto', 'banyo']):
        return 'furniture'
    if normalized_family in {'montaj', 'hizmet', 'service'} or any(term in normalized_family for term in ['montaj', 'hizmet', 'service']):
        return 'service'
    section_key = _canonical_section_key(defaults.get('section_key') or line.section_key)
    return SECTION_FAMILY_MAP.get(section_key, 'service' if section_key else '')


def _seller_profiles_for_template(quote):
    profiles = [profile for profile in get_seller_profiles(quote.organization) if profile.get('is_active', True)]
    if len(profiles) >= 2:
        return profiles[0], profiles[1]
    if len(profiles) == 1:
        fallback = get_default_seller_profiles()
        secondary = fallback[1] if normalize_seller_company_key(profiles[0].get('key')) != fallback[1]['key'] else fallback[0]
        return profiles[0], secondary
    defaults = get_default_seller_profiles()
    return defaults[0], defaults[1]


def _selected_seller_profile(quote):
    profiles = [profile for profile in get_seller_profiles(quote.organization) if profile.get('is_active', True)] or get_default_seller_profiles()
    selected_key = normalize_seller_company_key(getattr(quote, 'seller_company_key', '') or '')
    for profile in profiles:
        if normalize_seller_company_key(profile.get('key')) == selected_key:
            return profile
    return profiles[0]


def _selected_seller_side(quote):
    left_seller, right_seller = _seller_profiles_for_template(quote)
    selected = _selected_seller_profile(quote)
    selected_key = normalize_seller_company_key(selected.get('key'))
    if selected_key == normalize_seller_company_key(right_seller.get('key')):
        return 'right'
    return 'left'


def _fill_shared_header(ws, quote):
    config = quote.contract_config or {}
    customer = _customer_snapshot(config)
    prepared_by = _resolve_prepared_by_name(quote)
    currency_code = _quote_currency(quote)
    currency_note = _currency_note(quote)
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    doc_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)

    _set_cell_value(ws, 'L14', doc_date)
    _set_cell_value(ws, 'D23', customer.get('name') or getattr(quote.customer, 'name', ''))
    _set_cell_value(ws, 'D24', _join_tax(customer.get('tax_office'), customer.get('tax_number')))
    _set_cell_value(ws, 'D25', customer.get('address') or '')
    _set_cell_value(ws, 'D26', customer.get('authorized_person') or '')
    _set_cell_value(ws, 'D27', _join_contact(customer.get('phone'), customer.get('email')))
    _set_cell_value(ws, 'D30', quote.number)
    _ensure_header_value_row_layout(ws, 31)
    _set_cell_value(ws, 'D31', prepared_by)
    _ensure_header_value_row_layout(ws, 32, source_row=31)
    _set_cell_value(ws, 'D32', format_validity_text(quote.valid_until))
    _set_cell_value(ws, 'D33', ' • '.join(part for part in [get_quote_price_list_label(quote), currency_note] if part))
    _set_cell_value(ws, 'C16', _choice_text(_normalize_display_name(left_seller.get('display_name', '')), seller_key == left_seller.get('key')))
    _set_cell_value(ws, 'J16', _choice_text(_normalize_display_name(right_seller.get('display_name', '')), seller_key == right_seller.get('key')))
    _set_cell_value(ws, 'C17', f"VERGİ DAİRESİ: {left_seller.get('tax_office', '')}")
    _set_cell_value(ws, 'J17', f"VERGİ DAİRESİ: {right_seller.get('tax_office', '')}")
    _set_cell_value(ws, 'C18', f"VERGİ NO: {left_seller.get('tax_number', '')}")
    _set_cell_value(ws, 'J18', f"VERGİ NO: {right_seller.get('tax_number', '')}")
    _set_cell_value(ws, 'C19', _join_contact(left_seller.get('email', ''), left_seller.get('phone', '')))
    _set_cell_value(ws, 'J19', _join_contact(right_seller.get('email', ''), right_seller.get('phone', '')))


def _expand_template_for_line_counts(ws, template, quote):
    layout = deepcopy(template)
    expanded = False
    ordered_lines = list(quote.lines.select_related('product__category').order_by('sort_order', 'id'))

    for block_index, block in enumerate(layout['line_blocks']):
        lines = _select_lines_for_block(ordered_lines, block)
        extra_rows = max(0, len(lines) - block['capacity'])
        if extra_rows <= 0:
            continue

        insert_at = block['subtotal_row']
        _insert_styled_rows(ws, insert_at, extra_rows, block['subtotal_row'] - 1)
        block['capacity'] += extra_rows
        _shift_template_layout_rows(layout, insert_at, extra_rows, block_index)
        expanded = True

    return layout, expanded


def _shift_template_layout_rows(template, insert_at, offset, current_block_index):
    current_block = template['line_blocks'][current_block_index]
    current_block['subtotal_row'] += offset
    current_block['tax_row'] += offset
    current_block['grand_row'] += offset

    for next_block in template['line_blocks'][current_block_index + 1 :]:
        next_block['start_row'] += offset
        next_block['subtotal_row'] += offset
        next_block['tax_row'] += offset
        next_block['grand_row'] += offset

    template['commercial_rows'] = [row + offset if row >= insert_at else row for row in template['commercial_rows']]
    template['terms_rows'] = [row + offset if row >= insert_at else row for row in template['terms_rows']]
    template['bank_rows'] = [row + offset if row >= insert_at else row for row in template['bank_rows']]
    if template['bank_header_row'] >= insert_at:
        template['bank_header_row'] += offset
    if template['signature_row'] >= insert_at:
        template['signature_row'] += offset


def _insert_styled_rows(ws, insert_at, amount, source_row):
    ws.insert_rows(insert_at, amount)
    for row_offset in range(amount):
        target_row = insert_at + row_offset
        _copy_row_layout(ws, source_row, target_row)
        _clear_line_row(ws, target_row)


def _copy_row_layout(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    for column in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, column)
        target_cell = ws.cell(target_row, column)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = ''


def _fill_line_blocks(ws, quote, template):
    ordered_lines = list(quote.lines.select_related('product__category').order_by('sort_order', 'id'))
    currency_code = _quote_currency(quote)
    currency_symbol = _currency_symbol(currency_code)
    for block in template['line_blocks']:
        lines = _select_lines_for_block(ordered_lines, block)
        visible = bool(lines)
        for row in range(block['start_row'] - 2, block['grand_row'] + 1):
            ws.row_dimensions[row].hidden = not visible

        for offset in range(block['capacity']):
            row_number = block['start_row'] + offset
            has_line = offset < len(lines)
            ws.row_dimensions[row_number].hidden = not has_line
            if has_line:
                _clear_line_row(ws, row_number)
            else:
                _clear_line_row(ws, row_number, clear_visual=True)
                continue
            line = lines[offset]
            details = dict(line.details or {})
            attributes = details.get('attributes') if isinstance(details.get('attributes'), dict) else {}
            primary, secondary = _resolve_detail_texts(details, attributes)
            net_unit = _net_unit_price(line)
            line_total = Decimal(line.qty) * net_unit

            _set_cell_value(ws, f'C{row_number}', details.get('code') or getattr(line.product, 'sku', '') or '')
            _set_cell_value(ws, f'D{row_number}', line.name)
            _set_cell_value(ws, f'E{row_number}', primary)
            _set_cell_value(ws, f'F{row_number}', secondary)
            _set_cell_value(ws, f'G{row_number}', float(line.qty) if Decimal(line.qty or 0) > 0 else '')
            _set_currency_value(ws, f'H{row_number}', float(line.unit_price) if Decimal(line.unit_price or 0) > 0 else '', currency_code)
            _set_cell_value(ws, f'I{row_number}', float(Decimal(line.discount or 0)) if Decimal(line.discount or 0) > 0 else '')
            _set_cell_value(ws, f'J{row_number}', float(Decimal(getattr(line, 'discount_secondary', 0) or 0)) if Decimal(getattr(line, 'discount_secondary', 0) or 0) > 0 else '')
            _set_cell_value(ws, f'K{row_number}', line.unit or 'Adet')
            _set_currency_value(ws, f'L{row_number}', float(net_unit), currency_code)
            _set_currency_value(ws, f'M{row_number}', float(line_total), currency_code)

        subtotal_value = sum((_line_subtotal(line) for line in lines), Decimal('0'))
        tax_value = sum((_line_tax(line) for line in lines), Decimal('0'))
        grand_value = subtotal_value + tax_value
        for summary_row in [block['subtotal_row'], block['tax_row'], block['grand_row']]:
            for column in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                _clear_visual_cell(ws, f'{column}{summary_row}')
        _set_cell_value(ws, f'L{block["subtotal_row"]}', f'Tutar Toplam ({currency_symbol})' if visible else '')
        _set_cell_value(ws, f'L{block["tax_row"]}', _summary_tax_label(lines) if visible else '')
        _set_cell_value(ws, f'L{block["grand_row"]}', f'Yekün ({currency_symbol})' if visible else '')
        _set_currency_value(ws, f'M{block["subtotal_row"]}', float(subtotal_value) if visible else '', currency_code)
        _set_currency_value(ws, f'M{block["tax_row"]}', float(tax_value) if visible else '', currency_code)
        _set_currency_value(ws, f'M{block["grand_row"]}', float(grand_value) if visible else '', currency_code)


def _select_lines_for_block(ordered_lines, block):
    section_keys = {str(key or '').strip() for key in block.get('section_keys', []) if str(key or '').strip()}
    summary_keys = {str(key or '').strip() for key in block.get('summary_of_sections', []) if str(key or '').strip()}
    allowed_keys = section_keys | summary_keys
    if not allowed_keys:
        return []
    return [
        line
        for line in ordered_lines
        if _line_is_meaningful(line)
        and (
            _canonical_section_key(_category_section_key(line) or line.section_key) in allowed_keys
            or ('service' in allowed_keys and _category_template_family(line) == 'service')
        )
    ]


def _fill_commercial_rows(ws, quote, template):
    config = quote.contract_config or {}
    currency_code = _quote_currency(quote)
    prepared_by = _resolve_prepared_by_name(quote)
    commercial_values = [
        quote.number,
        prepared_by,
        quote.payment_terms or '',
        _format_date_text(quote.delivery_terms),
        format_validity_text(quote.valid_until),
        (config.get('delivery_type') or '').strip(),
        (config.get('payment_option') or '').strip(),
    ]
    for row, value in zip(template['commercial_rows'], commercial_values, strict=False):
        _set_cell_value(ws, f'D{row}', value)
        for column in ['L', 'M']:
            cell = _resolve_cell(ws, f'{column}{row}')
            if cell.value not in [None, '']:
                cell.number_format = _currency_number_format(currency_code)


def _fill_terms(ws, quote, template):
    terms = parse_terms_text((quote.contract_config or {}).get('terms_text'))
    for row in template['terms_rows']:
        _set_cell_value(ws, f'C{row}', '')
    for row, term in zip(template['terms_rows'], terms, strict=False):
        _set_cell_value(ws, f'C{row}', term)
    if quote.document_type == 'Contract':
        _fill_contract_notes(ws, quote, template)


def _fill_contract_notes(ws, quote, template):
    config = quote.contract_config or {}
    note_rows = list(template['terms_rows'][1:7])
    closing_row = template['terms_rows'][-1]
    notes = parse_contract_notes_text(config.get('contract_notes_text') or DEFAULT_CONTRACT_NOTES_TEXT)
    for index, row in enumerate(note_rows, start=1):
        text = notes[index - 1] if index - 1 < len(notes) else ''
        _set_cell_value(ws, f'I{row}', f'{index}- {text}'.rstrip())

    contract_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)
    _set_cell_value(
        ws,
        f'I{closing_row}',
        f'İşbu sözleşme {contract_date.strftime("%d.%m.%Y")} tarihinde iki nüsha olarak imzalanmış ve yürürlüğe girmiştir.',
    )


def _fill_bank_accounts(ws, quote, template):
    currency_symbol = _currency_symbol(_quote_currency(quote))
    left_seller, right_seller = _seller_profiles_for_template(quote)
    header_row = template['bank_header_row']
    if not template.get('expanded_layout'):
        _ensure_bank_header_layout(ws, header_row)
    _set_cell_value(ws, f'C{header_row}', f"{_normalize_display_name(left_seller.get('display_name', ''))} {currency_symbol}")
    _set_cell_value(ws, f'K{header_row}', f"{_normalize_display_name(right_seller.get('display_name', ''))} {currency_symbol}")
    ortka_accounts = left_seller.get('bank_accounts') or []
    ayka_accounts = right_seller.get('bank_accounts') or []
    for index, row in enumerate(template['bank_rows']):
        if not template.get('expanded_layout'):
            _ensure_bank_row_layout(ws, row)
        left = ortka_accounts[index] if index < len(ortka_accounts) else {}
        right = ayka_accounts[index] if index < len(ayka_accounts) else {}
        left_bank = _normalize_bank_name(left.get('bank', '')) if left.get('bank') and left.get('iban') else ''
        left_iban = left.get('iban', '') if left.get('bank') and left.get('iban') else ''
        right_bank = _normalize_bank_name(right.get('bank', '')) if right.get('bank') and right.get('iban') else ''
        right_iban = right.get('iban', '') if right.get('bank') and right.get('iban') else ''
        _set_cell_value(ws, f'C{row}', left_bank)
        _set_cell_value(ws, f'D{row}', left_iban)
        _set_cell_value(ws, f'K{row}', right_bank)
        _set_cell_value(ws, f'M{row}', right_iban)


def _fill_signature_block(ws, quote, template):
    customer = _customer_snapshot(quote.contract_config or {})
    profiles = {profile['key'].upper(): profile for profile in get_seller_profiles(quote.organization)}
    fallback_seller = _seller_profiles_for_template(quote)[1]
    seller = profiles.get((quote.seller_company_key or fallback_seller.get('key') or 'AYKA').strip().upper(), fallback_seller)
    row = template['signature_row']
    signature_label = str(seller.get('signature_label') or '').strip() or f"{_normalize_display_name(seller.get('display_name', ''))} ADINA İMZA"
    _set_cell_value(ws, f'C{row}', signature_label)
    _set_cell_value(ws, f'J{row}', customer.get('name') or getattr(quote.customer, 'name', '') or '')


def _apply_template_placeholders(ws, quote, template):
    context = _build_template_placeholder_context(quote, template)
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or '{' not in cell.value or '}' not in cell.value:
                continue
            _replace_placeholder_in_cell(cell, context)


def _replace_placeholder_in_cell(cell, context):
    source = str(cell.value or '')
    matches = list(PLACEHOLDER_PATTERN.finditer(source))
    if not matches:
        return

    if len(matches) == 1 and matches[0].span() == (0, len(source)):
        key = matches[0].group(1)
        if key in context:
            cell.value = _placeholder_typed_value(context.get(key))
        return

    rendered = source
    for match in matches:
        key = match.group(1)
        rendered = rendered.replace(match.group(0), _placeholder_display_value(context.get(key, '')))
    cell.value = rendered


def _placeholder_typed_value(value):
    if value in [None, '']:
        return ''
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.date()
    return value


def _placeholder_display_value(value):
    if value in [None, '']:
        return ''
    if isinstance(value, Decimal):
        return _format_decimal_text(value)
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    return str(value)


def _format_decimal_text(value):
    try:
        decimal_value = Decimal(str(value))
    except Exception:
        return str(value)
    rendered = f'{decimal_value:,.2f}'
    return rendered.replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_template_placeholder_context(quote, template):
    config = quote.contract_config or {}
    customer = _customer_snapshot(config)
    currency_code = _quote_currency(quote)
    exchange_rate = _quote_exchange_rate(quote)
    created_at = _timezone_fallback(quote.created_at)
    updated_at = _timezone_fallback(quote.updated_at)
    valid_until = _parse_date(quote.valid_until)
    selected_seller_key = (quote.seller_company_key or '').strip().upper()
    all_sellers = [profile for profile in get_seller_profiles(quote.organization) if profile.get('is_active', True)] or get_default_seller_profiles()
    selected_seller = next((profile for profile in all_sellers if profile.get('key') == selected_seller_key), all_sellers[0] if all_sellers else {})
    meaningful_lines = [
        line for line in quote.lines.select_related('product__category').order_by('sort_order', 'id')
        if _line_is_meaningful(line)
    ]
    document_title = 'TEKLİF' if quote.document_type == 'Quote' else 'SÖZLEŞME'

    context = {
        'belgeNo': quote.number,
        'belgeTuru': document_title,
        'olusturmaTarihi': created_at,
        'guncellenmeTarihi': updated_at,
        'gecerlilikTarihi': valid_until,
        'teslimTarihi': _parse_date(quote.delivery_terms) or quote.delivery_terms or '',
        'teslimTipi': (config.get('delivery_type') or '').strip(),
        'odemeKosulu': quote.payment_terms or '',
        'odemeTipi': (config.get('payment_option') or '').strip(),
        'hazirlayan': _resolve_prepared_by_name(quote),
        'saticiLogo': '',
        'fiyatListesiEtiketi': get_quote_price_list_label(quote),
        'paraBirimi': _currency_symbol(currency_code),
        'paraBirimiKodu': currency_code,
        'kur': exchange_rate,
        'araToplam': Decimal(quote.subtotal or 0),
        'iskontoToplami': Decimal(quote.discount_total or 0),
        'kdvToplami': Decimal(quote.tax_total or 0),
        'genelToplam': Decimal(quote.total or 0),
        'yekun': Decimal(quote.total or 0),
        'kdvOrani': Decimal(quote.vat_rate or 0),
        'notlar': quote.notes or '',
        'cariUnvani': customer.get('name') or getattr(quote.customer, 'name', '') or '',
        'cariAdi': customer.get('name') or getattr(quote.customer, 'name', '') or '',
        'yetkili': customer.get('authorized_person') or getattr(quote.customer, 'authorized_person', '') or '',
        'vergiDairesi': customer.get('tax_office') or getattr(quote.customer, 'tax_office', '') or '',
        'vergiNo': customer.get('tax_number') or getattr(quote.customer, 'tax_number', '') or '',
        'adres': customer.get('address') or getattr(quote.customer, 'address', '') or '',
        'telefon': customer.get('phone') or getattr(quote.customer, 'phone', '') or '',
        'email': customer.get('email') or getattr(quote.customer, 'email', '') or '',
        'sehir': getattr(quote.customer, 'city', '') or '',
        'ulke': getattr(quote.customer, 'country', '') or '',
        'cariParaBirimi': getattr(quote.customer, 'currency', '') or '',
        'cariOlcegi': getattr(quote.customer, 'size', '') or '',
    }

    context.update(_build_seller_placeholder_context(selected_seller, prefix='seciliSatici'))
    for index, seller in enumerate(all_sellers, start=1):
        context.update(_build_seller_placeholder_context(seller, prefix=f'saticiFirma{index}'))

    for index, line in enumerate(meaningful_lines, start=1):
        context.update(_build_line_placeholder_context(line, index))

    return context


def _build_seller_placeholder_context(profile, prefix):
    normalized = _normalize_seller_profile(profile or {})
    context = {
        f'{prefix}.kod': normalized.get('key', ''),
        f'{prefix}.kisaAd': normalized.get('short_name', ''),
        f'{prefix}.unvan': _normalize_display_name(normalized.get('display_name', '')),
        f'{prefix}.resmiUnvan': _normalize_display_name(normalized.get('legal_name', '')),
        f'{prefix}.vergiDairesi': normalized.get('tax_office', ''),
        f'{prefix}.vergiNo': normalized.get('tax_number', ''),
        f'{prefix}.mersisNo': normalized.get('mersis_number', ''),
        f'{prefix}.ticaretSicilNo': normalized.get('trade_registry_number', ''),
        f'{prefix}.adres': normalized.get('address', ''),
        f'{prefix}.sehir': normalized.get('city', ''),
        f'{prefix}.ulke': normalized.get('country', ''),
        f'{prefix}.telefon': normalized.get('phone', ''),
        f'{prefix}.email': normalized.get('email', ''),
        f'{prefix}.web': normalized.get('website', ''),
        f'{prefix}.kep': normalized.get('kep_address', ''),
        f'{prefix}.imzaAdi': normalized.get('signature_name', ''),
        f'{prefix}.imzaUnvani': normalized.get('signature_title', ''),
        f'{prefix}.imzaEtiketi': normalized.get('signature_label', ''),
        f'{prefix}.notlar': normalized.get('notes', ''),
        f'{prefix}.ibanBaslik1': _seller_bank_iban_label(normalized, 1),
        f'{prefix}.ibanBaslik2': _seller_bank_iban_label(normalized, 2),
    }
    for bank_index, bank_account in enumerate(normalized.get('bank_accounts') or [], start=1):
        context[f'{prefix}.banka{bank_index}.ad'] = _normalize_bank_name(bank_account.get('bank', ''))
        context[f'{prefix}.banka{bank_index}.iban'] = bank_account.get('iban', '')
        context[f'{prefix}.banka{bank_index}.iban2'] = bank_account.get('iban_2', '')
        context[f'{prefix}.banka{bank_index}.sube'] = bank_account.get('branch', '')
        context[f'{prefix}.banka{bank_index}.hesapSahibi'] = bank_account.get('account_holder', '')
        context[f'{prefix}.banka{bank_index}.paraBirimi'] = bank_account.get('currency', '')
        context[f'{prefix}.banka{bank_index}.paraBirimi2'] = bank_account.get('currency_2', '')
    for bank_index in range(1, 11):
        context.setdefault(f'{prefix}.banka{bank_index}.ad', '')
        context.setdefault(f'{prefix}.banka{bank_index}.iban', '')
        context.setdefault(f'{prefix}.banka{bank_index}.iban2', '')
        context.setdefault(f'{prefix}.banka{bank_index}.sube', '')
        context.setdefault(f'{prefix}.banka{bank_index}.hesapSahibi', '')
        context.setdefault(f'{prefix}.banka{bank_index}.paraBirimi', '')
        context.setdefault(f'{prefix}.banka{bank_index}.paraBirimi2', '')
    return context


def _build_line_placeholder_context(line, index):
    details = dict(line.details or {})
    attributes = details.get('attributes') if isinstance(details.get('attributes'), dict) else {}
    primary, secondary = _resolve_detail_texts(details, attributes)
    net_unit = _net_unit_price(line)
    line_total = Decimal(line.qty or 0) * net_unit
    prefix = f'kalem{index}'
    return {
        f'{prefix}.kod': details.get('code') or getattr(getattr(line, 'product', None), 'sku', '') or '',
        f'{prefix}.urun': line.name or '',
        f'{prefix}.ad': line.name or '',
        f'{prefix}.miktar': Decimal(line.qty or 0),
        f'{prefix}.birim': line.unit or 'Adet',
        f'{prefix}.listeFiyati': Decimal(line.unit_price or 0),
        f'{prefix}.birimFiyat': Decimal(line.unit_price or 0),
        f'{prefix}.netBirimFiyat': net_unit,
        f'{prefix}.tutar': line_total,
        f'{prefix}.olcu': primary,
        f'{prefix}.renk': secondary,
        f'{prefix}.detay1': primary,
        f'{prefix}.detay2': secondary,
        f'{prefix}.iskonto1': Decimal(line.discount or 0),
        f'{prefix}.iskonto2': Decimal(getattr(line, 'discount_secondary', 0) or 0),
        f'{prefix}.kdvOrani': Decimal(line.tax or 0),
    }


def _resolve_detail_texts(details, attributes):
    attribute_lines = []
    for key, value in (attributes or {}).items():
        if value in [None, '']:
            continue
        label = str(key).replace('_', ' ').strip().title()
        attribute_lines.append(f"{label}: {value}")
    primary = details.get('primary') or (attribute_lines[0] if attribute_lines else '')
    secondary_parts = []
    if details.get('secondary'):
        secondary_parts.append(str(details.get('secondary')))
    secondary_parts.extend(attribute_lines[1:3])
    return primary, ' | '.join(part for part in secondary_parts if part)


def _normalize_bank_name(value):
    normalized = {
        'Turkiye Is Bankasi': 'Türkiye İş Bankası',
        'Ziraat Bankasi': 'Ziraat Bankası',
        'Albaraka Turk': 'Albaraka Türk',
        'Albaraka Turk K.Bank': 'Albaraka Türk K.Bank.',
        'Albaraka Turk K.Bank.': 'Albaraka Türk K.Bank.',
        'Vakiflar Bankasi': 'Vakıflar Bankası',
    }
    text = str(value or '').strip()
    return normalized.get(text, text)


def _summary_tax_label(lines):
    tax_rates = []
    for line in lines:
        try:
            rate = Decimal(line.tax or 0)
        except Exception:
            rate = Decimal('0')
        if rate > 0:
            tax_rates.append(rate)

    unique_rates = sorted({rate.normalize() for rate in tax_rates})
    if len(unique_rates) == 1:
        rate_label = format(unique_rates[0], 'f')
        if '.' in rate_label:
            rate_label = rate_label.rstrip('0').rstrip('.')
        return f'K.D.V. %{rate_label}'
    return 'K.D.V.'


def _ensure_header_value_row_layout(ws, row_number, source_row=31):
    _unmerge_row_slice(ws, row_number, 4, 13)
    for column in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        _copy_bank_cell_layout(ws, f'{column}{source_row}', f'{column}{row_number}')
        _set_cell_value(ws, f'{column}{row_number}', '')


def _ensure_bank_row_layout(ws, row_number):
    _unmerge_row_slice(ws, row_number, 11, 14)
    _copy_bank_cell_layout(ws, f'C{row_number}', f'K{row_number}')
    _copy_bank_cell_layout(ws, f'C{row_number}', f'L{row_number}')
    _copy_bank_cell_layout(ws, f'D{row_number}', f'M{row_number}')
    for coordinate in [f'K{row_number}', f'L{row_number}', f'M{row_number}', f'N{row_number}']:
        _set_cell_value(ws, coordinate, '')
    _clear_visual_cell(ws, f'N{row_number}')
    for target_merge in [f'K{row_number}:L{row_number}']:
        if all(str(existing) != target_merge for existing in ws.merged_cells.ranges):
            ws.merge_cells(target_merge)


def _ensure_bank_header_layout(ws, row_number):
    _unmerge_row_slice(ws, row_number, 11, 14)
    _copy_bank_cell_layout(ws, f'C{row_number}', f'K{row_number}')
    _copy_bank_cell_layout(ws, f'D{row_number}', f'L{row_number}')
    _copy_bank_cell_layout(ws, f'D{row_number}', f'M{row_number}')
    for coordinate in [f'K{row_number}', f'L{row_number}', f'M{row_number}', f'N{row_number}']:
        _set_cell_value(ws, coordinate, '')
    _clear_visual_cell(ws, f'N{row_number}')
    target_merge = f'K{row_number}:M{row_number}'
    if all(str(existing) != target_merge for existing in ws.merged_cells.ranges):
        ws.merge_cells(target_merge)


def _unmerge_row_slice(ws, row_number, min_col, max_col):
    for merged_range in list(ws.merged_cells.ranges):
        overlaps_row = merged_range.min_row <= row_number <= merged_range.max_row
        overlaps_col = not (merged_range.max_col < min_col or merged_range.min_col > max_col)
        if overlaps_row and overlaps_col:
            ws.unmerge_cells(str(merged_range))


def _copy_bank_cell_layout(ws, source_coordinate, target_coordinate):
    source = _resolve_cell(ws, source_coordinate)
    target = _resolve_cell(ws, target_coordinate)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _user_display_name(user):
    if not user:
        return ''
    full_name = f"{(getattr(user, 'first_name', '') or '').strip()} {(getattr(user, 'last_name', '') or '').strip()}".strip()
    if not full_name and hasattr(user, 'get_full_name'):
        full_name = str(user.get_full_name() or '').strip()
    return full_name or str(getattr(user, 'username', '') or getattr(user, 'email', '') or '').strip()


def _resolve_prepared_by_name(quote):
    snapshot = dict((quote.contract_config or {}).get('prepared_by_snapshot') or {})
    snapshot_name = str(snapshot.get('name') or '').strip()
    related_name = _user_display_name(getattr(quote, 'prepared_by', None)) or _user_display_name(getattr(quote, 'owner', None))
    if snapshot_name and '@' not in snapshot_name:
        return snapshot_name
    return related_name or snapshot_name


def _normalize_display_name(value):
    normalized = {
        'ORTKA YAPI ELEMANLARI URETIM SAN. LTD. STI.': 'ORTKA YAPI ELEMANLARI ÜRETİM SAN. LTD. ŞTİ.',
        'AYKA KAPI SANAYI TICARET ANONIM SIRKETI': 'AYKA KAPI SANAYİ TİCARET ANONİM ŞİRKETİ',
    }
    text = str(value or '').strip()
    return normalized.get(text, text)


def _normalize_currency_code(currency):
    normalized = str(currency or 'TRY').strip().upper()
    return normalized if normalized in CURRENCY_SYMBOLS else 'TRY'


def _quote_currency(quote):
    return _normalize_currency_code(getattr(quote, 'currency', 'TRY'))


def _currency_symbol(currency):
    return CURRENCY_SYMBOLS.get(_normalize_currency_code(currency), '₺')


def _quote_exchange_rate(quote):
    config = quote.contract_config or {}
    try:
        rate = Decimal(str(config.get('exchange_rate') or 1))
    except Exception:
        rate = Decimal('1')
    return rate if rate > 0 else Decimal('1')


def _format_exchange_rate(rate: Decimal):
    return format(rate.normalize(), 'f').rstrip('0').rstrip('.') or '1'


def _currency_note(quote):
    currency_code = _quote_currency(quote)
    currency_symbol = _currency_symbol(currency_code)
    if currency_code == 'TRY':
        return f'Para Birimi: {currency_symbol}'
    return f'Para Birimi: {currency_symbol} | Kur: 1 {currency_symbol} = {_format_exchange_rate(_quote_exchange_rate(quote))} ₺'


def _currency_number_format(currency):
    return f'"{_currency_symbol(currency)}" #,##0.00'


def _resolve_cell(ws, coordinate):
    cell = ws[coordinate]
    if cell.__class__.__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col)
        return ws.cell(cell.row, cell.column)
    return cell


def _set_currency_value(ws, coordinate, value, currency):
    cell = _resolve_cell(ws, coordinate)
    cell.value = value
    if value not in [None, '']:
        cell.number_format = _currency_number_format(currency)


def _clear_visual_cell(ws, coordinate):
    cell = _resolve_cell(ws, coordinate)
    cell.value = ''
    cell.border = EMPTY_BORDER
    cell.fill = EMPTY_FILL
    cell.number_format = 'General'


def _clear_line_row(ws, row_number, clear_visual=False):
    for column in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        if clear_visual:
            _clear_visual_cell(ws, f'{column}{row_number}')
        else:
            _set_cell_value(ws, f'{column}{row_number}', '')


def _set_cell_value(ws, coordinate, value):
    cell = _resolve_cell(ws, coordinate)
    cell.value = value


def _default_seller_logo_path(profile):
    key = normalize_seller_company_key((profile or {}).get('key') or (profile or {}).get('short_name'))
    filename = {'AYKA': 'ayka-logo.png', 'ORTKA': 'ortka-logo.png'}.get(key)
    if not filename:
        return None
    for candidate in (
        Path(settings.MEDIA_ROOT) / 'seller-logos' / 'default' / filename,
        ASSETS_DIR / 'seller-logos' / 'default' / filename,
    ):
        if candidate.exists():
            return candidate
    return None


def _seller_logo_media_path(profile):
    raw = str((profile or {}).get('logo_url') or '').strip()
    if not raw:
        return _default_seller_logo_path(profile)
    media_prefix = settings.MEDIA_URL.rstrip('/') + '/'
    if raw.startswith(media_prefix):
        relative = raw[len(media_prefix):]
    elif raw.startswith('/media/'):
        relative = raw[len('/media/'):]
    elif raw.startswith('http://') or raw.startswith('https://'):
        marker = '/media/'
        if marker not in raw:
            return None
        relative = raw.split(marker, 1)[1]
    else:
        relative = raw.lstrip('/')
    candidate = Path(settings.MEDIA_ROOT) / relative
    if not candidate.exists():
        return _default_seller_logo_path(profile)
    if candidate.suffix.lower() not in {'.png', '.jpg', '.jpeg'}:
        return _default_seller_logo_path(profile)
    return candidate


def _fit_excel_image(image, max_width, max_height):
    width = float(getattr(image, 'width', max_width) or max_width)
    height = float(getattr(image, 'height', max_height) or max_height)
    if width <= 0 or height <= 0:
        image.width = max_width
        image.height = max_height
        return image
    scale = min(max_width / width, max_height / height)
    image.width = int(width * scale)
    image.height = int(height * scale)
    return image


def _scale_excel_image_to_width(image, target_width):
    width = float(getattr(image, 'width', target_width) or target_width)
    height = float(getattr(image, 'height', 0) or 0)
    if width <= 0 or height <= 0:
        image.width = int(target_width)
        return image
    scale = target_width / width
    image.width = int(target_width)
    image.height = int(height * scale)
    return image


def _crop_excel_image_content(image, padding=0):
    try:
        source = PILImage.open(BytesIO(image._data())).convert('RGBA')
    except Exception:
        return image

    bbox = _excel_image_content_bbox(source)
    if not bbox:
        return image

    left = max(0, bbox[0] - padding)
    top = max(0, bbox[1] - padding)
    right = min(source.width, bbox[2] + padding)
    bottom = min(source.height, bbox[3] + padding)
    if left <= 0 and top <= 0 and right >= source.width and bottom >= source.height:
        return image

    output = BytesIO()
    source.crop((left, top, right, bottom)).save(output, format='PNG')
    output.seek(0)
    return XLImage(output)


def _excel_image_content_bbox(source):
    alpha_bbox = source.getchannel('A').getbbox()
    if not alpha_bbox:
        return None

    pixels = source.load()
    bg_r, bg_g, bg_b, bg_a = pixels[0, 0]
    left, top, right, bottom = source.width, source.height, 0, 0
    found = False
    tolerance = 18

    for y in range(source.height):
        for x in range(source.width):
            r, g, b, a = pixels[x, y]
            if a <= 10:
                continue
            alpha_delta = abs(a - bg_a)
            color_delta = max(abs(r - bg_r), abs(g - bg_g), abs(b - bg_b))
            if alpha_delta <= tolerance and color_delta <= tolerance:
                continue
            left = min(left, x)
            top = min(top, y)
            right = max(right, x + 1)
            bottom = max(bottom, y + 1)
            found = True

    return (left, top, right, bottom) if found else alpha_bbox


def _crop_excel_image_vertical_content(image):
    try:
        source = PILImage.open(BytesIO(image._data())).convert('RGBA')
    except Exception:
        return image

    bbox = source.getchannel('A').getbbox()
    if not bbox:
        return image

    padding = 12
    top = max(0, bbox[1] - padding)
    bottom = min(source.height, bbox[3] + padding)
    if top <= 0 and bottom >= source.height:
        return image

    output = BytesIO()
    source.crop((0, top, source.width, bottom)).save(output, format='PNG')
    output.seek(0)
    return XLImage(output)


def _worksheet_printable_width_pixels(ws):
    orientation = str(getattr(ws.page_setup, 'orientation', '') or '').lower()
    paper_size = str(getattr(ws.page_setup, 'paperSize', '') or '')
    if paper_size == str(ws.PAPERSIZE_A4):
        page_width = 11.69 if orientation == 'landscape' else 8.27
    else:
        page_width = 11.69 if orientation == 'landscape' else 8.5
    margins = ws.page_margins
    printable_width = max(1, page_width - float(margins.left or 0) - float(margins.right or 0))
    return int(printable_width * 144)


def _clear_logo_frame(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            break
    else:
        min_col = max_col = cell.column
        min_row = max_row = cell.row
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            target = ws.cell(row, column)
            target.border = EMPTY_BORDER
            target.fill = EMPTY_FILL


def _strip_header_logos(ws):
    kept = []
    for image in list(getattr(ws, '_images', [])):
        anchor = getattr(image, 'anchor', None)
        marker = getattr(anchor, '_from', None)
        if marker and marker.row <= 12:
            continue
        kept.append(image)
    ws._images = kept


def _apply_seller_logo(ws, profile, seller_side):
    _strip_header_logos(ws)
    logo_path = _seller_logo_media_path(profile)
    if not logo_path:
        return
    image = XLImage(str(logo_path))
    image = _fit_excel_image(image, 220, 76)
    ws.add_image(image, 'K1' if seller_side == 'right' else 'B1')


def _fill_shared_header(ws, quote):
    config = quote.contract_config or {}
    customer = _customer_snapshot(config)
    prepared_by = _resolve_prepared_by_name(quote)
    currency_note = _currency_note(quote)
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    doc_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)

    _set_cell_value(ws, 'L14', doc_date)
    _set_cell_value(ws, 'D23', customer.get('name') or getattr(quote.customer, 'name', ''))
    _set_cell_value(ws, 'D24', _join_tax(customer.get('tax_office'), customer.get('tax_number')))
    _set_cell_value(ws, 'D25', customer.get('address') or '')
    _set_cell_value(ws, 'D26', customer.get('authorized_person') or '')
    _set_cell_value(ws, 'D27', _join_contact(customer.get('phone'), customer.get('email')))
    _set_cell_value(ws, 'D30', quote.number)
    _ensure_header_value_row_layout(ws, 31)
    _set_cell_value(ws, 'D31', prepared_by)
    _ensure_header_value_row_layout(ws, 32, source_row=31)
    _set_cell_value(ws, 'D32', format_validity_text(quote.valid_until))
    _set_cell_value(ws, 'D33', ' • '.join(part for part in [get_quote_price_list_label(quote), currency_note] if part))

    for coordinate in ['C16', 'C17', 'C18', 'C19', 'J16', 'J17', 'J18', 'J19']:
        _set_cell_value(ws, coordinate, '')

    target_prefix = 'J' if seller_side == 'right' else 'C'
    _set_cell_value(ws, f'{target_prefix}16', _normalize_display_name(seller.get('display_name', '')))
    _set_cell_value(ws, f'{target_prefix}17', f"VERGİ DAİRESİ : {seller.get('tax_office', '')}")
    _set_cell_value(ws, f'{target_prefix}18', f"VERGİ NO: {seller.get('tax_number', '')}")
    _set_cell_value(ws, f'{target_prefix}19', _join_contact(seller.get('email', ''), seller.get('phone', '')))
    _apply_seller_logo(ws, seller, seller_side)


def _fill_bank_accounts(ws, quote, template):
    currency_symbol = _currency_symbol(_quote_currency(quote))
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    header_row = template['bank_header_row']
    if not template.get('expanded_layout'):
        _ensure_bank_header_layout(ws, header_row)

    for coordinate in [f'C{header_row}', f'K{header_row}']:
        _set_cell_value(ws, coordinate, '')
    for row in template['bank_rows']:
        if not template.get('expanded_layout'):
            _ensure_bank_row_layout(ws, row)
        for coordinate in [f'C{row}', f'D{row}', f'K{row}', f'M{row}']:
            _set_cell_value(ws, coordinate, '')

    header_col = 'K' if seller_side == 'right' else 'C'
    bank_col = 'K' if seller_side == 'right' else 'C'
    iban_col = 'M' if seller_side == 'right' else 'D'

    _set_cell_value(ws, f'{header_col}{header_row}', f"{_normalize_display_name(seller.get('display_name', ''))} {currency_symbol}")
    for index, row in enumerate(template['bank_rows']):
        account = (seller.get('bank_accounts') or [])[index] if index < len(seller.get('bank_accounts') or []) else {}
        bank_name = _normalize_bank_name(account.get('bank', '')) if account.get('bank') and account.get('iban') else ''
        iban = account.get('iban', '') if account.get('bank') and account.get('iban') else ''
        _set_cell_value(ws, f'{bank_col}{row}', bank_name)
        _set_cell_value(ws, f'{iban_col}{row}', iban)


def _line_subtotal(line):
    return Decimal(line.qty) * _net_unit_price(line)


def _line_is_meaningful(line):
    try:
        return Decimal(line.qty or 0) > 0
    except Exception:
        return False


def _line_tax(line):
    return _line_subtotal(line) * (Decimal(line.tax or 0) / Decimal('100'))


def _net_unit_price(line):
    net_unit = Decimal(line.unit_price) * (Decimal('1') - (Decimal(line.discount or 0) / Decimal('100')))
    return net_unit * (Decimal('1') - (Decimal(getattr(line, 'discount_secondary', 0) or 0) / Decimal('100')))


def _customer_snapshot(config):
    snapshot = dict((config or {}).get('customer_snapshot') or {})
    return {
        'name': snapshot.get('name', ''),
        'tax_office': snapshot.get('tax_office') or snapshot.get('taxOffice', ''),
        'tax_number': snapshot.get('tax_number') or snapshot.get('taxNumber', ''),
        'address': snapshot.get('address', ''),
        'authorized_person': snapshot.get('authorized_person') or snapshot.get('authorizedPerson', ''),
        'phone': snapshot.get('phone', ''),
        'email': snapshot.get('email', ''),
    }


def _category_section_key(line):
    product = getattr(line, 'product', None)
    category = getattr(product, 'category', None) if product else None
    defaults = getattr(category, 'template_defaults', {}) or {}
    return _canonical_section_key(defaults.get('section_key', ''))


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _join_tax(tax_office, tax_number):
    left = (tax_office or '').strip()
    right = (tax_number or '').strip()
    if left and right:
        return f'{left} / {right}'
    return left or right


def _join_contact(first, second):
    return ' / '.join(part.strip() for part in [first, second] if part and str(part).strip())


def _timezone_fallback(value):
    if isinstance(value, datetime):
        return value.date()
    return date.today()


def _choice_text(label, selected):
    return f'X {label}' if label and selected else label


def format_validity_text(valid_until):
    parsed = _parse_date(valid_until)
    return parsed.strftime('%d.%m.%Y') if parsed else ''


def _format_date_text(value):
    parsed = _parse_date(value)
    return parsed.strftime('%d.%m.%Y') if parsed else str(value or '')


def _seller_logo_media_path(profile):
    raw = str((profile or {}).get('logo_url') or '').strip()
    if not raw:
        return _default_seller_logo_path(profile)
    media_prefix = settings.MEDIA_URL.rstrip('/') + '/'
    if raw.startswith(media_prefix):
        relative = raw[len(media_prefix):]
    elif raw.startswith('/media/'):
        relative = raw[len('/media/'):]
    elif raw.startswith('http://') or raw.startswith('https://'):
        marker = '/media/'
        if marker not in raw:
            return None
        relative = raw.split(marker, 1)[1]
    else:
        relative = raw.lstrip('/')
    candidate = Path(settings.MEDIA_ROOT) / relative
    if not candidate.exists():
        return _default_seller_logo_path(profile)
    if candidate.suffix.lower() not in {'.png', '.jpg', '.jpeg'}:
        return _default_seller_logo_path(profile)
    return candidate


def _strip_header_logos(ws):
    kept = []
    for image in list(getattr(ws, '_images', [])):
        anchor = getattr(image, 'anchor', None)
        marker = getattr(anchor, '_from', None)
        if marker and marker.row <= 12:
            continue
        kept.append(image)
    ws._images = kept


def _apply_seller_logo(ws, profile, seller_side):
    _strip_header_logos(ws)
    logo_path = _seller_logo_media_path(profile)
    if not logo_path:
        return
    image = XLImage(str(logo_path))
    image = _fit_excel_image(image, 220, 76)
    ws.add_image(image, 'K1' if seller_side == 'right' else 'B1')


def _fill_shared_header(ws, quote):
    config = quote.contract_config or {}
    customer = _customer_snapshot(config)
    prepared_by = _resolve_prepared_by_name(quote)
    currency_note = _currency_note(quote)
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    doc_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)

    _set_cell_value(ws, 'L14', doc_date)
    _set_cell_value(ws, 'D23', customer.get('name') or getattr(quote.customer, 'name', ''))
    _set_cell_value(ws, 'D24', _join_tax(customer.get('tax_office'), customer.get('tax_number')))
    _set_cell_value(ws, 'D25', customer.get('address') or '')
    _set_cell_value(ws, 'D26', customer.get('authorized_person') or '')
    _set_cell_value(ws, 'D27', _join_contact(customer.get('phone'), customer.get('email')))
    _set_cell_value(ws, 'D30', quote.number)
    _ensure_header_value_row_layout(ws, 31)
    _set_cell_value(ws, 'D31', prepared_by)
    _ensure_header_value_row_layout(ws, 32, source_row=31)
    _set_cell_value(ws, 'D32', format_validity_text(quote.valid_until))
    _set_cell_value(ws, 'D33', ' • '.join(part for part in [get_quote_price_list_label(quote), currency_note] if part))

    for coordinate in ['C16', 'C17', 'C18', 'C19', 'J16', 'J17', 'J18', 'J19']:
        _set_cell_value(ws, coordinate, '')

    target_prefix = 'J' if seller_side == 'right' else 'C'
    _set_cell_value(ws, f'{target_prefix}16', _normalize_display_name(seller.get('display_name', '')))
    _set_cell_value(ws, f'{target_prefix}17', f"VERGİ DAİRESİ: {seller.get('tax_office', '')}")
    _set_cell_value(ws, f'{target_prefix}18', f"VERGİ NO: {seller.get('tax_number', '')}")
    _set_cell_value(ws, f'{target_prefix}19', _join_contact(seller.get('email', ''), seller.get('phone', '')))
    _apply_seller_logo(ws, seller, seller_side)


def _fill_bank_accounts(ws, quote, template):
    currency_symbol = _currency_symbol(_quote_currency(quote))
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    header_row = template['bank_header_row']

    if not template.get('expanded_layout'):
        _ensure_bank_header_layout(ws, header_row)

    for coordinate in [f'C{header_row}', f'K{header_row}']:
        _set_cell_value(ws, coordinate, '')
    for row in template['bank_rows']:
        if not template.get('expanded_layout'):
            _ensure_bank_row_layout(ws, row)
        for coordinate in [f'C{row}', f'D{row}', f'K{row}', f'M{row}']:
            _set_cell_value(ws, coordinate, '')

    header_col = 'K' if seller_side == 'right' else 'C'
    bank_col = 'K' if seller_side == 'right' else 'C'
    iban_col = 'M' if seller_side == 'right' else 'D'

    _set_cell_value(ws, f'{header_col}{header_row}', f"{_normalize_display_name(seller.get('display_name', ''))} {currency_symbol}")
    bank_accounts = seller.get('bank_accounts') or []
    for index, row in enumerate(template['bank_rows']):
        account = bank_accounts[index] if index < len(bank_accounts) else {}
        bank_name = _normalize_bank_name(account.get('bank', '')) if account.get('bank') and account.get('iban') else ''
        iban = account.get('iban', '') if account.get('bank') and account.get('iban') else ''
        _set_cell_value(ws, f'{bank_col}{row}', bank_name)
        _set_cell_value(ws, f'{iban_col}{row}', iban)


MASTER_COMBINED_CONTRACT_PATH = ASSETS_DIR / 'contract-templates' / 'MASTER' / 'MASTER_COMBINED_CONTRACT.xlsx'
for _template in TEMPLATE_REGISTRY:
    if _template['template_key'] == 'contract_ck_ik_mob_montajli' and MASTER_COMBINED_CONTRACT_PATH.exists():
        _template.update(
            {
                'source_path': MASTER_COMBINED_CONTRACT_PATH,
                'sheet_name': '1',
                'master_contract_layout': True,
                'line_blocks': [
                    {'section_keys': ['steel_door'], 'start_row': 36, 'capacity': 2, 'subtotal_row': 38, 'tax_row': 39, 'grand_row': 40},
                    {'section_keys': ['interior_door'], 'start_row': 47, 'capacity': 3, 'subtotal_row': 50, 'tax_row': 51, 'grand_row': 52},
                    {'section_keys': ['kitchen'], 'start_row': 59, 'capacity': 1, 'subtotal_row': 60, 'tax_row': 61, 'grand_row': 62},
                    {'section_keys': ['wardrobe'], 'start_row': 70, 'capacity': 3, 'subtotal_row': 73, 'tax_row': 74, 'grand_row': 75},
                    {'section_keys': ['bathroom'], 'start_row': 84, 'capacity': 4, 'subtotal_row': 88, 'tax_row': 89, 'grand_row': 90},
                    {'section_keys': ['accessory'], 'start_row': 98, 'capacity': 5, 'subtotal_row': 103, 'tax_row': 104, 'grand_row': 105},
                    {'section_keys': ['laminate'], 'start_row': 109, 'capacity': 1, 'subtotal_row': 110, 'tax_row': 111, 'grand_row': 112},
                    {
                        'section_keys': ['service'],
                        'summary_of_sections': ['steel_door', 'interior_door', 'kitchen', 'wardrobe', 'bathroom', 'accessory', 'laminate'],
                        'start_row': 117,
                        'capacity': 11,
                        'subtotal_row': 128,
                        'tax_row': 129,
                        'grand_row': 130,
                    },
                ],
                'commercial_rows': [],
                'terms_rows': [160, 161, 162, 163, 164, 165, 166, 167, 168, 169],
                'bank_header_row': 177,
                'bank_rows': [178, 179, 180, 181],
                'signature_row': 205,
            }
        )
        break


def _current_export_template(quote):
    template = getattr(quote, '_current_export_template', None)
    return template if isinstance(template, dict) else {}


def _is_master_contract_template(quote, template=None):
    candidate = template or _current_export_template(quote)
    return bool(candidate.get('master_contract_layout'))


def _safe_unmerge(ws, cell_range):
    if cell_range in ws.merged_cells:
        ws.unmerge_cells(cell_range)


def _merge_value_row(ws, cell_range, value, source_cell):
    _safe_unmerge(ws, cell_range)
    ws.merge_cells(cell_range)
    target = ws[source_cell]
    target.value = value
    base_alignment = copy(target.alignment) if target.alignment else Alignment()
    target.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=base_alignment.text_rotation,
        wrap_text=base_alignment.wrap_text,
        shrink_to_fit=base_alignment.shrink_to_fit,
        indent=base_alignment.indent,
        relativeIndent=getattr(base_alignment, 'relativeIndent', 0),
        justifyLastLine=getattr(base_alignment, 'justifyLastLine', None),
        readingOrder=getattr(base_alignment, 'readingOrder', 0),
    )


def _normalize_master_contract_header(ws, seller):
    for coordinate in ['C16', 'C17', 'C18', 'C19', 'J16', 'J17', 'J18', 'J19']:
        _set_cell_value(ws, coordinate, '')
    _merge_value_row(ws, 'C16:L16', _normalize_display_name(seller.get('display_name', '')), 'C16')
    _merge_value_row(ws, 'C17:L17', f"VERGİ DAİRESİ : {seller.get('tax_office', '')}", 'C17')
    _merge_value_row(ws, 'C18:L18', f"VERGİ NO: {seller.get('tax_number', '')}", 'C18')
    _merge_value_row(ws, 'C19:L19', _join_contact(seller.get('email', ''), seller.get('phone', '')), 'C19')


def _normalize_master_contract_bank_block(ws):
    for row in [184, 185, 186, 187, 188]:
        for column in ['C', 'D', 'G', 'H', 'I', 'J', 'K', 'L']:
            _set_cell_value(ws, f'{column}{row}', '')
        ws.row_dimensions[row].hidden = True


def _fill_master_contract_commercial_rows(ws, quote):
    config = quote.contract_config or {}
    delivery_value = (config.get('delivery_type') or _format_date_text(quote.delivery_terms) or '').strip()
    payment_value = (config.get('payment_option') or quote.payment_terms or '').strip()
    for row in [146, 147, 150, 151, 152]:
        for column in ['F', 'H', 'I', 'J', 'K', 'L']:
            _set_cell_value(ws, f'{column}{row}', '')
    _set_cell_value(ws, 'F145', delivery_value)
    _set_cell_value(ws, 'F149', payment_value)


def _fill_master_contract_terms(ws, quote, template):
    terms_rows = list(template.get('terms_rows') or [])
    if len(terms_rows) < 2:
        return
    terms = parse_terms_text((quote.contract_config or {}).get('terms_text'))
    body_rows = terms_rows[:-1]
    closing_row = terms_rows[-1]
    for row in body_rows:
        _set_cell_value(ws, f'C{row}', '')
    for row, term in zip(body_rows, terms, strict=False):
        _set_cell_value(ws, f'C{row}', term)
    contract_date = _parse_date((quote.contract_config or {}).get('contract_date')) or _timezone_fallback(quote.created_at)
    _set_cell_value(ws, f'C{closing_row}', f'İşbu sözleşme {contract_date.strftime("%d.%m.%Y")} tarihinde iki nüsha olarak imzalanmış ve yürürlüğe girmiştir.')


def _fill_master_contract_bank_accounts(ws, quote, template):
    seller = _selected_seller_profile(quote)
    currency_symbol = _currency_symbol(_quote_currency(quote))
    _normalize_master_contract_bank_block(ws)
    _set_cell_value(ws, f"C{template['bank_header_row']}", f"{_normalize_display_name(seller.get('display_name', ''))} {currency_symbol}")
    _set_cell_value(ws, 'G177', _normalize_display_name(seller.get('display_name', '')))
    _set_cell_value(ws, 'G178', _join_tax(seller.get('tax_office', ''), seller.get('tax_number', '')))
    _set_cell_value(ws, 'G179', seller.get('address', ''))
    _set_cell_value(ws, 'G180', seller.get('phone', ''))
    _set_cell_value(ws, 'G181', seller.get('email', ''))
    bank_accounts = seller.get('bank_accounts') or []
    for index, row in enumerate(template['bank_rows']):
        account = bank_accounts[index] if index < len(bank_accounts) else {}
        _set_cell_value(ws, f'C{row}', _normalize_bank_name(account.get('bank', '')) if account.get('bank') and account.get('iban') else '')
        _set_cell_value(ws, f'D{row}', account.get('iban', '') if account.get('bank') and account.get('iban') else '')


def _fill_master_contract_signature(ws, quote, template):
    customer = _customer_snapshot(quote.contract_config or {})
    seller = _selected_seller_profile(quote)
    row = template['signature_row']
    signature_label = str(seller.get('signature_label') or '').strip() or f"{_normalize_display_name(seller.get('display_name', ''))} ADINA İMZA"
    _set_cell_value(ws, f'C{row}', signature_label)
    _set_cell_value(ws, f'G{row}', customer.get('name') or getattr(quote.customer, 'name', '') or '')


def _fill_shared_header(ws, quote):
    config = quote.contract_config or {}
    customer = _customer_snapshot(config)
    prepared_by = _resolve_prepared_by_name(quote)
    currency_note = _currency_note(quote)
    seller = _selected_seller_profile(quote)
    doc_date = _parse_date(config.get('contract_date')) or _timezone_fallback(quote.created_at)

    _set_cell_value(ws, 'L14', doc_date)
    _set_cell_value(ws, 'D22', customer.get('name') or getattr(quote.customer, 'name', ''))
    _set_cell_value(ws, 'D23', _join_tax(customer.get('tax_office'), customer.get('tax_number')))
    _set_cell_value(ws, 'D24', customer.get('address') or '')
    _set_cell_value(ws, 'D25', customer.get('authorized_person') or '')
    _set_cell_value(ws, 'D26', _join_contact(customer.get('phone'), customer.get('email')))
    _set_cell_value(ws, 'D29', quote.number)
    _set_cell_value(ws, 'D30', prepared_by)
    _set_cell_value(ws, 'D31', format_validity_text(quote.valid_until))
    _set_cell_value(ws, 'D32', ' • '.join(part for part in [get_quote_price_list_label(quote), currency_note] if part))

    if _is_master_contract_template(quote):
        _normalize_master_contract_header(ws, seller)
        _apply_seller_logo(ws, seller, 'left')
        return

    for coordinate in ['C16', 'C17', 'C18', 'C19', 'J16', 'J17', 'J18', 'J19']:
        _set_cell_value(ws, coordinate, '')

    target_prefix = 'J' if _selected_seller_side(quote) == 'right' else 'C'
    _set_cell_value(ws, f'{target_prefix}16', _normalize_display_name(seller.get('display_name', '')))
    _set_cell_value(ws, f'{target_prefix}17', f"VERGİ DAİRESİ: {seller.get('tax_office', '')}")
    _set_cell_value(ws, f'{target_prefix}18', f"VERGİ NO: {seller.get('tax_number', '')}")
    _set_cell_value(ws, f'{target_prefix}19', _join_contact(seller.get('email', ''), seller.get('phone', '')))
    _apply_seller_logo(ws, seller, _selected_seller_side(quote))


def _fill_commercial_rows(ws, quote, template):
    if _is_master_contract_template(quote, template):
        _fill_master_contract_commercial_rows(ws, quote)
        return

    config = quote.contract_config or {}
    currency_code = _quote_currency(quote)
    prepared_by = _resolve_prepared_by_name(quote)
    commercial_values = [
        quote.number,
        prepared_by,
        quote.payment_terms or '',
        _format_date_text(quote.delivery_terms),
        format_validity_text(quote.valid_until),
        (config.get('delivery_type') or '').strip(),
        (config.get('payment_option') or '').strip(),
    ]
    for row, value in zip(template['commercial_rows'], commercial_values, strict=False):
        _set_cell_value(ws, f'D{row}', value)
        for column in ['L', 'M']:
            cell = _resolve_cell(ws, f'{column}{row}')
            if cell.value not in [None, '']:
                cell.number_format = _currency_number_format(currency_code)


def _fill_terms(ws, quote, template):
    if _is_master_contract_template(quote, template):
        _fill_master_contract_terms(ws, quote, template)
        return

    terms = parse_terms_text((quote.contract_config or {}).get('terms_text'))
    for row in template['terms_rows']:
        _set_cell_value(ws, f'C{row}', '')
    for row, term in zip(template['terms_rows'], terms, strict=False):
        _set_cell_value(ws, f'C{row}', term)
    if quote.document_type == 'Contract' and template.get('terms_rows'):
        _fill_contract_notes(ws, quote, template)


def _fill_bank_accounts(ws, quote, template):
    if _is_master_contract_template(quote, template):
        _fill_master_contract_bank_accounts(ws, quote, template)
        return

    currency_symbol = _currency_symbol(_quote_currency(quote))
    seller = _selected_seller_profile(quote)
    seller_side = _selected_seller_side(quote)
    header_row = template['bank_header_row']

    if not template.get('expanded_layout'):
        _ensure_bank_header_layout(ws, header_row)

    for coordinate in [f'C{header_row}', f'K{header_row}']:
        _set_cell_value(ws, coordinate, '')
    for row in template['bank_rows']:
        if not template.get('expanded_layout'):
            _ensure_bank_row_layout(ws, row)
        for coordinate in [f'C{row}', f'D{row}', f'K{row}', f'M{row}']:
            _set_cell_value(ws, coordinate, '')

    header_col = 'K' if seller_side == 'right' else 'C'
    bank_col = 'K' if seller_side == 'right' else 'C'
    iban_col = 'M' if seller_side == 'right' else 'D'

    _set_cell_value(ws, f'{header_col}{header_row}', f"{_normalize_display_name(seller.get('display_name', ''))} {currency_symbol}")
    bank_accounts = seller.get('bank_accounts') or []
    for index, row in enumerate(template['bank_rows']):
        account = bank_accounts[index] if index < len(bank_accounts) else {}
        bank_name = _normalize_bank_name(account.get('bank', '')) if account.get('bank') and account.get('iban') else ''
        iban = account.get('iban', '') if account.get('bank') and account.get('iban') else ''
        _set_cell_value(ws, f'{bank_col}{row}', bank_name)
        _set_cell_value(ws, f'{iban_col}{row}', iban)


def _fill_signature_block(ws, quote, template):
    if _is_master_contract_template(quote, template):
        _fill_master_contract_signature(ws, quote, template)
        return

    customer = _customer_snapshot(quote.contract_config or {})
    profiles = {profile['key'].upper(): profile for profile in get_seller_profiles(quote.organization)}
    fallback_seller = _seller_profiles_for_template(quote)[1]
    seller = profiles.get((quote.seller_company_key or fallback_seller.get('key') or 'AYKA').strip().upper(), fallback_seller)
    row = template['signature_row']
    signature_label = str(seller.get('signature_label') or '').strip() or f"{_normalize_display_name(seller.get('display_name', ''))} ADINA İMZA"
    _set_cell_value(ws, f'C{row}', signature_label)
    _set_cell_value(ws, f'J{row}', customer.get('name') or getattr(quote.customer, 'name', '') or '')
